"""
build_player_profiles.py
────────────────────────
Generates four Excel files using The Athletic's 18 player roles:

  Player Profiles/Attackers.xlsx
  Player Profiles/Midfielders.xlsx
  Player Profiles/Defenders.xlsx
  Player Profiles/Goalkeepers.xlsx

Each file contains:
  • One sheet per role (players sorted by role fit score)
  • Group-level Z-score anomaly sheet  (vs all players in the position group)
  • Role-level  Z-score anomaly sheet  (vs peers in the same role)
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import norm
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Output directory ──────────────────────────────────────────────────────────
OUT_DIR = Path("Player Profiles")
OUT_DIR.mkdir(exist_ok=True)

WYSCOUT_DIR = Path("data/Wyscout DB")
MIN_MINUTES  = 400
Z_GROUP_THRESH = 1.80   # anomaly threshold at position-group level
Z_ROLE_THRESH  = 1.60   # anomaly threshold at role level (smaller pools)

# ══════════════════════════════════════════════════════════════════════════════
# 1.  ROLE DEFINITIONS (The Athletic's 18 roles)
# ══════════════════════════════════════════════════════════════════════════════

ROLE_METRICS: dict[str, dict[str, float]] = {
    # ── Attackers ─────────────────────────────────────────────────
    "Finisher": {
        "Goals per 90": 3.0, "Non-penalty goals per 90": 2.5,
        "xG per 90": 2.5, "Goal conversion, %": 2.0,
        "Shots per 90": 1.5, "Shots on target, %": 1.5,
        "Touches in box per 90": 1.0,
    },
    "Target": {
        "Head goals per 90": 3.0, "Aerial duels won, %": 2.5,
        "Aerial duels per 90": 2.0, "Touches in box per 90": 1.5,
        "Fouls suffered per 90": 1.5, "xG per 90": 1.0,
    },
    "Roamer": {
        "Progressive runs per 90": 2.5, "Successful dribbles, %": 2.0,
        "Dribbles per 90": 2.0, "Accelerations per 90": 1.5,
        "Goals per 90": 1.5, "Offensive duels won, %": 1.5,
        "xG per 90": 1.0,
    },
    "Wide threat": {
        "Crosses per 90": 2.5, "Accurate crosses, %": 2.0,
        "Deep completed crosses per 90": 2.0, "xA per 90": 1.5,
        "Assists per 90": 1.5, "Shot assists per 90": 1.0,
    },
    "Unlocker": {
        "Key passes per 90": 3.0, "Smart passes per 90": 2.5,
        "xA per 90": 2.5, "Through passes per 90": 2.0,
        "Passes to penalty area per 90": 2.0, "Shot assists per 90": 1.5,
    },
    "Outlet": {
        "Received passes per 90": 2.5, "Passes per 90": 2.0,
        "Accurate passes, %": 2.0, "Fouls suffered per 90": 1.5,
        "Progressive passes per 90": 1.5,
        "Successful attacking actions per 90": 1.0,
    },
    # ── Midfielders ────────────────────────────────────────────────
    "Box crasher": {
        "Touches in box per 90": 3.0, "Goals per 90": 2.5,
        "xG per 90": 2.5, "Progressive runs per 90": 2.0,
        "Successful attacking actions per 90": 1.5, "Shots per 90": 1.5,
    },
    "Creator": {
        "Key passes per 90": 3.0, "xA per 90": 3.0,
        "Smart passes per 90": 2.0, "Shot assists per 90": 2.0,
        "Passes to penalty area per 90": 1.5, "Deep completions per 90": 1.5,
        "Second assists per 90": 1.0,
    },
    "Orchestrator": {
        "Passes per 90": 2.5, "Accurate passes, %": 2.0,
        "Progressive passes per 90": 2.5, "Passes to final third per 90": 2.0,
        "Received passes per 90": 1.5, "Accurate progressive passes, %": 1.5,
    },
    "Box to box": {
        "Defensive duels per 90": 2.0, "Defensive duels won, %": 2.0,
        "Successful defensive actions per 90": 2.0, "Goals per 90": 1.5,
        "Forward passes per 90": 1.5, "Duels won, %": 1.5,
        "Progressive passes per 90": 1.0,
    },
    "Distributor": {
        "Long passes per 90": 3.0, "Accurate long passes, %": 2.5,
        "Passes per 90": 2.0, "Accurate passes, %": 2.0,
        "Average pass length, m": 2.0, "Received passes per 90": 1.0,
    },
    # ── Defenders ──────────────────────────────────────────────────
    "Anchor": {
        "Interceptions per 90": 3.0, "PAdj Interceptions": 2.5,
        "Defensive duels won, %": 2.0, "Successful defensive actions per 90": 2.0,
        "Defensive duels per 90": 1.5, "Duels won, %": 1.5,
    },
    "Spreader": {
        "Accurate passes, %": 2.5, "Progressive passes per 90": 2.5,
        "Accurate progressive passes, %": 2.0, "Passes to final third per 90": 2.0,
        "Long passes per 90": 1.5, "Accurate long passes, %": 1.5,
    },
    "Aggressor": {
        "Defensive duels per 90": 2.5, "Sliding tackles per 90": 2.5,
        "PAdj Sliding tackles": 2.0, "Duels per 90": 1.5,
        "Aerial duels per 90": 1.5, "Fouls per 90": 1.0,
    },
    "Safety": {
        "Successful defensive actions per 90": 2.5,
        "Defensive duels won, %": 2.5, "Aerial duels won, %": 2.0,
        "Interceptions per 90": 2.0, "Shots blocked per 90": 1.5,
        "Duels won, %": 1.5,
    },
    "Progressor": {
        "Progressive passes per 90": 2.5, "Passes to final third per 90": 2.5,
        "Forward passes per 90": 2.0, "Key passes per 90": 1.5,
        "xA per 90": 1.5, "Accurate passes, %": 1.0,
    },
    "Overlapper": {
        "Crosses per 90": 2.5, "Progressive runs per 90": 2.5,
        "Assists per 90": 2.0, "Offensive duels per 90": 2.0,
        "Deep completed crosses per 90": 1.5, "Accelerations per 90": 1.5,
    },
    "Builder": {
        "Accurate passes, %": 2.5, "Passes per 90": 2.0,
        "Accurate long passes, %": 2.0, "Average pass length, m": 2.0,
        "Received long passes per 90": 1.5, "Back passes per 90": 1.0,
    },
}

GK_METRICS: dict[str, float] = {
    "Save rate, %": 3.0, "Prevented goals per 90": 2.5,
    "Exits per 90": 2.0, "xG against per 90": -2.0,
    "Aerial duels per 90": 1.5, "Back passes received as GK per 90": 1.0,
    "Accurate passes, %": 2.0, "Accurate long passes, %": 1.5,
}

# Position code → PositionFamily
POSITION_FAMILY: dict[str, str] = {
    "CF": "Central attacker", "SS": "Central attacker",
    "LW": "Wide attacker", "RW": "Wide attacker",
    "LWF": "Wide attacker", "RWF": "Wide attacker", "WF": "Wide attacker",
    "AMF": "Advanced midfielder", "LAMF": "Advanced midfielder", "RAMF": "Advanced midfielder",
    "CMF": "Deep midfielder",
    "LCM": "Deep midfielder", "RCM": "Deep midfielder",
    "LCMF": "Deep midfielder", "RCMF": "Deep midfielder",
    "DMF": "Deep midfielder",
    "LDM": "Deep midfielder", "RDM": "Deep midfielder",
    "LDMF": "Deep midfielder", "RDMF": "Deep midfielder",
    "LB": "Wide defender", "RB": "Wide defender",
    "LWB": "Wide defender", "RWB": "Wide defender",
    "CB": "Central defender",
    "LCB": "Central defender", "RCB": "Central defender",
    "GK": "Goalkeeper",
}

FAMILY_TO_ROLES: dict[str, list[str]] = {
    "Central attacker":   ["Finisher", "Target"],
    "Wide attacker":      ["Roamer", "Wide threat", "Unlocker", "Outlet"],
    "Advanced midfielder":["Box crasher", "Creator", "Orchestrator"],
    "Deep midfielder":    ["Box to box", "Distributor"],
    "Central defender":   ["Anchor", "Spreader", "Aggressor"],
    "Wide defender":      ["Safety", "Progressor", "Overlapper", "Builder"],
}

FAMILY_TO_GROUP: dict[str, str] = {
    "Central attacker": "Attackers", "Wide attacker": "Attackers",
    "Advanced midfielder": "Midfielders", "Deep midfielder": "Midfielders",
    "Central defender": "Defenders", "Wide defender": "Defenders",
    "Goalkeeper": "Goalkeepers",
}

# Key metrics for group-level anomaly detection
GROUP_KEY_METRICS: dict[str, list[str]] = {
    "Attackers": [
        "Goals per 90", "xG per 90", "Assists per 90", "xA per 90",
        "Shots per 90", "Touches in box per 90", "Key passes per 90",
        "Progressive runs per 90", "Dribbles per 90", "Successful dribbles, %",
        "Aerial duels won, %", "Crosses per 90", "Smart passes per 90",
    ],
    "Midfielders": [
        "Passes per 90", "Accurate passes, %", "Progressive passes per 90",
        "Defensive duels per 90", "Interceptions per 90", "Key passes per 90",
        "xA per 90", "Goals per 90", "Duels won, %", "Touches in box per 90",
        "Long passes per 90", "Successful defensive actions per 90",
    ],
    "Defenders": [
        "Successful defensive actions per 90", "Defensive duels per 90",
        "Defensive duels won, %", "Interceptions per 90",
        "Aerial duels per 90", "Aerial duels won, %",
        "Accurate passes, %", "Progressive passes per 90",
        "Sliding tackles per 90", "Duels won, %",
    ],
    "Goalkeepers": [
        "Save rate, %", "Prevented goals per 90", "Exits per 90",
        "xG against per 90", "Aerial duels per 90",
        "Back passes received as GK per 90",
        "Accurate passes, %", "Accurate long passes, %",
    ],
}

BIO_COLS = ["Player", "Team", "League", "Position", "PositionFamily",
            "Age", "Height", "Foot", "Minutes played", "Matches played",
            "Market value", "Contract expires",
            "Birth country", "Passport country"]

# ══════════════════════════════════════════════════════════════════════════════
# 2.  STYLING CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

THEME: dict[str, dict] = {
    "Attackers":   {"header": "9C1700", "light": "FFF0ED", "accent": "E53935"},
    "Midfielders": {"header": "0D3B6E", "light": "EEF4FF", "accent": "1565C0"},
    "Defenders":   {"header": "1B5E20", "light": "EDFAEE", "accent": "2E7D32"},
    "Goalkeepers": {"header": "5D4037", "light": "FFF8E1", "accent": "F57F17"},
}
WHITE = "FFFFFF"
ANOMALY_RED  = "FF4B4B"
ANOMALY_AMBER = "FFA726"


def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header(ws, row: int, ncols: int, bg: str, font_size: int = 9):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=WHITE, size=font_size)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin()


def _row_style(ws, r: int, ncols: int, bg: str, bold_cols: set[int] | None = None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center" if c > 4 else "left",
                                   vertical="center")
        cell.font = Font(bold=(bold_cols and c in bold_cols), size=9)
        cell.border = _thin()


def _col_widths(ws, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze(ws, cell: str = "A2"):
    ws.freeze_panes = cell

# ══════════════════════════════════════════════════════════════════════════════
# 3.  DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════

NON_METRIC = {
    "Player", "Team", "Team within selected timeframe", "Position",
    "Age", "Market value", "Contract expires", "Birth country",
    "Passport country", "Foot", "Height", "Weight", "On loan",
}


def _load_file(path: Path) -> pd.DataFrame:
    try:
        df = pd.read_excel(path)
        df.columns = [str(c).strip() for c in df.columns]
        # First position only
        if "Position" in df.columns:
            df["Position"] = (
                df["Position"].astype(str).str.split(",").str[0].str.strip()
            )
        df["League"] = path.stem
        return df
    except Exception:
        return pd.DataFrame()


def load_all_wyscout() -> pd.DataFrame:
    files = sorted(WYSCOUT_DIR.glob("*.xlsx"))
    print(f"  Loading {len(files)} Wyscout files …")
    frames = [_load_file(p) for p in files]
    frames = [f for f in frames if not f.empty]
    df = pd.concat(frames, ignore_index=True)

    # Numeric coercion
    for col in df.columns:
        if col not in NON_METRIC:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Minutes filter
    if "Minutes played" in df.columns:
        df = df.loc[df["Minutes played"].fillna(0) >= MIN_MINUTES].copy()

    # Map position → family → group
    df["PositionFamily"] = df["Position"].map(POSITION_FAMILY).fillna("Other")
    df["PositionGroup"]  = df["PositionFamily"].map(FAMILY_TO_GROUP).fillna("Other")

    df = df.loc[df["PositionGroup"] != "Other"].reset_index(drop=True)
    print(f"  → {len(df):,} players after {MIN_MINUTES}-min filter")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 4.  ROLE SCORE COMPUTATION
# ══════════════════════════════════════════════════════════════════════════════

def _z_percentile(z: np.ndarray) -> np.ndarray:
    return norm.cdf(z) * 100


def _weighted_z_score(
    grp: pd.DataFrame, blueprint: dict[str, float]
) -> pd.Series:
    """Weighted z-score composite (returns raw z, not percentile)."""
    available = [(m, w) for m, w in blueprint.items() if m in grp.columns]
    if not available:
        return pd.Series(0.0, index=grp.index)
    total_w = sum(w for _, w in available)
    z_comp = pd.Series(0.0, index=grp.index)
    for metric, weight in available:
        col = grp[metric].fillna(0)
        mu  = col.mean()
        sig = col.std() or 1e-9
        z_comp += (weight / total_w) * (col - mu) / sig
    return z_comp


def compute_role_scores(df: pd.DataFrame) -> pd.DataFrame:
    """Add RoleScore_<role> columns and PrimaryRole / PrimaryRoleScore."""
    df = df.copy()
    role_score_cols: list[str] = []

    for family, roles in FAMILY_TO_ROLES.items():
        mask = df["PositionFamily"] == family
        if mask.sum() < 5:
            continue
        grp = df.loc[mask]

        scores: dict[str, pd.Series] = {}
        for role in roles:
            z = _weighted_z_score(grp, ROLE_METRICS[role])
            pct = pd.Series(_z_percentile(z.values), index=grp.index)
            col_name = f"RoleScore_{role}"
            df.loc[mask, col_name] = pct.values
            scores[role] = pct
            if col_name not in role_score_cols:
                role_score_cols.append(col_name)

        role_df = pd.DataFrame(scores)
        df.loc[mask, "PrimaryRole"]      = role_df.idxmax(axis=1).values
        df.loc[mask, "PrimaryRoleScore"] = role_df.max(axis=1).round(1).values

    return df


# ══════════════════════════════════════════════════════════════════════════════
# 5.  Z-SCORE ANOMALY ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def compute_anomalies(
    df: pd.DataFrame, metrics: list[str], threshold: float, label: str
) -> pd.DataFrame:
    """
    Compute per-player z-scores, peak_z, anomaly_score, anomaly_type.
    Returns only anomaly rows (peak_z >= threshold).
    """
    metrics = [m for m in metrics if m in df.columns]
    if not metrics or len(df) < 5:
        return pd.DataFrame()

    result = df.copy()
    X = result[metrics].apply(pd.to_numeric, errors="coerce").fillna(0).values.astype(float)
    mu  = X.mean(axis=0)
    sig = X.std(axis=0)
    sig = np.where(sig == 0, 1e-9, sig)
    Z   = (X - mu) / sig

    for i, m in enumerate(metrics):
        result[f"_z_{m}"] = Z[:, i]

    z_cols = [f"_z_{m}" for m in metrics]
    result["_peak_z"]          = result[z_cols].max(axis=1)
    result["_mean_z"]          = result[z_cols].mean(axis=1)
    result["_anomaly_breadth"] = (result[z_cols] >= threshold).sum(axis=1)
    result["_anomaly_score"]   = (
        0.45 * result["_peak_z"].clip(lower=0)
        + 0.35 * result["_anomaly_breadth"]
        + 0.20 * result["_mean_z"].clip(lower=0)
    )

    def _classify(row) -> str:
        peak    = float(row["_peak_z"])
        breadth = int(row["_anomaly_breadth"])
        age     = float(pd.to_numeric(row.get("Age"), errors="coerce") or 99)
        if breadth >= 5:
            return "Multi-dimensional"
        if peak >= threshold * 1.6 and breadth <= 2:
            return "Specialist Elite"
        if peak >= threshold and age <= 22:
            return "Age-adjusted Gem"
        if breadth >= 3:
            return "Consistent Overperformer"
        return "Emerging Talent"

    result["_anomaly_type"] = result.apply(_classify, axis=1)
    result["_context"] = label

    anomalies = result.loc[result["_peak_z"] >= threshold].copy()
    return anomalies.sort_values("_anomaly_score", ascending=False)


# ══════════════════════════════════════════════════════════════════════════════
# 6.  EXCEL WRITING
# ══════════════════════════════════════════════════════════════════════════════

DISPLAY_COLS_ATK = [
    "Player", "Team", "League", "Position", "PositionFamily", "Age",
    "Minutes played", "PrimaryRole", "PrimaryRoleScore",
    "Goals per 90", "xG per 90", "Assists per 90", "xA per 90",
    "Shots per 90", "Touches in box per 90",
    "Key passes per 90", "Progressive runs per 90",
    "Dribbles per 90", "Successful dribbles, %",
    "Aerial duels won, %", "Crosses per 90", "Accurate crosses, %",
]

DISPLAY_COLS_MID = [
    "Player", "Team", "League", "Position", "PositionFamily", "Age",
    "Minutes played", "PrimaryRole", "PrimaryRoleScore",
    "Passes per 90", "Accurate passes, %", "Progressive passes per 90",
    "Passes to final third per 90", "Key passes per 90",
    "xA per 90", "Goals per 90",
    "Defensive duels per 90", "Defensive duels won, %",
    "Interceptions per 90", "Long passes per 90", "Accurate long passes, %",
]

DISPLAY_COLS_DEF = [
    "Player", "Team", "League", "Position", "PositionFamily", "Age",
    "Minutes played", "PrimaryRole", "PrimaryRoleScore",
    "Successful defensive actions per 90", "Defensive duels per 90",
    "Defensive duels won, %", "Interceptions per 90",
    "Aerial duels per 90", "Aerial duels won, %",
    "Accurate passes, %", "Progressive passes per 90",
    "Passes to final third per 90", "Crosses per 90",
    "Sliding tackles per 90",
]

DISPLAY_COLS_GK = [
    "Player", "Team", "League", "Position", "Age", "Height",
    "Minutes played", "Matches played",
    "Save rate, %", "Prevented goals per 90", "Exits per 90",
    "xG against per 90", "Aerial duels per 90",
    "Back passes received as GK per 90",
    "Accurate passes, %", "Accurate long passes, %",
    "Conceded goals per 90",
]

ANOMALY_DISPLAY = [
    "Player", "Team", "League", "Position", "PositionFamily", "Age",
    "Minutes played", "PrimaryRole",
    "_anomaly_type", "_peak_z", "_anomaly_breadth", "_anomaly_score", "_context",
]
ANOMALY_DISPLAY_GK = [
    "Player", "Team", "League", "Position", "Age", "Minutes played",
    "_anomaly_type", "_peak_z", "_anomaly_breadth", "_anomaly_score", "_context",
]


def _available(df: pd.DataFrame, cols: list[str]) -> list[str]:
    return [c for c in cols if c in df.columns]


def _prep_display(df: pd.DataFrame, cols: list[str], round_dp: int = 2) -> pd.DataFrame:
    cols = _available(df, cols)
    out  = df[cols].copy()
    num  = [c for c in cols if out[c].dtype.kind in "f"]
    out[num] = out[num].round(round_dp)
    return out.reset_index(drop=True)


def _write_main_sheet(
    ws, df: pd.DataFrame, theme_key: str, role: str | None = None
):
    """Write a formatted player table onto an existing worksheet."""
    ncols = len(df.columns)
    ws.row_dimensions[1].height = 30
    _header(ws, 1, ncols, THEME[theme_key]["header"])

    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        bg = THEME[theme_key]["light"] if r_idx % 2 == 0 else WHITE
        for c_idx, val in enumerate(row.values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="left" if c_idx <= 3 else "center",
                vertical="center",
            )
            cell.font = Font(size=9, bold=(c_idx == 1))
            cell.border = _thin()

    _freeze(ws, "A2")


def _write_anomaly_sheet(ws, df: pd.DataFrame, theme_key: str):
    """Anomaly table with colour-coded peak z column."""
    ncols = len(df.columns)
    ws.row_dimensions[1].height = 30
    _header(ws, 1, ncols, "1A1A2E")

    peak_col_idx = (
        df.columns.tolist().index("_peak_z") + 1
        if "_peak_z" in df.columns else None
    )
    type_col_idx = (
        df.columns.tolist().index("_anomaly_type") + 1
        if "_anomaly_type" in df.columns else None
    )

    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        bg = "F8F8FF" if r_idx % 2 == 0 else WHITE
        for c_idx, val in enumerate(row.values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="left" if c_idx <= 3 else "center",
                vertical="center",
            )
            cell.font = Font(size=9, bold=(c_idx == 1))
            cell.border = _thin()

        # Highlight peak_z
        if peak_col_idx:
            pz = row.get("_peak_z", 0)
            try:
                pz = float(pz)
            except Exception:
                pz = 0.0
            colour = ANOMALY_RED if pz >= 2.5 else (ANOMALY_AMBER if pz >= 2.0 else "4CAF50")
            cell = ws.cell(row=r_idx, column=peak_col_idx)
            cell.font = Font(size=9, bold=True, color=colour)

    _freeze(ws, "A2")


def _set_widths_auto(ws, df: pd.DataFrame, base: float = 12.0):
    for i, col in enumerate(df.columns, 1):
        width = max(len(str(col)) * 1.1, base)
        if col in ("Player", "Team", "League"):
            width = max(width, 22.0)
        elif col.startswith("_") or col in ("PrimaryRole", "PositionFamily"):
            width = max(width, 18.0)
        ws.column_dimensions[get_column_letter(i)].width = min(width, 30.0)


# ══════════════════════════════════════════════════════════════════════════════
# 7.  BUILD EACH FILE
# ══════════════════════════════════════════════════════════════════════════════

def _build_group_file(
    df_group: pd.DataFrame,
    group: str,
    roles: list[str],
    display_cols: list[str],
    filename: str,
):
    out_path = OUT_DIR / filename
    print(f"  Building {filename}  ({len(df_group):,} players) …")

    # ── Group-level anomalies ─────────────────────────────────────
    group_anomalies = compute_anomalies(
        df_group,
        GROUP_KEY_METRICS[group],
        Z_GROUP_THRESH,
        label=f"{group} group",
    )

    # ── Role-level anomalies ──────────────────────────────────────
    role_anom_frames: list[pd.DataFrame] = []
    for role in roles:
        mask = df_group["PrimaryRole"] == role
        sub  = df_group.loc[mask]
        if len(sub) < 5:
            continue
        ra = compute_anomalies(
            sub,
            list(ROLE_METRICS[role].keys()),
            Z_ROLE_THRESH,
            label=role,
        )
        if not ra.empty:
            role_anom_frames.append(ra)
    role_anomalies = (
        pd.concat(role_anom_frames, ignore_index=True)
        if role_anom_frames
        else pd.DataFrame()
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # ── Overview ──────────────────────────────────────────────
        overview = _prep_display(
            df_group.sort_values("PrimaryRoleScore", ascending=False),
            display_cols,
        )
        overview.to_excel(writer, sheet_name=f"All {group}", index=False)
        ws = writer.sheets[f"All {group}"]
        _write_main_sheet(ws, overview, group)
        _set_widths_auto(ws, overview)

        # ── One sheet per role ────────────────────────────────────
        for role in roles:
            mask   = df_group["PrimaryRole"] == role
            sub    = df_group.loc[mask]
            if sub.empty:
                continue

            role_score_col = f"RoleScore_{role}"
            sort_col = role_score_col if role_score_col in sub.columns else "PrimaryRoleScore"

            # Build role-specific display: bio + role metrics
            role_metric_cols = list(ROLE_METRICS[role].keys())
            role_display = _available(
                sub,
                ["Player", "Team", "League", "Position", "Age",
                 "Minutes played", "PrimaryRoleScore"] + role_metric_cols
            )

            sheet_df = _prep_display(sub.sort_values(sort_col, ascending=False), role_display)
            sname = role[:31]
            sheet_df.to_excel(writer, sheet_name=sname, index=False)
            ws = writer.sheets[sname]
            _write_main_sheet(ws, sheet_df, group, role=role)
            _set_widths_auto(ws, sheet_df)

        # ── Group Anomalies ───────────────────────────────────────
        if not group_anomalies.empty:
            ga_disp = _prep_display(
                group_anomalies, ANOMALY_DISPLAY if group != "Goalkeepers" else ANOMALY_DISPLAY_GK
            )
            ga_disp.to_excel(writer, sheet_name="Group Anomalies", index=False)
            ws = writer.sheets["Group Anomalies"]
            _write_anomaly_sheet(ws, ga_disp, group)
            _set_widths_auto(ws, ga_disp)
        else:
            pd.DataFrame(columns=["No anomalies found"]).to_excel(
                writer, sheet_name="Group Anomalies", index=False
            )

        # ── Role Anomalies ────────────────────────────────────────
        if not role_anomalies.empty:
            ra_disp = _prep_display(
                role_anomalies, ANOMALY_DISPLAY if group != "Goalkeepers" else ANOMALY_DISPLAY_GK
            )
            ra_disp.to_excel(writer, sheet_name="Role Anomalies", index=False)
            ws = writer.sheets["Role Anomalies"]
            _write_anomaly_sheet(ws, ra_disp, group)
            _set_widths_auto(ws, ra_disp)
        else:
            pd.DataFrame(columns=["No anomalies found"]).to_excel(
                writer, sheet_name="Role Anomalies", index=False
            )

    print(f"    Saved → {out_path}")


def build_gk_file(df_gk: pd.DataFrame):
    out_path = OUT_DIR / "Goalkeepers.xlsx"
    print(f"  Building Goalkeepers.xlsx  ({len(df_gk):,} players) …")

    # GK composite score
    z = _weighted_z_score(df_gk, GK_METRICS)
    df_gk = df_gk.copy()
    df_gk["GKScore"] = _z_percentile(z.values).round(1)

    # GK style: Shot stopper / Sweeper / Ball-playing
    gk_styles = {
        "Shot Stopper":    {"Save rate, %": 3.0, "Prevented goals per 90": 2.5},
        "Sweeper GK":      {"Exits per 90": 3.0, "Aerial duels per 90": 2.5},
        "Ball-playing GK": {"Accurate passes, %": 3.0, "Accurate long passes, %": 2.5},
        "Command":         {"Aerial duels per 90": 2.0, "Shots against per 90": 2.0},
    }
    style_scores: dict[str, pd.Series] = {}
    for sname, weights in gk_styles.items():
        z_s = _weighted_z_score(df_gk, weights)
        style_scores[sname] = pd.Series(_z_percentile(z_s.values), index=df_gk.index)
    style_df = pd.DataFrame(style_scores)
    df_gk["GKStyle"]      = style_df.idxmax(axis=1).values
    df_gk["GKStyleScore"] = style_df.max(axis=1).round(1).values

    # Anomalies
    group_anom = compute_anomalies(
        df_gk, GROUP_KEY_METRICS["Goalkeepers"], Z_GROUP_THRESH, label="Goalkeepers"
    )
    role_anom_frames = []
    for style in gk_styles:
        mask = df_gk["GKStyle"] == style
        sub  = df_gk.loc[mask]
        if len(sub) < 5:
            continue
        ra = compute_anomalies(
            sub, list(gk_styles[style].keys()), Z_ROLE_THRESH, label=style
        )
        if not ra.empty:
            role_anom_frames.append(ra)
    role_anom = pd.concat(role_anom_frames) if role_anom_frames else pd.DataFrame()

    gk_display_cols = _available(df_gk, DISPLAY_COLS_GK + ["GKScore", "GKStyle", "GKStyleScore"])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # All GKs
        overview = _prep_display(df_gk.sort_values("GKScore", ascending=False), gk_display_cols)
        overview.to_excel(writer, sheet_name="All Goalkeepers", index=False)
        ws = writer.sheets["All Goalkeepers"]
        _write_main_sheet(ws, overview, "Goalkeepers")
        _set_widths_auto(ws, overview)

        # Per-style sheets
        for style in gk_styles:
            mask   = df_gk["GKStyle"] == style
            sub    = df_gk.loc[mask]
            if sub.empty:
                continue
            sub_disp = _prep_display(
                sub.sort_values("GKStyleScore", ascending=False),
                gk_display_cols,
            )
            sname = style[:31]
            sub_disp.to_excel(writer, sheet_name=sname, index=False)
            ws = writer.sheets[sname]
            _write_main_sheet(ws, sub_disp, "Goalkeepers")
            _set_widths_auto(ws, sub_disp)

        # Group Anomalies
        if not group_anom.empty:
            ga = _prep_display(group_anom, ANOMALY_DISPLAY_GK)
            ga.to_excel(writer, sheet_name="Group Anomalies", index=False)
            ws = writer.sheets["Group Anomalies"]
            _write_anomaly_sheet(ws, ga, "Goalkeepers")
            _set_widths_auto(ws, ga)

        # Role Anomalies
        if not role_anom.empty:
            ra = _prep_display(role_anom, ANOMALY_DISPLAY_GK)
            ra.to_excel(writer, sheet_name="Role Anomalies", index=False)
            ws = writer.sheets["Role Anomalies"]
            _write_anomaly_sheet(ws, ra, "Goalkeepers")
            _set_widths_auto(ws, ra)

    print(f"    Saved → {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# 8.  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("Loading Wyscout data …")
    df = load_all_wyscout()

    print("Computing role scores …")
    df = compute_role_scores(df)

    # ── Split by group ────────────────────────────────────────────
    df_atk = df.loc[df["PositionGroup"] == "Attackers"].copy()
    df_mid = df.loc[df["PositionGroup"] == "Midfielders"].copy()
    df_def = df.loc[df["PositionGroup"] == "Defenders"].copy()
    df_gk  = df.loc[df["PositionGroup"] == "Goalkeepers"].copy()

    print(f"\n  Attackers:   {len(df_atk):,}")
    print(f"  Midfielders: {len(df_mid):,}")
    print(f"  Defenders:   {len(df_def):,}")
    print(f"  Goalkeepers: {len(df_gk):,}\n")

    print("Writing Excel files …")

    _build_group_file(
        df_atk, "Attackers",
        roles=["Finisher", "Target", "Roamer", "Wide threat", "Unlocker", "Outlet"],
        display_cols=DISPLAY_COLS_ATK,
        filename="Attackers.xlsx",
    )
    _build_group_file(
        df_mid, "Midfielders",
        roles=["Box crasher", "Creator", "Orchestrator", "Box to box", "Distributor"],
        display_cols=DISPLAY_COLS_MID,
        filename="Midfielders.xlsx",
    )
    _build_group_file(
        df_def, "Defenders",
        roles=["Anchor", "Spreader", "Aggressor", "Safety", "Progressor", "Overlapper", "Builder"],
        display_cols=DISPLAY_COLS_DEF,
        filename="Defenders.xlsx",
    )
    build_gk_file(df_gk)

    print(f"\nDone. All files saved to:  {OUT_DIR.resolve()}")


if __name__ == "__main__":
    main()
