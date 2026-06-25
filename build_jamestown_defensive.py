"""
build_jamestown_defensive.py  (xlsxwriter-accelerated)
────────────────────────────
JamesTown Analytics — Defensive Player Intelligence Report
Covers: GK · CB · FB · DM

New composite defensive indexes computed here:

  DIS  – Defensive Impact Score
         Weighted percentile rank of core defensive output:
         Successful def actions/90 (3×) · Def duels won % (2.5×) ·
         Interceptions/90 (2×) · PAdj Interceptions (2×) ·
         Aerial duels won % (1.5×) · Shots blocked/90 (1×)

  ADI  – Aerial Dominance Index
         Aerial duels/90 × aerial duels won % / 100 → rank (0–100)
         Captures both volume and quality of aerial play.

  PADS – Pressure-Adjusted Defensive Score
         PAdj Interceptions + PAdj Sliding tackles + Shots blocked/90
         Removes possession-context bias; true defensive contribution.

  PDS  – Positional Discipline Score
         Inverted percentile of (Fouls/90 × 0.6 + Yellow cards/90 × 0.3
         + Red cards/90 × 0.1) — high PDS = disciplined, low = sloppy.

  JDR  – JamesTown Defensive Rating (0–100)
         Master defensive rank: position-specific weighted composite
         of DIS (40%) · PADS (25%) · ADI (20%) · PDS (15%)
         Analogous to SQS in the Lamberts model.

  DVI  – Defensive Value Index  (= JDR Rank − Market Value Rank)
         Positive → player performs above what the market charges.
         Identical logic to the Lamberts Index but tuned for defenders.

Usage:
  python build_jamestown_defensive.py
  python build_jamestown_defensive.py --leagues Czech Slovakia --min-minutes 500
  python build_jamestown_defensive.py --output data/JamesTown_Defensive.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config ─────────────────────────────────────────────────────────────────────

ROOT = Path(__file__).parent
WYSCOUT_DIR = ROOT / "Wyscout Files"

SKIP_FILES = {
    "FCHK Model V3 - Loaded Leagues", "FCHK Model V3 - Model Input",
    "FCHK Model V3 - Player Scores", "FCHK Model V3 - Player Styles",
    "FCHK Model V3 - Recruitment Scores", "FCHK Model V3 - Smart Club Closeness",
    "FCHK Model V3 - Summary", "FCHK Model V3 Scores", "FCHK Scouting Report",
    "Leagues Overview", "Wyscout Anomaly Report", "Wyscout Full Scouting Report",
}

DEFAULT_MIN_MINUTES = 400
DEFAULT_MAX_AGE = 30
DEFAULT_BUDGET = 1_000_000

# Defensive positions only
DEFENSIVE_POSITIONS = {"GK", "CB", "FB", "DM"}

POS_MAP: dict[str, str] = {
    "GK": "GK",
    "CB": "CB", "LCB": "CB", "RCB": "CB",
    "LB": "FB", "RB": "FB", "LWB": "FB", "RWB": "FB",
    "DMF": "DM", "LDM": "DM", "RDM": "DM", "LDMF": "DM", "RDMF": "DM",
}

# Hradec defensive squad benchmarks (Czech top-flight percentile)
HRADEC_DEFENSE: dict[str, float] = {
    "GK": 61.5,
    "CB": 33.3,
    "FB": 81.8,
    "DM": 36.8,
}

# ── Colors ─────────────────────────────────────────────────────────────────────

C = {
    "navy":    "0D1B2A",
    "gold":    "C9A84C",
    "teal":    "0B6E4F",       # JamesTown primary accent
    "steel":   "1B4F72",       # secondary header
    "elite":   "1A5276",
    "high":    "1E8449",
    "value":   "117A65",
    "fair":    "626567",
    "over":    "922B21",
    "white":   "FFFFFF",
    "light":   "EBF5FB",
    "ice":     "E8F8F5",       # alternate row for defensive tables
    "header":  "0B6E4F",       # teal headers for JamesTown brand
    "upgrade": "1A5276",
    "rota":    "7D6608",
    "depth":   "117A65",
    "jt_dark": "05433C",       # dark JamesTown green
    "dvi_pos": "1E8449",       # positive DVI
    "dvi_neg": "922B21",       # negative DVI
}

# ── Data loading ───────────────────────────────────────────────────────────────

def load_leagues(leagues: list[str] | None, min_minutes: int) -> pd.DataFrame:
    if leagues is None:
        paths = sorted(p for p in WYSCOUT_DIR.glob("*.xlsx") if p.stem not in SKIP_FILES)
    else:
        paths = [WYSCOUT_DIR / f"{lg}.xlsx" for lg in leagues]

    frames: list[pd.DataFrame] = []
    for path in paths:
        if not path.exists():
            print(f"  [warn] {path} not found — skipping")
            continue
        lg = path.stem
        try:
            df = pd.read_excel(path)
        except Exception as e:
            print(f"  [warn] Could not read {path.name}: {e}")
            continue
        df = df.copy()
        df["_League"] = lg
        frames.append(df)
        print(f"  Loaded {lg}: {len(df)} rows")

    if not frames:
        raise RuntimeError(f"No Wyscout files found in {WYSCOUT_DIR}")

    raw = pd.concat(frames, ignore_index=True)

    mins_col = next((c for c in ["Minutes played", "MinutesPlayed", "Minutes"] if c in raw.columns), None)
    raw["_minutes"] = pd.to_numeric(raw[mins_col], errors="coerce").fillna(0) if mins_col else 0
    raw = raw[raw["_minutes"] >= min_minutes].copy()
    print(f"  → {len(raw)} players after {min_minutes}+ min filter")
    return raw.reset_index(drop=True)


def map_position(pos_str: str) -> str:
    if not isinstance(pos_str, str):
        return "Other"
    first = pos_str.split(",")[0].strip()
    return POS_MAP.get(first, "Other")


def add_position_group(df: pd.DataFrame) -> pd.DataFrame:
    pos_col = next((c for c in ["Position", "Pos"] if c in df.columns), None)
    if pos_col:
        df["_pos_group"] = df[pos_col].apply(map_position)
        df["_full_position"] = df[pos_col].fillna("Unknown")
    else:
        df["_pos_group"] = "Other"
        df["_full_position"] = "Unknown"
    return df


# ── Defensive index helpers ────────────────────────────────────────────────────

def _pct_rank(series: pd.Series, ascending: bool = True) -> pd.Series:
    """Return 0–100 percentile rank; higher is always better after inversion."""
    return series.rank(pct=True, ascending=ascending) * 100


def _get(df: pd.DataFrame, col: str) -> pd.Series:
    if col in df.columns:
        return pd.to_numeric(df[col], errors="coerce").fillna(0)
    return pd.Series(0.0, index=df.index)


def _weighted_score(components: list[tuple[pd.Series, float]]) -> pd.Series:
    total_w = sum(w for _, w in components)
    score = sum(s * w for s, w in components)
    return score / total_w if total_w > 0 else score


# ── DIS: Defensive Impact Score ────────────────────────────────────────────────

def compute_dis(grp: pd.DataFrame) -> pd.Series:
    """
    Weighted blend of core defensive output columns, each ranked 0–100 within
    the position group before combining.
    Weights tuned to reward volume + efficiency equally.
    """
    components = [
        (_pct_rank(_get(grp, "Successful defensive actions per 90")), 3.0),
        (_pct_rank(_get(grp, "Defensive duels won, %")),             2.5),
        (_pct_rank(_get(grp, "Interceptions per 90")),               2.0),
        (_pct_rank(_get(grp, "PAdj Interceptions")),                 2.0),
        (_pct_rank(_get(grp, "Aerial duels won, %")),                1.5),
        (_pct_rank(_get(grp, "Shots blocked per 90")),               1.0),
    ]
    return _weighted_score(components).round(2)


# ── ADI: Aerial Dominance Index ────────────────────────────────────────────────

def compute_adi(grp: pd.DataFrame) -> pd.Series:
    """
    Raw ADI = aerial duels per 90 × (aerial duels won % / 100).
    This combines volume with success rate into a single "net aerial wins/90"
    number, which is then percentile-ranked within the position group.
    """
    volume  = _get(grp, "Aerial duels per 90")
    pct_won = _get(grp, "Aerial duels won, %") / 100.0
    raw_adi = volume * pct_won
    return _pct_rank(raw_adi).round(2)


# ── PADS: Pressure-Adjusted Defensive Score ────────────────────────────────────

def compute_pads(grp: pd.DataFrame) -> pd.Series:
    """
    Uses Wyscout's possession-adjusted (PAdj) metrics to strip out the bias
    from playing in a high-possession team.  Higher PADS = defender contributes
    more defensive actions relative to how often the opposition has the ball.
    """
    components = [
        (_pct_rank(_get(grp, "PAdj Interceptions")),      2.5),
        (_pct_rank(_get(grp, "PAdj Sliding tackles")),    2.5),
        (_pct_rank(_get(grp, "Shots blocked per 90")),    1.5),
        (_pct_rank(_get(grp, "Defensive duels per 90")),  1.0),
    ]
    return _weighted_score(components).round(2)


# ── PDS: Positional Discipline Score ──────────────────────────────────────────

def compute_pds(grp: pd.DataFrame) -> pd.Series:
    """
    A high PDS means the player defends cleanly without giving away dangerous
    fouls or cards.  The composite foul-cost metric is ranked low-is-better
    (ascending=False → inverted rank so high rank = disciplined).
    """
    fouls   = _get(grp, "Fouls per 90")
    yellows = _get(grp, "Yellow cards per 90")
    reds    = _get(grp, "Red cards per 90")
    cost    = fouls * 0.6 + yellows * 0.3 + reds * 0.1
    # Invert: lowest cost → highest rank
    return _pct_rank(cost, ascending=False).round(2)


# ── JDR: JamesTown Defensive Rating ───────────────────────────────────────────

def compute_jdr(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compute all five indexes position-by-position, then combine into JDR.
    JDR = DIS(40%) + PADS(25%) + ADI(20%) + PDS(15%), all already 0–100.
    Final JDR is re-ranked as a percentile within the position group.
    """
    df = df.copy()
    for col in ["_DIS", "_ADI", "_PADS", "_PDS", "_JDR_raw", "_JDR"]:
        df[col] = np.nan

    for pos in DEFENSIVE_POSITIONS:
        mask = df["_pos_group"] == pos
        if mask.sum() == 0:
            continue
        grp = df.loc[mask].copy()

        dis  = compute_dis(grp)
        adi  = compute_adi(grp)
        pads = compute_pads(grp)
        pds  = compute_pds(grp)

        jdr_raw = dis * 0.40 + pads * 0.25 + adi * 0.20 + pds * 0.15

        df.loc[mask, "_DIS"]     = dis.values
        df.loc[mask, "_ADI"]     = adi.values
        df.loc[mask, "_PADS"]    = pads.values
        df.loc[mask, "_PDS"]     = pds.values
        df.loc[mask, "_JDR_raw"] = jdr_raw.values

    # JDR = percentile within position (mirrors SQS Rank in Lamberts model)
    for pos in DEFENSIVE_POSITIONS:
        mask = df["_pos_group"] == pos
        if mask.sum() == 0:
            continue
        df.loc[mask, "_JDR"] = (
            df.loc[mask, "_JDR_raw"].rank(pct=True) * 100
        ).round(2)

    return df


# ── Market value & DVI ────────────────────────────────────────────────────────

def compute_mv_rank(df: pd.DataFrame) -> pd.DataFrame:
    mv_col = next((c for c in ["Market value", "MarketValue"] if c in df.columns), None)
    mv = pd.to_numeric(df[mv_col], errors="coerce").fillna(0) if mv_col else pd.Series(0.0, index=df.index)
    df["_mkt_val"] = mv
    df["_mv_rank"] = mv.rank(pct=True) * 100
    return df


def compute_dvi(df: pd.DataFrame) -> pd.DataFrame:
    """DVI = JDR Rank − Market Value Rank  (positive = undervalued)."""
    df["_DVI"] = (df["_JDR"] - df["_mv_rank"]).round(2)

    def tier(v: float) -> str:
        if v >= 30:  return "ELITE VALUE"
        if v >= 20:  return "HIGH VALUE"
        if v >= 10:  return "VALUE"
        if v >= 0:   return "FAIR VALUE"
        return "OVERPRICED"

    df["_tier"] = df["_DVI"].apply(tier)
    return df


def compute_vs_hradec(df: pd.DataFrame) -> pd.DataFrame:
    vs_col, status_col = [], []
    for _, row in df.iterrows():
        pos = row.get("_pos_group", "Other")
        hradec_q = HRADEC_DEFENSE.get(pos, 50.0)
        jdr = float(row.get("_JDR", 50.0) or 50.0)
        gap = round(jdr - hradec_q, 1)
        vs_col.append(gap)
        if gap > 0:
            status_col.append("CLEAR UPGRADE")
        elif gap > -10:
            status_col.append("ROTATIONAL / COVER")
        else:
            status_col.append("DEPTH")
    df["_vs_hradec"] = vs_col
    df["_status"] = status_col
    return df


def model_value(mkt: float, jdr: float) -> float:
    ratio = max(jdr / 50.0, 0.1)
    return round(mkt * ratio, 0)


def val_ratio_str(mkt: float, model: float) -> str:
    if mkt <= 0:
        return "N/A"
    return f"{model / mkt:.1f}×"


# ── Master table ──────────────────────────────────────────────────────────────

# Extra defensive stats to surface in the output
DEF_STAT_MAP = {
    "Def Actions/90":     "Successful defensive actions per 90",
    "Def Duels/90":       "Defensive duels per 90",
    "Def Duel Won %":     "Defensive duels won, %",
    "Aerial/90":          "Aerial duels per 90",
    "Aerial Won %":       "Aerial duels won, %",
    "Interceptions/90":   "Interceptions per 90",
    "PAdj Intercept":     "PAdj Interceptions",
    "Slides/90":          "Sliding tackles per 90",
    "PAdj Slides":        "PAdj Sliding tackles",
    "Blocks/90":          "Shots blocked per 90",
    "Fouls/90":           "Fouls per 90",
    "Yellow/90":          "Yellow cards per 90",
    # GK specific
    "Save %":             "Save rate, %",
    "Prevented Gls/90":   "Prevented goals per 90",
    "Exits/90":           "Exits per 90",
    # Passing — matters for modern defenders
    "Prog Pass/90":       "Progressive passes per 90",
    "Acc Pass %":         "Accurate passes, %",
    "Acc Long Pass %":    "Accurate long passes, %",
}

OUTPUT_COLS = [
    "Player", "Team", "League", "Pos", "Full Position", "Age",
    "Contract", "Exp?", "Mkt Val (€)", "Model Val (€)", "Val Ratio",
    "Tier", "JDR", "DVI", "DIS", "ADI", "PADS", "PDS",
    "Status", "vs Hradec", "Minutes",
] + list(DEF_STAT_MAP.keys())


def build_master(df: pd.DataFrame) -> pd.DataFrame:
    contract_col = next((c for c in ["Contract expires", "ContractExpires"] if c in df.columns), None)
    rows: list[dict] = []

    for _, r in df.iterrows():
        mkt = float(r.get("_mkt_val", 0) or 0)
        jdr = float(r.get("_JDR", 0) or 0)
        mod_val = model_value(mkt, jdr)

        contract = r.get(contract_col) if contract_col else None
        if pd.notna(contract):
            try:
                contract = pd.to_datetime(contract).strftime("%Y-%m-%d")
            except Exception:
                contract = str(contract)
        else:
            contract = None

        exp_year = None
        if contract:
            try:
                exp_year = str(pd.to_datetime(contract).year)
            except Exception:
                pass

        row: dict = {
            "Player":         r.get("Player", ""),
            "Team":           r.get("Team", ""),
            "League":         r.get("_League", ""),
            "Pos":            r.get("_pos_group", ""),
            "Full Position":  r.get("_full_position", ""),
            "Age":            int(r.get("Age", 0)) if pd.notna(r.get("Age")) else "",
            "Contract":       contract,
            "Exp?":           exp_year,
            "Mkt Val (€)":    int(mkt) if mkt > 0 else 0,
            "Model Val (€)":  int(mod_val),
            "Val Ratio":      val_ratio_str(mkt, mod_val),
            "Tier":           r.get("_tier", ""),
            "JDR":            round(float(r.get("_JDR",  0) or 0), 2),
            "DVI":            round(float(r.get("_DVI",  0) or 0), 2),
            "DIS":            round(float(r.get("_DIS",  0) or 0), 2),
            "ADI":            round(float(r.get("_ADI",  0) or 0), 2),
            "PADS":           round(float(r.get("_PADS", 0) or 0), 2),
            "PDS":            round(float(r.get("_PDS",  0) or 0), 2),
            "Status":         r.get("_status", ""),
            "vs Hradec":      r.get("_vs_hradec", 0),
            "Minutes":        int(r.get("_minutes", 0)),
        }

        for out_col, in_col in DEF_STAT_MAP.items():
            val = r.get(in_col, 0)
            try:
                row[out_col] = round(float(val), 2) if pd.notna(val) else 0
            except Exception:
                row[out_col] = 0

        rows.append(row)

    master = pd.DataFrame(rows)
    master = master.sort_values("DVI", ascending=False).reset_index(drop=True)
    return master


# ── openpyxl helpers (small summary sheets) ───────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        try:
            max_len = max(
                len(str(col_cells[0].value or "")),
                *(len(str(c.value or "")) for c in col_cells[1:10]),
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)
        except Exception:
            pass


# ── xlsxwriter helpers (large data sheets) ────────────────────────────────────

def _xf(wb_x, bold=False, font_color="#000000", bg_color=None,
         size=9, italic=False, align="center", wrap=False, border=True):
    """Create an xlsxwriter format object."""
    props: dict = {"font_size": size, "bold": bold, "italic": italic,
                   "font_color": font_color, "align": align, "valign": "vcenter"}
    if wrap:
        props["text_wrap"] = True
    if bg_color:
        props["bg_color"] = bg_color
        props["pattern"] = 1
    if border:
        props.update({"border": 1, "border_color": "#CCCCCC"})
    return wb_x.add_format(props)


def write_data_sheet_fast(wb_x, sheet_name: str, title: str, subtitle: str,
                           df: pd.DataFrame) -> None:
    """Write a large DataFrame to an xlsxwriter sheet with styled header rows."""
    if df.empty:
        ws = wb_x.add_worksheet(sheet_name)
        ws.write(0, 0, title)
        return

    ws = wb_x.add_worksheet(sheet_name)
    ws.set_zoom(85)

    cols = list(df.columns)
    n_cols = len(cols)

    # Row 1 — title
    fmt_title = _xf(wb_x, bold=True, font_color="#FFFFFF", bg_color=f"#{C['jt_dark']}",
                    size=13, align="left", border=False)
    ws.merge_range(0, 0, 0, n_cols - 1, title, fmt_title)
    ws.set_row(0, 22)

    # Row 2 — subtitle
    fmt_sub = _xf(wb_x, italic=True, font_color=f"#{C['gold']}", bg_color=f"#{C['jt_dark']}",
                  size=9, align="left", border=False)
    ws.merge_range(1, 0, 1, n_cols - 1, subtitle, fmt_sub)
    ws.set_row(1, 14)

    # Row 3 — header
    fmt_hdr = _xf(wb_x, bold=True, font_color="#FFFFFF", bg_color=f"#{C['header']}",
                  size=9, wrap=True)
    for j, col in enumerate(cols):
        ws.write(2, j, col, fmt_hdr)
    ws.set_row(2, 28)

    # Freeze below header
    ws.freeze_panes(3, 0)

    # Pre-build formats
    TIER_BG = {
        "ELITE VALUE": f"#{C['elite']}",
        "HIGH VALUE":  f"#{C['high']}",
        "VALUE":       f"#{C['value']}",
        "FAIR VALUE":  f"#{C['fair']}",
        "OVERPRICED":  f"#{C['over']}",
    }
    STATUS_BG = {
        "CLEAR UPGRADE":      f"#{C['upgrade']}",
        "ROTATIONAL / COVER": f"#{C['rota']}",
        "DEPTH":              f"#{C['depth']}",
    }

    tier_idx   = cols.index("Tier")   if "Tier"   in cols else -1
    status_idx = cols.index("Status") if "Status" in cols else -1
    dvi_idx    = cols.index("DVI")    if "DVI"    in cols else -1
    jdr_idx    = cols.index("JDR")    if "JDR"    in cols else -1

    fmt_even  = _xf(wb_x, bg_color=f"#{C['ice']}")
    fmt_odd   = _xf(wb_x, bg_color=f"#{C['white']}")
    fmt_bold_e = _xf(wb_x, bold=True, bg_color=f"#{C['ice']}")
    fmt_bold_o = _xf(wb_x, bold=True, bg_color=f"#{C['white']}")
    fmt_dvi_pos_e = _xf(wb_x, bold=True, font_color=f"#{C['dvi_pos']}", bg_color=f"#{C['ice']}")
    fmt_dvi_neg_e = _xf(wb_x, bold=True, font_color=f"#{C['dvi_neg']}", bg_color=f"#{C['ice']}")
    fmt_dvi_pos_o = _xf(wb_x, bold=True, font_color=f"#{C['dvi_pos']}", bg_color=f"#{C['white']}")
    fmt_dvi_neg_o = _xf(wb_x, bold=True, font_color=f"#{C['dvi_neg']}", bg_color=f"#{C['white']}")
    tier_fmts  = {k: _xf(wb_x, bold=True, font_color="#FFFFFF", bg_color=v)
                  for k, v in TIER_BG.items()}
    status_fmts = {k: _xf(wb_x, bold=True, font_color="#FFFFFF", bg_color=v)
                   for k, v in STATUS_BG.items()}

    for i, row_vals in enumerate(df.itertuples(index=False)):
        xrow = i + 3
        even = (i % 2 == 0)
        base_fmt = fmt_even if even else fmt_odd
        bold_fmt = fmt_bold_e if even else fmt_bold_o
        dvi_pos_f = fmt_dvi_pos_e if even else fmt_dvi_pos_o
        dvi_neg_f = fmt_dvi_neg_e if even else fmt_dvi_neg_o

        for j, val in enumerate(row_vals):
            # Pick format
            if j == tier_idx:
                fmt = tier_fmts.get(str(val), base_fmt)
            elif j == status_idx:
                fmt = status_fmts.get(str(val), base_fmt)
            elif j == dvi_idx:
                try:
                    fmt = dvi_pos_f if float(val or 0) >= 0 else dvi_neg_f
                except Exception:
                    fmt = base_fmt
            elif j == jdr_idx:
                fmt = bold_fmt
            else:
                fmt = base_fmt

            if val is None or (isinstance(val, float) and np.isnan(val)):
                ws.write_blank(xrow, j, None, fmt)
            elif isinstance(val, (int, float)):
                ws.write_number(xrow, j, val, fmt)
            else:
                ws.write_string(xrow, j, str(val), fmt)

    # Column widths (estimate)
    for j, col in enumerate(cols):
        ws.set_column(j, j, min(max(len(col) + 2, 10), 22))


# ── openpyxl-based small-sheet writer (kept for backward compat) ──────────────

def write_data_sheet(ws, title: str, subtitle: str, df: pd.DataFrame) -> None:
    """openpyxl writer — use only for small sheets (< 2000 rows)."""
    ws.append([title])
    n_cols = max(len(df.columns), 10)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws["A1"].font = Font(bold=True, color=C["white"], size=13)
    ws["A1"].fill = _fill(C["jt_dark"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.append([subtitle])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["jt_dark"])
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    if df.empty:
        return

    ws.append(list(df.columns))
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font = Font(bold=True, color=C["white"], size=9)
        cell.fill = _fill(C["header"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[hdr_row].height = 28

    cols = list(df.columns)
    tier_idx   = cols.index("Tier")   + 1 if "Tier"   in cols else None
    status_idx = cols.index("Status") + 1 if "Status" in cols else None
    dvi_idx    = cols.index("DVI")    + 1 if "DVI"    in cols else None
    jdr_idx    = cols.index("JDR")    + 1 if "JDR"    in cols else None

    TIER_HEX = {
        "ELITE VALUE": C["elite"], "HIGH VALUE": C["high"], "VALUE": C["value"],
        "FAIR VALUE": C["fair"], "OVERPRICED": C["over"],
    }
    STATUS_HEX = {
        "CLEAR UPGRADE": C["upgrade"], "ROTATIONAL / COVER": C["rota"], "DEPTH": C["depth"],
    }

    for i, row_vals in enumerate(df.itertuples(index=False), start=1):
        ws.append(list(row_vals))
        data_row = ws.max_row
        bg = C["ice"] if i % 2 == 0 else C["white"]
        for cell in ws[data_row]:
            cell.font = Font(size=9)
            cell.fill = _fill(bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()
        if tier_idx:
            tc = ws.cell(data_row, tier_idx)
            h = TIER_HEX.get(str(tc.value or ""))
            if h:
                tc.fill = _fill(h); tc.font = Font(bold=True, color=C["white"], size=9)
        if status_idx:
            sc = ws.cell(data_row, status_idx)
            h = STATUS_HEX.get(str(sc.value or ""))
            if h:
                sc.fill = _fill(h); sc.font = Font(bold=True, color=C["white"], size=9)
        if dvi_idx:
            dc = ws.cell(data_row, dvi_idx)
            try:
                v = float(dc.value or 0)
                dc.font = Font(bold=True, size=9,
                               color=C["dvi_pos"] if v >= 0 else C["dvi_neg"])
            except Exception:
                pass
        if jdr_idx:
            ws.cell(data_row, jdr_idx).font = Font(bold=True, size=9)

    ws.freeze_panes = f"A{hdr_row + 1}"
    _autofit(ws)


# ── README ────────────────────────────────────────────────────────────────────

def build_readme(ws, leagues: list[str], total: int, clear: int, budget: int) -> None:
    ws.title = "README"
    ws.sheet_view.showGridLines = False

    ws.append(["JAMESTOWN ANALYTICS — DEFENSIVE INTELLIGENCE REPORT  ·  FC Hradec Králové 2025–26"])
    ws.merge_cells("A1:T1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=15)
    ws["A1"].fill = _fill(C["jt_dark"])
    ws.row_dimensions[1].height = 30

    league_label = f"{len(leagues)} leagues" if len(leagues) > 5 else " + ".join(leagues)
    ws.append([f"Waltzing Analytics  ·  JamesTown methodology  ·  "
               f"{league_label}  ·  Positions: GK · CB · FB · DM  ·  Budget ≤ €{budget:,}"])
    ws.merge_cells("A2:T2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=10)
    ws["A2"].fill = _fill(C["jt_dark"])
    ws.row_dimensions[2].height = 18

    ws.append([None])

    ws.append([None, "WORKBOOK STRUCTURE"])
    ws[f"B{ws.max_row}"].font = Font(bold=True, size=11, color=C["jt_dark"])
    ws.row_dimensions[ws.max_row].height = 20

    ws.append([None, "Sheet", "Contents"])
    hdr = ws.max_row
    for col_letter in "BC":
        cell = ws[f"{col_letter}{hdr}"]
        cell.font = Font(bold=True, color=C["white"])
        cell.fill = _fill(C["header"])
        cell.alignment = Alignment(horizontal="left")

    structure = [
        ("README",                "This guide — methodology and index definitions"),
        ("Defensive Priority",    f"{clear} clear upgrades ranked by DVI (Defensive Value Index)"),
        ("Elite Defenders",       "Players with DVI ≥ 30 — strongest buy signals"),
        ("JamesTown Analysis",    "Tier and position breakdown with league statistics"),
        ("Top 5 per Position",    "Best 5 GK / CB / FB / DM by DVI"),
        ("All Defenders",         f"Full {total} candidate database with all indexes"),
        ("GK",                    "Goalkeeper targets — JDR + Save % + Prevented goals"),
        ("CB",                    "Centre-back targets — DIS + PADS + ADI"),
        ("FB",                    "Full-back targets — DIS + ADI + discipline"),
        ("DM",                    "Defensive mid targets — DIS + PADS + PDS"),
        ("Defensive Metrics",     "Deep-dive: all five indexes side by side for every player"),
        ("Expiring 2026",         "Contract expiry opportunities — free or discounted"),
        ("Budget Planner",        f"Build defensive unit within €{budget:,}"),
        ("Squad",                 "Hradec Králové current defensive quality benchmarks"),
    ]
    for sheet, desc in structure:
        ws.append([None, sheet, desc])
        row = ws.max_row
        ws[f"B{row}"].font = Font(bold=True, color=C["jt_dark"])

    ws.append([None])
    ws.append([None, "DEFENSIVE INDEX DEFINITIONS"])
    ws[f"B{ws.max_row}"].font = Font(bold=True, size=11, color=C["jt_dark"])

    defs = [
        ("DIS — Defensive Impact Score",
         "Core output rank (0–100 within position).  "
         "Weights: Successful def actions/90 ×3 · Def duels won% ×2.5 · "
         "Interceptions/90 ×2 · PAdj Interceptions ×2 · Aerial won% ×1.5 · Blocks/90 ×1"),

        ("ADI — Aerial Dominance Index",
         "Net aerial wins per 90 = (Aerial duels/90) × (Aerial won% / 100), then ranked 0–100.  "
         "Combines volume with success quality — a player who wins 80% of 5 aerials beats "
         "one who wins 50% of 6."),

        ("PADS — Pressure-Adjusted Defensive Score",
         "Uses Wyscout PAdj metrics to neutralise possession-context bias.  "
         "Weights: PAdj Interceptions ×2.5 · PAdj Sliding tackles ×2.5 · Blocks/90 ×1.5 · "
         "Def duels/90 ×1.  High PADS = defender works hard regardless of team possession."),

        ("PDS — Positional Discipline Score",
         "Inverted rank of foul cost = Fouls/90×0.6 + Yellow/90×0.3 + Red/90×0.1.  "
         "High PDS = defends cleanly, low foul rate, rarely booked.  "
         "Crucial for players expected to play week-to-week in a back four."),

        ("JDR — JamesTown Defensive Rating",
         "Master 0–100 composite: DIS×40% + PADS×25% + ADI×20% + PDS×15%, "
         "re-percentiled within position.  Analogous to SQS Rank in the Lamberts model."),

        ("DVI — Defensive Value Index",
         "JDR Rank − Market Value Rank.  Positive = undervalued; negative = overpriced.  "
         "ELITE ≥30 · HIGH ≥20 · VALUE ≥10 · FAIR 0–9 · OVERPRICED <0"),

        ("vs Hradec",
         "Target JDR − weakest Hradec starter at same position.  "
         ">0 = CLEAR UPGRADE  ·  −10 to 0 = ROTATIONAL  ·  <−10 = DEPTH"),
    ]
    for term, desc in defs:
        ws.append([None, term, desc])
        row = ws.max_row
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].fill = _fill(C["ice"])

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 90


# ── JamesTown Analysis sheet ──────────────────────────────────────────────────

def build_jt_analysis(ws, master: pd.DataFrame) -> None:
    ws.title = "JamesTown Analysis"
    ws.sheet_view.showGridLines = False

    ws.append(["JAMESTOWN ANALYTICS — DEFENSIVE INDEX BREAKDOWN"])
    ws.merge_cells("A1:N1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=13)
    ws["A1"].fill = _fill(C["jt_dark"])
    ws.row_dimensions[1].height = 22

    ws.append(["DVI = JDR Rank − Market Value Rank  ·  ELITE ≥30  ·  HIGH ≥20  ·  VALUE ≥10  ·  FAIR 0–9  ·  OVER <0"])
    ws.merge_cells("A2:N2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["jt_dark"])
    ws.row_dimensions[2].height = 14

    ws.append([None])

    # Overall DVI stats
    ws.append(["OVERALL DVI STATISTICS"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)

    dvi = master["DVI"]
    overall = [
        ("Total Defenders",  len(master)),
        ("Mean DVI",         round(dvi.mean(), 2)),
        ("Median DVI",       round(dvi.median(), 2)),
        ("Max DVI",          round(dvi.max(), 2)),
        ("Min DVI",          round(dvi.min(), 2)),
        ("Std Dev",          round(dvi.std(), 2)),
    ]

    # Position breakdown
    pos_stats = master.groupby("Pos")[["JDR", "DIS", "PADS", "ADI", "PDS"]].mean().round(2)

    ws.append(["Metric", "Value", None, "Tier", "Count", "%", None, "Position", "Avg JDR", "Avg DIS", "Avg PADS"])
    hdr = ws.max_row
    for cell in ws[hdr]:
        if cell.value:
            cell.font = Font(bold=True, color=C["white"])
            cell.fill = _fill(C["header"])

    tier_counts = master["Tier"].value_counts()
    tier_order = ["ELITE VALUE", "HIGH VALUE", "VALUE", "FAIR VALUE", "OVERPRICED"]
    pos_order = ["GK", "CB", "FB", "DM"]

    for i, (metric, val) in enumerate(overall):
        row = ws.max_row + 1
        ws.append([metric, val])

        tier_name = tier_order[i] if i < len(tier_order) else ""
        cnt = int(tier_counts.get(tier_name, 0))
        pct = round(cnt / len(master) * 100, 1) if len(master) > 0 else 0
        ws.cell(row, 4).value = tier_name
        ws.cell(row, 5).value = cnt
        ws.cell(row, 6).value = f"{pct}%"

        if i < len(pos_order):
            p = pos_order[i]
            if p in pos_stats.index:
                ws.cell(row, 8).value  = p
                ws.cell(row, 9).value  = float(pos_stats.loc[p, "JDR"])
                ws.cell(row, 10).value = float(pos_stats.loc[p, "DIS"])
                ws.cell(row, 11).value = float(pos_stats.loc[p, "PADS"])

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["H"].width = 12
    for col in ["I", "J", "K"]:
        ws.column_dimensions[col].width = 14


# ── Top 5 per position ────────────────────────────────────────────────────────

def build_top5(ws, master: pd.DataFrame) -> None:
    ws.title = "Top 5 per Position"
    ws.sheet_view.showGridLines = False

    ws.append(["JAMESTOWN ANALYTICS — TOP 5 DEFENDERS PER POSITION  ·  Ranked by DVI"])
    ws.merge_cells("A1:N1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=13)
    ws["A1"].fill = _fill(C["jt_dark"])
    ws.row_dimensions[1].height = 22

    ws.append(["Highest Defensive Value Index per position group across all recruitment leagues"])
    ws.merge_cells("A2:N2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["jt_dark"])

    medals = ["🥇", "🥈", "🥉", "4.", "5."]
    pos_labels = {
        "GK": "GK — GOALKEEPER",
        "CB": "CB — CENTRE-BACK",
        "FB": "FB — FULL-BACK",
        "DM": "DM — DEFENSIVE MID",
    }

    for pos, label in pos_labels.items():
        grp = master[master["Pos"] == pos].head(5)
        if grp.empty:
            continue

        ws.append([f"  {label}"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=14)
        ws[f"A{ws.max_row}"].font = Font(bold=True, color=C["white"], size=10)
        ws[f"A{ws.max_row}"].fill = _fill(C["navy"])
        ws.row_dimensions[ws.max_row].height = 18

        ws.append(["#", "Player", "Team", "League", "Age", "Contract",
                   "Mkt Val (€)", "JDR", "DVI", "DIS", "ADI", "PADS", "PDS", "Status"])
        hdr = ws.max_row
        for cell in ws[hdr]:
            cell.font = Font(bold=True, color=C["white"], size=9)
            cell.fill = _fill(C["header"])
            cell.alignment = Alignment(horizontal="center")

        for i, (_, r) in enumerate(grp.iterrows()):
            ws.append([
                medals[i], r["Player"], r["Team"], r["League"],
                r["Age"], r["Contract"], r["Mkt Val (€)"],
                r["JDR"], r["DVI"], r["DIS"], r["ADI"], r["PADS"], r["PDS"],
                r["Status"],
            ])
            dr = ws.max_row
            bg = C["ice"] if i % 2 == 0 else C["white"]
            for cell in ws[dr]:
                cell.font = Font(size=9)
                cell.fill = _fill(bg)
                cell.alignment = Alignment(horizontal="center")
            ws.cell(dr, 1).font = Font(bold=True, size=10)

        ws.append([None])

    _autofit(ws)


# ── Defensive Metrics deep-dive sheet ────────────────────────────────────────

DEF_METRIC_COLS = [
    "Player", "Team", "League", "Pos", "Age",
    "JDR", "DVI", "DIS", "ADI", "PADS", "PDS",
    "Def Actions/90", "Def Duels/90", "Def Duel Won %",
    "Aerial/90", "Aerial Won %",
    "Interceptions/90", "PAdj Intercept",
    "Slides/90", "PAdj Slides", "Blocks/90",
    "Fouls/90", "Yellow/90",
]


def build_defensive_metrics(wb_x, master: pd.DataFrame) -> None:
    """Fast xlsxwriter version of the defensive metrics deep-dive sheet."""
    cols = [c for c in DEF_METRIC_COLS if c in master.columns]
    df = master[cols].sort_values("JDR", ascending=False).reset_index(drop=True)
    write_data_sheet_fast(
        wb_x, "Defensive Metrics",
        "JAMESTOWN ANALYTICS — DEFENSIVE METRICS DEEP DIVE",
        "All five JamesTown indexes + underlying raw metrics — sorted by JDR",
        df,
    )


# ── Squad sheet ───────────────────────────────────────────────────────────────

def build_squad(ws) -> None:
    ws.title = "Squad"
    ws.sheet_view.showGridLines = False

    ws.append(["FC HRADEC KRÁLOVÉ — CURRENT DEFENSIVE QUALITY BENCHMARKS"])
    ws.merge_cells("A1:L1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=13)
    ws["A1"].fill = _fill(C["jt_dark"])
    ws.row_dimensions[1].height = 22

    ws.append(["Czech top-flight percentile ranks  ·  Higher = better  ·  Basis for vs Hradec calculation"])
    ws.merge_cells("A2:L2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["jt_dark"])

    ws.append([None])
    ws.append(["DEFENSIVE SQUAD QUALITY"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)

    ws.append(["Position", "Hradec Quality %", "Note"])
    hdr = ws.max_row
    for cell in ws[hdr]:
        if cell.value:
            cell.font = Font(bold=True, color=C["white"])
            cell.fill = _fill(C["header"])
            cell.alignment = Alignment(horizontal="center")

    pos_notes = {
        "GK": "Adam Zadrazil — starter",
        "CB": "Daniel Horak — weakest starter",
        "FB": "Martin Suchomel — starter",
        "DM": "Jakub Elbel — starter",
    }
    for pos, quality in HRADEC_DEFENSE.items():
        ws.append([pos, quality, pos_notes.get(pos, "")])
        row = ws.max_row
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40


# ── Budget Planner ────────────────────────────────────────────────────────────

def build_budget_planner(ws, master: pd.DataFrame, budget: int) -> None:
    ws.title = "Budget Planner"
    ws.sheet_view.showGridLines = False

    ws.append([f"DEFENSIVE BUDGET PLANNER  ·  FC Hradec Králové  ·  Cap: €{budget:,}"])
    ws.merge_cells("A1:O1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=13)
    ws["A1"].fill = _fill(C["jt_dark"])
    ws.row_dimensions[1].height = 22

    ws.append(["ELITE VALUE + HIGH VALUE clear upgrades sorted by market value (cheapest first)."])
    ws.merge_cells("A2:O2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["jt_dark"])

    ws.append([None, None, None, None, "Total Budget", budget])

    eligible = master[
        (master["Tier"].isin(["ELITE VALUE", "HIGH VALUE"])) &
        (master["Status"] == "CLEAR UPGRADE")
    ].sort_values("Mkt Val (€)").reset_index(drop=True)

    cols = ["#", "Player", "Pos", "Team", "League", "Age", "Contract",
            "Mkt Val (€)", "Model Val (€)", "Val Ratio", "JDR", "DVI",
            "DIS", "PADS", "vs Hradec", "Running Total (€)"]

    ws.append(cols)
    hdr = ws.max_row
    for cell in ws[hdr]:
        cell.font = Font(bold=True, color=C["white"], size=9)
        cell.fill = _fill(C["header"])
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hdr].height = 28

    running = 0
    for i, (_, r) in enumerate(eligible.iterrows(), 1):
        mv = int(r["Mkt Val (€)"])
        running += mv
        over = running > budget
        ws.append([
            i, r["Player"], r["Pos"], r["Team"], r["League"],
            r["Age"], r["Contract"],
            mv, int(r["Model Val (€)"]), r["Val Ratio"],
            r["JDR"], r["DVI"], r["DIS"], r["PADS"],
            r["vs Hradec"], running,
        ])
        dr = ws.max_row
        for cell in ws[dr]:
            cell.font = Font(size=9, color="922B21" if over else "000000")
            cell.fill = _fill("FDE8E8" if over else (C["ice"] if i % 2 == 0 else C["white"]))
            cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = f"A{hdr + 1}"
    _autofit(ws)


# ── Main ───────────────────────────────────────────────────────────────────────

def run(leagues: list[str] | None, min_minutes: int, max_age: int,
        budget: int, output: Path) -> None:
    print(f"\n{'='*62}")
    print("  JamesTown Analytics — Defensive Intelligence Report")
    print(f"  Positions: GK · CB · FB · DM")
    print(f"  Leagues: {'ALL' if leagues is None else leagues}")
    print(f"  Min minutes: {min_minutes}  |  Max age: {max_age}")
    print(f"{'='*62}\n")

    print("Loading Wyscout files…")
    raw = load_leagues(leagues, min_minutes)

    if "Age" in raw.columns:
        raw = raw[pd.to_numeric(raw["Age"], errors="coerce").fillna(99) <= max_age]
        print(f"  → {len(raw)} players after age ≤ {max_age} filter")

    raw = add_position_group(raw)
    raw = raw[raw["_pos_group"].isin(DEFENSIVE_POSITIONS)].copy()
    print(f"  → {len(raw)} defensive players (GK/CB/FB/DM)")

    print("Computing JamesTown defensive indexes…")
    raw = compute_jdr(raw)
    raw = compute_mv_rank(raw)
    raw = compute_dvi(raw)
    raw = compute_vs_hradec(raw)

    print("Building master table…")
    master = build_master(raw)
    master_out = master[[c for c in OUTPUT_COLS if c in master.columns]]

    total     = len(master)
    clear     = int((master["Status"] == "CLEAR UPGRADE").sum())
    elite     = int((master["Tier"] == "ELITE VALUE").sum())
    print(f"  → {total} defenders total")
    print(f"  → {clear} clear upgrades")
    print(f"  → {elite} ELITE VALUE")

    print(f"\nWriting workbook → {output}")
    output.parent.mkdir(parents=True, exist_ok=True)

    league_list = leagues if leagues else sorted(master["League"].unique().tolist())

    # ── Phase 1: large data sheets via xlsxwriter (fast) ──────────────────────
    print("  Building data sheets (xlsxwriter)…")
    xls_path = output.with_suffix(".tmp.xlsx")

    wb_x = xlsxwriter.Workbook(str(xls_path), {"constant_memory": False})

    prio_df = master_out[master_out["Status"] == "CLEAR UPGRADE"].copy()
    print(f"    Defensive Priority ({len(prio_df)} rows)…")
    write_data_sheet_fast(
        wb_x, "Defensive Priority",
        f"DEFENSIVE PRIORITY LIST — {len(prio_df)} Clear Upgrades — Ranked by DVI",
        "All are clear JDR improvements on current Hradec starters  ·  Sorted by Defensive Value Index",
        prio_df,
    )

    elite_df = master_out[master_out["Tier"] == "ELITE VALUE"].copy()
    print(f"    Elite Defenders ({len(elite_df)} rows)…")
    write_data_sheet_fast(
        wb_x, "Elite Defenders",
        f"ELITE DEFENSIVE VALUE — {len(elite_df)} Players with DVI ≥ 30",
        "Strongest defensive buy signals  ·  Sorted by DVI",
        elite_df,
    )

    print(f"    All Defenders ({total} rows)…")
    write_data_sheet_fast(
        wb_x, "All Defenders",
        f"ALL DEFENSIVE TARGETS — {total} Candidates  ·  {len(league_list)} leagues",
        "Use column filters to narrow by position, league, age, tier  ·  Ranked by DVI",
        master_out,
    )

    pos_config = [
        ("GK", "GOALKEEPER TARGETS"),
        ("CB", "CENTRE-BACK TARGETS"),
        ("FB", "FULL-BACK TARGETS"),
        ("DM", "DEFENSIVE MID TARGETS"),
    ]
    for pos, label in pos_config:
        grp = master_out[master_out["Pos"] == pos].copy()
        print(f"    {pos} ({len(grp)} rows)…")
        write_data_sheet_fast(
            wb_x, pos,
            f"{label}  ·  {len(grp)} candidates",
            "Sorted by DVI  ·  All JamesTown defensive indexes included",
            grp,
        )

    exp_df = master_out[master_out["Exp?"].astype(str) == "2026"].copy()
    print(f"    Expiring 2026 ({len(exp_df)} rows)…")
    write_data_sheet_fast(
        wb_x, "Expiring 2026",
        f"EXPIRING CONTRACTS 2026 — {len(exp_df)} DEFENSIVE PLAYERS  ·  Free/cut-price",
        "",
        exp_df,
    )

    print("    Defensive Metrics…")
    build_defensive_metrics(wb_x, master)

    wb_x.close()
    print("  xlsxwriter sheets done.")

    # ── Phase 2: summary/structural sheets via openpyxl ───────────────────────
    print("  Building summary sheets (openpyxl)…")
    from openpyxl import load_workbook
    wb = load_workbook(str(xls_path))

    # Insert README at front
    ws_readme = wb.create_sheet("README", 0)
    build_readme(ws_readme, league_list, total, clear, budget)

    # JamesTown Analysis
    ws_analysis = wb.create_sheet("JamesTown Analysis", 3)
    build_jt_analysis(ws_analysis, master)

    # Top 5 per Position
    ws_top5 = wb.create_sheet("Top 5 per Position", 4)
    build_top5(ws_top5, master_out)

    # Budget Planner
    ws_budget = wb.create_sheet("Budget Planner")
    build_budget_planner(ws_budget, master, budget)

    # Squad
    ws_squad = wb.create_sheet("Squad")
    build_squad(ws_squad)

    # Remove xlsxwriter default sheet if present (named "Sheet1")
    if "Sheet1" in wb.sheetnames:
        del wb["Sheet1"]

    wb.save(output)
    xls_path.unlink(missing_ok=True)

    size_mb = output.stat().st_size / 1_048_576
    print(f"\nDone. {size_mb:.1f} MB → {output.resolve()}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="JamesTown Analytics — Defensive Intelligence Report from Wyscout data"
    )
    parser.add_argument("--leagues", nargs="+", default=None,
                        help="League names (without .xlsx). Omit for ALL leagues.")
    parser.add_argument("--min-minutes", type=int, default=DEFAULT_MIN_MINUTES)
    parser.add_argument("--max-age",     type=int, default=DEFAULT_MAX_AGE)
    parser.add_argument("--budget",      type=int, default=DEFAULT_BUDGET)
    parser.add_argument("--output", type=Path,
                        default=ROOT / "data" / "JamesTown_Defensive_Report.xlsx")
    args = parser.parse_args()
    run(
        leagues=args.leagues,
        min_minutes=args.min_minutes,
        max_age=args.max_age,
        budget=args.budget,
        output=args.output,
    )
