"""
screening_app.py
─────────────────
Long-list screening tool for FC Hradec Králové.

Flow:
  1. Choose data source (IMPECT or Wyscout) and leagues/season
  2. Apply filters (position, age, minutes, score thresholds, bio)
  3. Preview the resulting long list in an interactive table
  4. Export to Excel — one sheet per position group, summary dashboard sheet

Run:  streamlit run screening_app.py
"""
from __future__ import annotations

import io
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st

# ── Project imports ────────────────────────────────────────────────────────────
APP_DIR = Path(__file__).parent

from wyscout_model import (
    build_wyscout_dashboard_data,
    available_leagues as wyscout_available_leagues,
    ALL_SCORE_COLS as WYSCOUT_SCORE_COLS,
    WYSCOUT_POSITION_MAP,
)

# ── Constants ──────────────────────────────────────────────────────────────────
IMPECT_SCORE_COLS = [
    "CompositeRecruitmentScore",
    "DecisionScore",
    "ValueRecruitmentScore",
    "ScoringThreatScore",
    "CreativeProgressionScore",
    "DefensiveDisruptionScore",
    "PressingScore",
    "BallSecurityScore",
    "ExpectedThreatScore",
    "ASA_GoalsAddedScore",
    "AgeResaleScore",
    "PerformanceReliabilityScore",
    "SuccessProbability",
]

POSITION_ORDER = ["GK", "CB", "FB", "DM", "CM", "AM", "W", "ST"]

POSITION_COLORS = {
    "GK": "#667085", "CB": "#2f5f98", "FB": "#00a6a6",
    "DM": "#6b8e23", "CM": "#2f855a", "AM": "#d97706",
    "W": "#e76f51",  "ST": "#c2410c",
}

# Key display columns per source
WYSCOUT_DISPLAY_COLS = [
    "Player", "Team", "Position", "Age", "_League",
    "Minutes played", "Market value", "Contract expires",
    "Foot", "Height", "Birth country", "Passport country",
]
IMPECT_DISPLAY_COLS = [
    "PlayerName", "TeamName", "PositionGroup", "AgeYears", "MinutesPlayed",
    "LeagueLabel", "CountryLabel", "SeasonLabel",
]

# Score display labels
SCORE_LABELS = {
    "CompositeRecruitmentScore": "Composite",
    "DecisionScore": "Decision",
    "ValueRecruitmentScore": "Value",
    "ScoringThreatScore": "Scoring",
    "CreativeProgressionScore": "Creative",
    "DefensiveDisruptionScore": "Defensive",
    "PressingScore": "Pressing",
    "BallSecurityScore": "Ball Security",
    "ExpectedThreatScore": "Exp. Threat",
    "ASA_GoalsAddedScore": "Goals Added",
    "AerialScore": "Aerial",
    "SetPieceScore": "Set Piece",
    "AgeResaleScore": "Age/Resale",
    "PerformanceReliabilityScore": "Reliability",
    "SuccessProbability": "Success Prob.",
}


# ── Data loaders ───────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def _load_impect() -> pd.DataFrame:
    req = APP_DIR / "data" / "FCHK Model V3 - Recruitment Scores.xlsx"
    if not req.exists():
        return pd.DataFrame()
    df = pd.read_excel(req)
    df.columns = [str(c).strip() for c in df.columns]
    # Merge supplementary sheets
    for fname in [
        "FCHK Model V3 - Player Scores.xlsx",
        "FCHK Model V3 - Player Styles.xlsx",
    ]:
        p = APP_DIR / "data" / fname
        if p.exists():
            sup = pd.read_excel(p)
            sup.columns = [str(c).strip() for c in sup.columns]
            key_cols = [c for c in ["PlayerName", "TeamName", "PositionGroup"] if c in df.columns and c in sup.columns]
            if key_cols:
                new_cols = [c for c in sup.columns if c not in df.columns]
                if new_cols:
                    df = df.merge(sup[key_cols + new_cols], on=key_cols, how="left")
    return df


@st.cache_data(show_spinner=False, ttl=600)
def _load_wyscout(leagues_key: tuple[str, ...] | None, min_minutes: int) -> pd.DataFrame:
    leagues = list(leagues_key) if leagues_key else None
    return build_wyscout_dashboard_data(min_minutes=min_minutes, leagues=leagues)


def get_data(source: str, leagues_key: tuple | None, min_minutes: int) -> pd.DataFrame:
    if source == "Wyscout":
        cache_key = ("scr_ws", leagues_key, min_minutes)
        if st.session_state.get("_scr_cache_key") != cache_key:
            with st.spinner("Loading Wyscout model…"):
                df = _load_wyscout(leagues_key, min_minutes)
            st.session_state["_scr_data"] = df
            st.session_state["_scr_cache_key"] = cache_key
        return st.session_state.get("_scr_data", pd.DataFrame())
    else:
        if "_scr_impect" not in st.session_state:
            with st.spinner("Loading IMPECT model…"):
                st.session_state["_scr_impect"] = _load_impect()
        return st.session_state.get("_scr_impect", pd.DataFrame())


# ── Excel export ───────────────────────────────────────────────────────────────

def _autofit(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 50)


def build_excel(df: pd.DataFrame, source: str, filters_summary: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    red_fill    = PatternFill("solid", fgColor="EE3A27")
    white_font  = Font(color="FFFFFF", bold=True, size=11)
    bold_font   = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws_sum["A1"] = "FC Hradec Králové — Long List Screening"
    ws_sum["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws_sum["A1"].fill = header_fill
    ws_sum["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[1].height = 30
    ws_sum.merge_cells("A1:H1")

    ws_sum["A2"] = f"Source: {source}  |  {filters_summary}  |  Total players: {len(df)}"
    ws_sum["A2"].font = Font(italic=True, size=9, color="888888")
    ws_sum.merge_cells("A2:H2")

    # Position breakdown table
    pos_col = "PositionGroup" if "PositionGroup" in df.columns else "Position"
    pos_counts = df[pos_col].value_counts().reindex(POSITION_ORDER).dropna().astype(int)
    score_col  = "CompositeRecruitmentScore" if "CompositeRecruitmentScore" in df.columns else None

    headers = ["Position", "Players", "Avg Score", "Top Score", "Median Age"]
    for ci, h in enumerate(headers, start=1):
        cell = ws_sum.cell(row=4, column=ci, value=h)
        cell.fill = red_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for ri, (pos, cnt) in enumerate(pos_counts.items(), start=5):
        sub = df.loc[df[pos_col] == pos]
        avg_score = round(sub[score_col].mean(), 1) if score_col else "—"
        top_score = round(sub[score_col].max(),  1) if score_col else "—"
        med_age   = round(sub["AgeYears"].median(), 1) if "AgeYears" in sub.columns else "—"
        row_data  = [pos, cnt, avg_score, top_score, med_age]
        pos_fill  = PatternFill("solid", fgColor=POSITION_COLORS.get(pos, "CCCCCC").lstrip("#"))
        for ci, val in enumerate(row_data, start=1):
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if ci == 1:
                cell.fill = pos_fill
                cell.font = Font(color="FFFFFF", bold=True)

    for col_cells in ws_sum.columns:
        ws_sum.column_dimensions[col_cells[0].column_letter].width = 18

    # ── Per-position sheets ────────────────────────────────────────────────────
    score_cols = [c for c in (WYSCOUT_SCORE_COLS if source == "Wyscout" else IMPECT_SCORE_COLS) if c in df.columns]
    bio_cols = [c for c in (
        WYSCOUT_DISPLAY_COLS if source == "Wyscout" else IMPECT_DISPLAY_COLS
    ) if c in df.columns]

    display_order = bio_cols + [c for c in score_cols if c not in bio_cols]

    positions = [p for p in POSITION_ORDER if p in (pos_counts.index.tolist() if len(pos_counts) else df[pos_col].unique())]
    if not positions:
        positions = sorted(df[pos_col].dropna().unique())

    for pos in positions:
        sub = df.loc[df[pos_col] == pos].copy()
        if sub.empty:
            continue
        if score_col and score_col in sub.columns:
            sub = sub.sort_values(score_col, ascending=False)

        out_cols = [c for c in display_order if c in sub.columns]
        sheet_df = sub[out_cols].reset_index(drop=True)

        ws = wb.create_sheet(title=pos)
        pos_color = POSITION_COLORS.get(pos, "#888888").lstrip("#")
        pos_fill  = PatternFill("solid", fgColor=pos_color)

        # Title row
        ws.cell(row=1, column=1, value=f"{pos} — {len(sub)} players")
        ws.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=12)
        ws.cell(row=1, column=1).fill = pos_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_cols))

        # Header row
        for ci, col_name in enumerate(out_cols, start=1):
            cell = ws.cell(row=2, column=ci, value=SCORE_LABELS.get(col_name, col_name))
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows with score colouring
        score_indices = {col: i + 1 for i, col in enumerate(out_cols) if col in score_cols}
        for ri, row in enumerate(dataframe_to_rows(sheet_df, index=False, header=False), start=3):
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if ci > 2 else "left", wrap_text=False)
                # Score gradient: red-white-green
                col_name = out_cols[ci - 1]
                if col_name in score_cols and isinstance(val, (int, float)) and not pd.isna(val):
                    v = max(0, min(100, float(val)))
                    if v >= 70:
                        r, g, b = int(255 - (v - 70) / 30 * 100), 200, int(150 - (v - 70) / 30 * 50)
                    elif v >= 40:
                        r, g, b = 255, int(150 + (v - 40) / 30 * 50), 150
                    else:
                        r, g, b = 255, int(100 + v / 40 * 50), 100
                    cell.fill = PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")

            if ri % 2 == 0:
                for ci_alt in range(1, len(out_cols) + 1):
                    existing = ws.cell(row=ri, column=ci_alt).fill
                    if existing.patternType == "none":
                        ws.cell(row=ri, column=ci_alt).fill = PatternFill("solid", fgColor="F5F5F0")

        _autofit(ws)

    # ── Full list sheet ────────────────────────────────────────────────────────
    ws_all = wb.create_sheet(title="All Players")
    out_cols_all = [c for c in display_order if c in df.columns]
    all_df = df[out_cols_all].copy()
    if score_col and score_col in all_df.columns:
        all_df = all_df.sort_values(score_col, ascending=False)

    ws_all.cell(row=1, column=1, value=f"All Players — {len(all_df)} total")
    ws_all.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=12)
    ws_all.cell(row=1, column=1).fill = header_fill
    ws_all.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_cols_all))
    ws_all.row_dimensions[1].height = 24

    for ci, col_name in enumerate(out_cols_all, start=1):
        cell = ws_all.cell(row=2, column=ci, value=SCORE_LABELS.get(col_name, col_name))
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(dataframe_to_rows(all_df.reset_index(drop=True), index=False, header=False), start=3):
        for ci, val in enumerate(row, start=1):
            ws_all.cell(row=ri, column=ci, value=val)

    _autofit(ws_all)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Score filter widget ────────────────────────────────────────────────────────

def score_filter_block(df: pd.DataFrame, source: str) -> dict[str, tuple[float, float]]:
    """Render score threshold sliders, return {col: (min, max)} dict."""
    cols = [c for c in (WYSCOUT_SCORE_COLS if source == "Wyscout" else IMPECT_SCORE_COLS) if c in df.columns]
    thresholds: dict[str, tuple[float, float]] = {}
    if not cols:
        return thresholds

    st.markdown("**Score thresholds** *(0–100)*")
    n_cols = 2
    groups = [cols[i:i + n_cols] for i in range(0, len(cols), n_cols)]
    for group in groups:
        row = st.columns(len(group))
        for col_widget, score_col in zip(row, group):
            with col_widget:
                label = SCORE_LABELS.get(score_col, score_col)
                val = st.slider(label, 0, 100, (0, 100), key=f"scr_score_{score_col}", label_visibility="visible")
                if val != (0, 100):
                    thresholds[score_col] = val
    return thresholds


# ── App layout ─────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="FCHK Long List Screener",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
/* ── FiveThirtyEight-style base ── */
html, body, [data-testid="stApp"] {
    background-color: #f9f8f3;
    color: #1a1a1a;
    font-family: "Atlas Grotesk", "Inter", sans-serif;
}
[data-testid="stSidebar"] {
    background-color: #1a1a2e;
}
[data-testid="stSidebar"] * {
    color: #e0e0e0 !important;
}
[data-testid="stSidebar"] label, [data-testid="stSidebar"] .stMarkdown p {
    color: #b0b0b0 !important;
    font-size: 0.75rem !important;
    text-transform: uppercase;
    letter-spacing: 0.06em;
}
.page-title {
    font-size: 2.2rem;
    font-weight: 800;
    color: #1a1a1a;
    border-left: 5px solid #ee3a27;
    padding-left: 14px;
    margin-bottom: 4px;
    line-height: 1.15;
}
.page-sub {
    font-size: 0.85rem;
    color: #666;
    padding-left: 19px;
    margin-bottom: 20px;
}
.stat-card {
    background: white;
    border-radius: 6px;
    padding: 16px 20px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.08);
    text-align: center;
}
.stat-card .stat-num {
    font-size: 2.4rem;
    font-weight: 800;
    color: #ee3a27;
    line-height: 1;
}
.stat-card .stat-label {
    font-size: 0.72rem;
    color: #888;
    text-transform: uppercase;
    letter-spacing: 0.07em;
    margin-top: 4px;
}
.section-hdr {
    font-size: 0.68rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: #aaa;
    margin: 18px 0 6px 0;
    padding-bottom: 4px;
    border-bottom: 1px solid #333;
}
.pos-badge {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 0.7rem;
    font-weight: 700;
    color: white;
    margin-right: 4px;
}
/* Score bar */
.score-bar-wrap { background: #eee; border-radius: 3px; height: 6px; width: 100%; }
.score-bar-fill { height: 6px; border-radius: 3px; }
div[data-testid="stDataFrame"] { border: 1px solid #e0e0e0; border-radius: 6px; }
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════
#  SIDEBAR — all filters
# ════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("<div style='padding: 18px 0 10px 0;'>", unsafe_allow_html=True)
    st.image("https://upload.wikimedia.org/wikipedia/en/thumb/3/3c/FC_Hradec_Kr%C3%A1lov%C3%A9_logo.svg/200px-FC_Hradec_Kr%C3%A1lov%C3%A9_logo.svg.png", width=56)
    st.markdown(
        "<div style='font-size:1.1rem;font-weight:800;color:#fff;margin-top:6px;'>Long List Screener</div>"
        "<div style='font-size:0.7rem;color:#888;margin-bottom:12px;'>FC Hradec Králové · Scouting</div>",
        unsafe_allow_html=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)

    # ── Data source ────────────────────────────────────────────────
    st.markdown("<div class='section-hdr'>Data Source</div>", unsafe_allow_html=True)
    source = st.radio("Source", ["IMPECT", "Wyscout"], horizontal=True, key="scr_source", label_visibility="collapsed")

    leagues_key: tuple | None = None
    min_minutes = 400

    if source == "Wyscout":
        all_leagues = wyscout_available_leagues()
        sel_leagues = st.multiselect(
            "Leagues", all_leagues, default=[],
            placeholder="All leagues",
            key="scr_leagues",
            label_visibility="collapsed",
        )
        leagues_key = tuple(sorted(sel_leagues)) if sel_leagues else None
        min_minutes = st.number_input("Min. minutes played", 0, 3000, 400, 50, key="scr_min_min")
    else:
        st.markdown(
            "<div style='font-size:0.72rem;color:#777;padding:4px 0 8px 0;'>IMPECT model — 16 European leagues</div>",
            unsafe_allow_html=True,
        )

    # ── Position ───────────────────────────────────────────────────
    st.markdown("<div class='section-hdr'>Position</div>", unsafe_allow_html=True)
    sel_positions = st.multiselect(
        "Positions", POSITION_ORDER,
        default=[], placeholder="All positions",
        key="scr_positions", label_visibility="collapsed",
    )

    # ── Age ────────────────────────────────────────────────────────
    st.markdown("<div class='section-hdr'>Age</div>", unsafe_allow_html=True)
    age_range = st.slider("Age range", 15, 40, (16, 32), key="scr_age", label_visibility="collapsed")

    # ── Minutes ────────────────────────────────────────────────────
    st.markdown("<div class='section-hdr'>Minutes</div>", unsafe_allow_html=True)
    min_min_filter = st.number_input("Min. minutes (filter)", 0, 5000, 400, 50, key="scr_mins_filter")

    # ── Biographical filters ───────────────────────────────────────
    st.markdown("<div class='section-hdr'>Biographical</div>", unsafe_allow_html=True)
    foot_opts = ["Either", "Right", "Left"]
    sel_foot = st.selectbox("Preferred foot", foot_opts, key="scr_foot", label_visibility="collapsed")

    nationality_input = st.text_input(
        "Nationality / passport (comma-sep)", key="scr_nationality",
        placeholder="e.g. Czech, Slovak",
        label_visibility="collapsed",
    )
    on_loan_only = st.checkbox("On loan only", key="scr_on_loan")
    contract_before = st.text_input("Contract expires before", placeholder="e.g. 2026", key="scr_contract", label_visibility="collapsed")

    # ── Market value ───────────────────────────────────────────────
    st.markdown("<div class='section-hdr'>Market Value (€M)</div>", unsafe_allow_html=True)
    mv_range = st.slider("Market value €M", 0.0, 50.0, (0.0, 50.0), 0.5, key="scr_mv", label_visibility="collapsed")

    # ── Max list size ──────────────────────────────────────────────
    st.markdown("<div class='section-hdr'>List Size</div>", unsafe_allow_html=True)
    max_list = st.number_input("Max players per position", 5, 500, 50, 5, key="scr_max_list", label_visibility="collapsed")

    st.divider()
    run_btn = st.button("🔍  Build Long List", type="primary", use_container_width=True, key="scr_run")


# ════════════════════════════════════════════════════════════════════
#  MAIN AREA
# ════════════════════════════════════════════════════════════════════

st.markdown(
    "<div class='page-title'>Long List Screener</div>"
    "<div class='page-sub'>Filter, score, and export player long lists from IMPECT or Wyscout data.</div>",
    unsafe_allow_html=True,
)

# ── Score threshold section (full width, before table) ─────────────
st.markdown("---")

# Load data
raw_df = get_data(source, leagues_key, min_minutes)

if raw_df.empty:
    st.warning("No data loaded. Check that data files are present in `data/`.")
    st.stop()

# ── Score filter (only shown when data is loaded) ──────────────────
with st.expander("⚙️  Score thresholds", expanded=False):
    score_thresholds = score_filter_block(raw_df, source)

# ════════════════════════════════════════════════════════════════════
#  APPLY FILTERS
# ════════════════════════════════════════════════════════════════════

df = raw_df.copy()

# Position column
pos_col = "PositionGroup" if "PositionGroup" in df.columns else "Position"
age_col = "AgeYears" if "AgeYears" in df.columns else "Age"
mins_col = "MinutesPlayed" if "MinutesPlayed" in df.columns else "Minutes played"
name_col = "PlayerName" if "PlayerName" in df.columns else "Player"

# Position filter
if sel_positions:
    df = df.loc[df[pos_col].isin(sel_positions)]

# Age filter
if age_col in df.columns:
    age_s = pd.to_numeric(df[age_col], errors="coerce")
    df = df.loc[age_s.between(age_range[0], age_range[1])]

# Minutes filter
if mins_col in df.columns:
    mins_s = pd.to_numeric(df[mins_col], errors="coerce").fillna(0)
    df = df.loc[mins_s >= min_min_filter]

# Score thresholds
for col, (lo, hi) in score_thresholds.items():
    if col in df.columns:
        s = pd.to_numeric(df[col], errors="coerce")
        df = df.loc[s.between(lo, hi)]

# Foot filter
if sel_foot != "Either" and "Foot" in df.columns:
    df = df.loc[df["Foot"].fillna("").str.lower().str.startswith(sel_foot.lower())]

# Nationality filter
if nationality_input.strip():
    terms = [t.strip().lower() for t in nationality_input.split(",") if t.strip()]
    nat_cols = [c for c in ["Birth country", "Passport country", "Nationality", "CountryLabel"] if c in df.columns]
    if nat_cols and terms:
        mask = pd.Series(False, index=df.index)
        for nc in nat_cols:
            for t in terms:
                mask |= df[nc].fillna("").str.lower().str.contains(t, regex=False)
        df = df.loc[mask]

# On loan filter
if on_loan_only and "On loan" in df.columns:
    df = df.loc[df["On loan"].fillna("").astype(str).str.lower().isin(["yes", "true", "1"])]

# Contract filter
if contract_before.strip() and "Contract expires" in df.columns:
    try:
        year_limit = int(contract_before.strip())
        contract_years = pd.to_numeric(
            df["Contract expires"].fillna("").astype(str).str[:4], errors="coerce"
        )
        df = df.loc[contract_years <= year_limit]
    except ValueError:
        pass

# Market value filter
if "Market value" in df.columns:
    mv_euros = pd.to_numeric(
        df["Market value"].fillna("").astype(str)
        .str.replace("€", "", regex=False)
        .str.replace("M", "e6", regex=False)
        .str.replace("K", "e3", regex=False)
        .str.replace(",", "", regex=False),
        errors="coerce",
    )
    mv_m = mv_euros / 1e6
    low_m, high_m = mv_range
    if low_m > 0 or high_m < 50.0:
        df = df.loc[mv_m.fillna(0).between(low_m, high_m)]

# Sort by composite score descending
score_sort = "CompositeRecruitmentScore" if "CompositeRecruitmentScore" in df.columns else None
if score_sort:
    df = df.sort_values(score_sort, ascending=False)

# Cap per position
if max_list < 500:
    capped_frames = []
    for pos, grp in df.groupby(pos_col):
        capped_frames.append(grp.head(max_list))
    df = pd.concat(capped_frames, ignore_index=True) if capped_frames else df

df = df.reset_index(drop=True)

# ════════════════════════════════════════════════════════════════════
#  KPI CARDS
# ════════════════════════════════════════════════════════════════════

n_total   = len(df)
n_pos     = df[pos_col].nunique() if pos_col in df.columns else 0
avg_score = round(df["CompositeRecruitmentScore"].mean(), 1) if "CompositeRecruitmentScore" in df.columns and n_total > 0 else "—"
avg_age   = round(pd.to_numeric(df[age_col], errors="coerce").mean(), 1) if age_col in df.columns and n_total > 0 else "—"

k1, k2, k3, k4 = st.columns(4)
for col_w, num, label in [
    (k1, n_total,   "Players on list"),
    (k2, n_pos,     "Positions covered"),
    (k3, avg_score, "Avg composite score"),
    (k4, avg_age,   "Avg age"),
]:
    col_w.markdown(
        f"<div class='stat-card'>"
        f"<div class='stat-num'>{num}</div>"
        f"<div class='stat-label'>{label}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

st.markdown("<br>", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════
#  POSITION BREAKDOWN
# ════════════════════════════════════════════════════════════════════

if n_total > 0 and pos_col in df.columns:
    pos_counts = df[pos_col].value_counts().reindex(POSITION_ORDER).dropna().astype(int)
    bar_cols = st.columns(len(pos_counts))
    for i, (pos, cnt) in enumerate(pos_counts.items()):
        color = POSITION_COLORS.get(pos, "#aaa")
        bar_cols[i].markdown(
            f"<div style='text-align:center;'>"
            f"<div style='background:{color};color:white;font-weight:800;font-size:1.1rem;"
            f"border-radius:6px;padding:8px 0;'>{cnt}</div>"
            f"<div style='font-size:0.7rem;color:#888;margin-top:4px;font-weight:700;'>{pos}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )

st.markdown("<br>", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════
#  TABLE TABS — one per position group
# ════════════════════════════════════════════════════════════════════

if n_total == 0:
    st.info("No players match the current filters. Adjust the criteria in the sidebar.")
else:
    score_cols_present = [c for c in (WYSCOUT_SCORE_COLS if source == "Wyscout" else IMPECT_SCORE_COLS) if c in df.columns]
    bio_cols_present   = [c for c in (WYSCOUT_DISPLAY_COLS if source == "Wyscout" else IMPECT_DISPLAY_COLS) if c in df.columns]

    display_cols = bio_cols_present + [c for c in score_cols_present if c not in bio_cols_present]

    available_positions = [p for p in POSITION_ORDER if p in df[pos_col].values] if pos_col in df.columns else []
    tab_labels = ["All"] + available_positions
    tabs = st.tabs(tab_labels)

    for tab, label in zip(tabs, tab_labels):
        with tab:
            subset = df if label == "All" else df.loc[df[pos_col] == label]
            show_cols = [c for c in display_cols if c in subset.columns]

            if subset.empty:
                st.info(f"No {label} players in the current list.")
                continue

            # Rename for display
            rename_map = {c: SCORE_LABELS.get(c, c) for c in show_cols}
            display_df = subset[show_cols].copy().rename(columns=rename_map)

            # Round score columns
            for c in score_cols_present:
                disp_name = SCORE_LABELS.get(c, c)
                if disp_name in display_df.columns:
                    display_df[disp_name] = pd.to_numeric(display_df[disp_name], errors="coerce").round(1)

            st.dataframe(
                display_df.reset_index(drop=True),
                use_container_width=True,
                height=min(600, 56 + len(display_df) * 36),
                column_config={
                    SCORE_LABELS.get(c, c): st.column_config.ProgressColumn(
                        label=SCORE_LABELS.get(c, c),
                        min_value=0, max_value=100,
                        format="%.1f",
                    )
                    for c in score_cols_present if SCORE_LABELS.get(c, c) in display_df.columns
                },
            )
            st.caption(f"{len(display_df)} players · sorted by Composite score")

# ════════════════════════════════════════════════════════════════════
#  EXPORT
# ════════════════════════════════════════════════════════════════════

st.markdown("---")
st.markdown("### Export Long List")

export_col, info_col = st.columns([2, 3])

with export_col:
    if n_total > 0:
        filters_summary = (
            f"Pos: {', '.join(sel_positions) or 'All'} | "
            f"Age: {age_range[0]}–{age_range[1]} | "
            f"Min min: {min_min_filter} | "
            f"Scores filtered: {len(score_thresholds)}"
        )
        excel_bytes = build_excel(df, source, filters_summary)
        st.download_button(
            label="⬇️  Download Excel (.xlsx)",
            data=excel_bytes,
            file_name="FCHK_Long_List.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.button("⬇️  Download Excel", disabled=True, use_container_width=True)

with info_col:
    st.markdown(
        "<div style='font-size:0.8rem;color:#888;padding-top:8px;'>"
        "Excel file contains: <b>Summary</b> sheet (position breakdown) + "
        "one sheet per position group (colour-coded scores) + <b>All Players</b> sheet."
        "</div>",
        unsafe_allow_html=True,
    )
