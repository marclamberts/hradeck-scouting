"""
pass_danger_app.py
──────────────────
Wyscout Pass Danger Index — standalone Streamlit app.

The Pass Danger Index (PDI) measures how threatening a player's passing is,
built from five sub-dimensions scored 0–100 within position group:

  Incisiveness   — key passes, through passes, passes to penalty area
  Creativity     — smart passes, xA, shot assists
  Progression    — progressive passes, passes to final third, deep completions
  Delivery       — crosses to box, deep completed crosses
  Accuracy       — weighted accuracy across dangerous pass types

Each dimension is 0–100 (percentile within position group).
PDI = position-weighted composite of the five dimensions.

Run: streamlit run pass_danger_app.py
"""
from __future__ import annotations

import io
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import norm
import streamlit as st

APP_DIR = Path(__file__).parent

from wyscout_model import available_leagues as wyscout_available_leagues

# ── Metric definitions ─────────────────────────────────────────────────────────

# (metric, weight) blueprints per PDI dimension
DIMENSIONS: dict[str, list[tuple[str, float]]] = {
    "Incisiveness": [
        ("Key passes per 90",                   3.0),
        ("Through passes per 90",               2.5),
        ("Passes to penalty area per 90",        2.0),
        ("Accurate through passes, %",           1.5),
        ("Accurate passes to penalty area, %",   1.5),
    ],
    "Creativity": [
        ("xA per 90",                            3.5),
        ("Smart passes per 90",                  2.5),
        ("Shot assists per 90",                  2.0),
        ("Accurate smart passes, %",             1.5),
        ("Second assists per 90",                1.0),
    ],
    "Progression": [
        ("Progressive passes per 90",            3.0),
        ("Passes to final third per 90",          2.5),
        ("Deep completions per 90",               2.0),
        ("Accurate progressive passes, %",        1.5),
        ("Accurate passes to final third, %",     1.0),
    ],
    "Delivery": [
        ("Deep completed crosses per 90",         3.0),
        ("Crosses to goalie box per 90",          2.5),
        ("Accurate crosses, %",                   2.0),
        ("Key passes per 90",                     1.0),
    ],
    "Accuracy": [
        ("Accurate passes, %",                    2.0),
        ("Accurate short / medium passes, %",     1.5),
        ("Accurate long passes, %",               1.5),
        ("Accurate forward passes, %",            2.0),
        ("Accurate progressive passes, %",        1.5),
    ],
}

# Position relevance weights for each dimension
POSITION_WEIGHTS: dict[str, dict[str, float]] = {
    "ST":  {"Incisiveness": 0.8, "Creativity": 1.0, "Progression": 0.8, "Delivery": 0.6, "Accuracy": 0.8},
    "W":   {"Incisiveness": 1.2, "Creativity": 1.2, "Progression": 1.0, "Delivery": 1.5, "Accuracy": 0.9},
    "AM":  {"Incisiveness": 1.5, "Creativity": 1.6, "Progression": 1.1, "Delivery": 0.7, "Accuracy": 1.0},
    "CM":  {"Incisiveness": 1.2, "Creativity": 1.1, "Progression": 1.4, "Delivery": 0.8, "Accuracy": 1.2},
    "DM":  {"Incisiveness": 0.8, "Creativity": 0.7, "Progression": 1.5, "Delivery": 0.6, "Accuracy": 1.4},
    "FB":  {"Incisiveness": 0.9, "Creativity": 1.0, "Progression": 1.2, "Delivery": 1.8, "Accuracy": 1.0},
    "CB":  {"Incisiveness": 0.5, "Creativity": 0.5, "Progression": 1.3, "Delivery": 0.6, "Accuracy": 1.5},
    "GK":  {"Incisiveness": 0.3, "Creativity": 0.3, "Progression": 1.0, "Delivery": 0.4, "Accuracy": 1.8},
}

DIMENSION_COLORS = {
    "Incisiveness": "#ee3a27",
    "Creativity":   "#d97706",
    "Progression":  "#2f855a",
    "Delivery":     "#2f5f98",
    "Accuracy":     "#6b21a8",
}

POSITION_ORDER = ["GK", "CB", "FB", "DM", "CM", "AM", "W", "ST"]
POSITION_COLORS = {
    "GK": "#667085", "CB": "#2f5f98", "FB": "#00a6a6",
    "DM": "#6b8e23", "CM": "#2f855a", "AM": "#d97706",
    "W":  "#e76f51",  "ST": "#c2410c",
}

ALL_METRICS = list({m for dim in DIMENSIONS.values() for m, _ in dim})

# ── Data loading ───────────────────────────────────────────────────────────────

WYSCOUT_DB = APP_DIR / "data" / "Wyscout DB"


@st.cache_data(show_spinner=False)
def load_wyscout_raw(
    leagues_key: tuple[str, ...] | None,
    min_minutes: int,
) -> pd.DataFrame:
    if not WYSCOUT_DB.exists():
        return pd.DataFrame()
    files = sorted(WYSCOUT_DB.glob("*.xlsx"))
    if leagues_key:
        files = [f for f in files if f.stem in leagues_key]

    frames = []
    for path in files:
        try:
            df = pd.read_excel(path)
            df.columns = [str(c).strip() for c in df.columns]
            for col in ["Position", "Pos"]:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.split(",").str[0].str.strip()
            df["_League"] = path.stem
            frames.append(df)
        except Exception:
            continue

    if not frames:
        return pd.DataFrame()

    raw = pd.concat(frames, ignore_index=True)

    # Numeric coercion
    skip = {"Player", "Team", "Position", "Birth country", "Passport country",
            "Foot", "On loan", "Team within selected timeframe",
            "Contract expires", "_League"}
    for col in raw.columns:
        if col not in skip:
            c = pd.to_numeric(raw[col], errors="coerce")
            if c.notna().any():
                raw[col] = c

    # Minutes filter
    mins_col = next((c for c in ["Minutes played", "MinutesPlayed"] if c in raw.columns), None)
    if mins_col:
        raw = raw.loc[pd.to_numeric(raw[mins_col], errors="coerce").fillna(0) >= min_minutes]

    # Position group
    from wyscout_model import WYSCOUT_POSITION_MAP
    pos_col = next((c for c in ["Position", "Pos"] if c in raw.columns), None)
    if pos_col:
        raw["PositionGroup"] = raw[pos_col].map(WYSCOUT_POSITION_MAP).fillna("Other")

    return raw.reset_index(drop=True)


# ── Scoring engine ─────────────────────────────────────────────────────────────

def _z_to_pct(z: np.ndarray) -> np.ndarray:
    return norm.cdf(z) * 100


def compute_pdi(df: pd.DataFrame) -> pd.DataFrame:
    """Add dimension scores (0–100) and PDI to every row."""
    df = df.copy()
    dim_cols = list(DIMENSIONS.keys())

    scored_frames: list[pd.DataFrame] = []
    for pos_group, grp in df.groupby("PositionGroup"):
        grp = grp.copy()
        pos_str = str(pos_group)
        pos_rel = POSITION_WEIGHTS.get(pos_str, {d: 1.0 for d in dim_cols})

        for dim, blueprint in DIMENSIONS.items():
            available = [(m, w) for m, w in blueprint if m in grp.columns]
            if not available:
                grp[dim] = 50.0
                continue
            total_w = sum(w for _, w in available)
            z = pd.Series(0.0, index=grp.index)
            for metric, weight in available:
                col = pd.to_numeric(grp[metric], errors="coerce").fillna(0)
                mu  = col.mean()
                sig = col.std() or 1e-9
                z  += (weight / total_w) * (col - mu) / sig
            grp[dim] = _z_to_pct(z.values)

        # PDI = position-weighted average of dimensions
        total_rel = sum(pos_rel.get(d, 1.0) for d in dim_cols)
        pdi = sum(grp[d] * pos_rel.get(d, 1.0) for d in dim_cols) / total_rel
        grp["PDI"] = pdi.clip(0, 100)

        scored_frames.append(grp)

    return pd.concat(scored_frames, ignore_index=True) if scored_frames else df


# ── Excel export ───────────────────────────────────────────────────────────────

def build_excel(df: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    white_bold  = Font(color="FFFFFF", bold=True)
    thin        = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _autofit(ws):
        for col_cells in ws.columns:
            w = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(w + 3, 45)

    def _score_fill(v: float) -> PatternFill:
        v = max(0, min(100, v))
        if v >= 70:
            r, g, b = int(155 + (v - 70) / 30 * 50), 200, 120
        elif v >= 40:
            r, g, b = 255, int(180 + (v - 40) / 30 * 20), 120
        else:
            r, g, b = 255, int(100 + v / 40 * 80), 100
        return PatternFill("solid", fgColor=f"{int(r):02X}{int(g):02X}{int(b):02X}")

    bio_cols   = ["Player", "Team", "Position", "PositionGroup", "Age",
                  "_League", "Minutes played", "Market value", "Contract expires",
                  "Foot", "Birth country"]
    score_cols = ["PDI"] + list(DIMENSIONS.keys())
    metric_cols = [m for m in ALL_METRICS if m in df.columns]
    all_cols   = [c for c in bio_cols + score_cols + metric_cols if c in df.columns]

    pos_col = "PositionGroup"

    positions = [p for p in POSITION_ORDER if p in df[pos_col].values]

    # All Players sheet
    ws_all = wb.active
    ws_all.title = "All Players"
    sub = df.sort_values("PDI", ascending=False)[all_cols].reset_index(drop=True)

    ws_all.cell(1, 1, f"Pass Danger Index — All Players ({len(sub)})")
    ws_all.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=13)
    ws_all.cell(1, 1).fill = header_fill
    ws_all.row_dimensions[1].height = 26
    ws_all.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_cols))

    for ci, col in enumerate(all_cols, 1):
        c = ws_all.cell(2, ci, col)
        c.fill = header_fill; c.font = white_bold; c.border = thin
        c.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(dataframe_to_rows(sub, index=False, header=False), 3):
        for ci, val in enumerate(row, 1):
            cell = ws_all.cell(ri, ci, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if ci > 2 else "left")
            if all_cols[ci - 1] in score_cols and isinstance(val, (int, float)) and not pd.isna(val):
                cell.fill = _score_fill(float(val))
    _autofit(ws_all)

    # Per-position sheets
    for pos in positions:
        sub_pos = df.loc[df[pos_col] == pos].sort_values("PDI", ascending=False)
        if sub_pos.empty:
            continue
        out = sub_pos[[c for c in all_cols if c in sub_pos.columns]].reset_index(drop=True)
        ws = wb.create_sheet(title=pos)
        pos_fill = PatternFill("solid", fgColor=POSITION_COLORS.get(pos, "888888").lstrip("#"))

        ws.cell(1, 1, f"{pos} — {len(sub_pos)} players")
        ws.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=12)
        ws.cell(1, 1).fill = pos_fill
        ws.row_dimensions[1].height = 24
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out.columns))

        for ci, col in enumerate(out.columns, 1):
            c = ws.cell(2, ci, col)
            c.fill = header_fill; c.font = white_bold; c.border = thin
            c.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(dataframe_to_rows(out, index=False, header=False), 3):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(ri, ci, val)
                cell.border = thin
                cell.alignment = Alignment(horizontal="center" if ci > 2 else "left")
                if out.columns[ci - 1] in score_cols and isinstance(val, (int, float)) and not pd.isna(val):
                    cell.fill = _score_fill(float(val))
        _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════
#  APP
# ════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Pass Danger Index",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
html, body, [data-testid="stApp"] {
    background-color: #f9f8f3;
    color: #1a1a1a;
}
[data-testid="stSidebar"] { background-color: #1a1a2e; }
[data-testid="stSidebar"] * { color: #e0e0e0 !important; }
[data-testid="stSidebar"] label, [data-testid="stSidebar"] .stMarkdown p {
    color: #b0b0b0 !important;
    font-size: 0.73rem !important;
    text-transform: uppercase;
    letter-spacing: 0.06em;
}
.page-title {
    font-size: 2.2rem; font-weight: 800;
    border-left: 5px solid #ee3a27;
    padding-left: 14px; margin-bottom: 4px;
}
.page-sub { font-size: 0.85rem; color: #666; padding-left: 19px; margin-bottom: 20px; }
.stat-card {
    background: white; border-radius: 6px;
    padding: 14px 18px; box-shadow: 0 1px 4px rgba(0,0,0,0.08);
    text-align: center;
}
.stat-num  { font-size: 2.2rem; font-weight: 800; color: #ee3a27; line-height: 1; }
.stat-label { font-size: 0.7rem; color: #888; text-transform: uppercase; letter-spacing:.07em; margin-top:4px; }
.dim-card {
    background: white; border-radius: 6px;
    padding: 12px 16px; box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    margin-bottom: 8px;
}
.sbar-hdr {
    font-size: 0.65rem; font-weight: 700; text-transform: uppercase;
    letter-spacing: .1em; color: #aaa; margin: 14px 0 4px;
    padding-bottom: 3px; border-bottom: 1px solid #333;
}
</style>
""", unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown(
        "<div style='padding:16px 0 10px'>"
        "<div style='font-size:1.1rem;font-weight:800;color:#fff;'>Pass Danger Index</div>"
        "<div style='font-size:0.7rem;color:#888;margin-bottom:10px;'>Wyscout · FC Hradec Králové</div>"
        "</div>",
        unsafe_allow_html=True,
    )

    # Data
    st.markdown("<div class='sbar-hdr'>Leagues</div>", unsafe_allow_html=True)
    all_leagues = wyscout_available_leagues()
    sel_leagues = st.multiselect(
        "Leagues", all_leagues, default=[],
        placeholder="All leagues", key="pdi_leagues",
        label_visibility="collapsed",
    )
    leagues_key = tuple(sorted(sel_leagues)) if sel_leagues else None
    min_minutes = st.number_input("Min. minutes", 0, 3000, 500, 50, key="pdi_min_min")

    # Filters
    st.markdown("<div class='sbar-hdr'>Position</div>", unsafe_allow_html=True)
    sel_pos = st.multiselect("Position", POSITION_ORDER, default=[], placeholder="All", key="pdi_pos", label_visibility="collapsed")

    st.markdown("<div class='sbar-hdr'>Age</div>", unsafe_allow_html=True)
    age_range = st.slider("Age", 15, 40, (16, 35), key="pdi_age", label_visibility="collapsed")

    st.markdown("<div class='sbar-hdr'>PDI threshold</div>", unsafe_allow_html=True)
    pdi_min = st.slider("Min PDI", 0, 100, 0, key="pdi_threshold", label_visibility="collapsed")

    st.markdown("<div class='sbar-hdr'>Results</div>", unsafe_allow_html=True)
    top_n = st.number_input("Top N per position", 5, 300, 30, 5, key="pdi_top_n")

    st.markdown("<div class='sbar-hdr'>Highlight dimension</div>", unsafe_allow_html=True)
    highlight_dim = st.selectbox(
        "Dimension", ["PDI (overall)"] + list(DIMENSIONS.keys()),
        key="pdi_highlight", label_visibility="collapsed",
    )


# ── Load & score ───────────────────────────────────────────────────────────────

cache_key = ("pdi", leagues_key, min_minutes)
if st.session_state.get("_pdi_cache_key") != cache_key:
    with st.spinner("Loading Wyscout data…"):
        raw = load_wyscout_raw(leagues_key, min_minutes)
    if raw.empty:
        st.error("No Wyscout data found in `data/Wyscout DB/`.")
        st.stop()
    with st.spinner("Computing Pass Danger Index…"):
        scored = compute_pdi(raw)
    st.session_state["_pdi_data"]      = scored
    st.session_state["_pdi_cache_key"] = cache_key

df = st.session_state["_pdi_data"].copy()

# ── Apply filters ──────────────────────────────────────────────────────────────

pos_col  = "PositionGroup"
age_col  = "Age"
mins_col = "Minutes played"

if sel_pos:
    df = df.loc[df[pos_col].isin(sel_pos)]

age_s = pd.to_numeric(df[age_col], errors="coerce")
df = df.loc[age_s.between(age_range[0], age_range[1])]

df = df.loc[df["PDI"] >= pdi_min]

# Cap top N per position
if top_n < 300:
    frames = [grp.sort_values("PDI", ascending=False).head(top_n)
              for _, grp in df.groupby(pos_col)]
    df = pd.concat(frames, ignore_index=True) if frames else df

df = df.sort_values("PDI", ascending=False).reset_index(drop=True)

# ── Page header ────────────────────────────────────────────────────────────────

st.markdown(
    "<div class='page-title'>Pass Danger Index</div>"
    "<div class='page-sub'>"
    "Measures how threatening a player's passing is — "
    "Incisiveness · Creativity · Progression · Delivery · Accuracy"
    "</div>",
    unsafe_allow_html=True,
)

# ── KPI row ────────────────────────────────────────────────────────────────────

n_total  = len(df)
avg_pdi  = round(df["PDI"].mean(), 1) if n_total else "—"
top_pdi  = round(df["PDI"].max(), 1)  if n_total else "—"
n_leagues_shown = df["_League"].nunique() if "_League" in df.columns else "—"

k1, k2, k3, k4 = st.columns(4)
for col_w, num, lbl in [
    (k1, n_total,         "Players"),
    (k2, avg_pdi,         "Avg PDI"),
    (k3, top_pdi,         "Top PDI"),
    (k4, n_leagues_shown, "Leagues"),
]:
    col_w.markdown(
        f"<div class='stat-card'><div class='stat-num'>{num}</div>"
        f"<div class='stat-label'>{lbl}</div></div>",
        unsafe_allow_html=True,
    )

st.markdown("<br>", unsafe_allow_html=True)

# ── Dimension legend ───────────────────────────────────────────────────────────

with st.expander("📐 How PDI dimensions are calculated", expanded=False):
    dcols = st.columns(len(DIMENSIONS))
    for col_w, (dim, blueprint) in zip(dcols, DIMENSIONS.items()):
        color = DIMENSION_COLORS[dim]
        metrics_html = "".join(
            f"<li style='font-size:.72rem;color:#555;'>{m} <span style='color:#aaa;font-size:.65rem;'>(×{w})</span></li>"
            for m, w in blueprint
        )
        col_w.markdown(
            f"<div class='dim-card'>"
            f"<div style='font-weight:800;color:{color};font-size:.85rem;margin-bottom:4px;'>{dim}</div>"
            f"<ul style='margin:0;padding-left:14px;'>{metrics_html}</ul>"
            f"</div>",
            unsafe_allow_html=True,
        )

# ── Main table ─────────────────────────────────────────────────────────────────

if n_total == 0:
    st.info("No players match the current filters.")
    st.stop()

sort_col = "PDI" if highlight_dim == "PDI (overall)" else highlight_dim

bio_cols    = [c for c in ["Player", "Team", "_League", "Position", "PositionGroup", "Age",
                            "Minutes played", "Market value", "Contract expires", "Foot",
                            "Birth country"] if c in df.columns]
dim_cols    = ["PDI"] + list(DIMENSIONS.keys())
show_cols   = bio_cols + [c for c in dim_cols if c in df.columns]

positions_present = [p for p in POSITION_ORDER if p in df[pos_col].values]
tab_labels  = ["All"] + positions_present
tabs        = st.tabs(tab_labels)

for tab, label in zip(tabs, tab_labels):
    with tab:
        sub = df if label == "All" else df.loc[df[pos_col] == label]
        sub = sub.sort_values(sort_col, ascending=False)
        out_cols = [c for c in show_cols if c in sub.columns]
        display  = sub[out_cols].copy().reset_index(drop=True)

        for dc in dim_cols:
            if dc in display.columns:
                display[dc] = pd.to_numeric(display[dc], errors="coerce").round(1)

        st.dataframe(
            display,
            use_container_width=True,
            height=min(640, 56 + len(display) * 36),
            column_config={
                dc: st.column_config.ProgressColumn(
                    label=dc, min_value=0, max_value=100, format="%.1f",
                )
                for dc in dim_cols if dc in display.columns
            },
        )
        st.caption(f"{len(display)} players · sorted by {sort_col}")

# ── Top 10 leaderboard per dimension ──────────────────────────────────────────

st.markdown("---")
st.markdown("### Top 10 by Dimension")

dim_tabs = st.tabs(list(DIMENSIONS.keys()) + ["PDI"])
for tab, dim in zip(dim_tabs, list(DIMENSIONS.keys()) + ["PDI"]):
    with tab:
        if dim not in df.columns:
            st.info("No data.")
            continue
        top10 = df.nlargest(10, dim)[
            [c for c in ["Player", "Team", "_League", "PositionGroup", "Age", dim] + list(DIMENSIONS.keys()) if c in df.columns]
        ].reset_index(drop=True)
        top10.index = top10.index + 1
        col_cfg = {
            d: st.column_config.ProgressColumn(label=d, min_value=0, max_value=100, format="%.1f")
            for d in (list(DIMENSIONS.keys()) + ["PDI"]) if d in top10.columns
        }
        st.dataframe(top10, use_container_width=True, column_config=col_cfg)

# ── Export ─────────────────────────────────────────────────────────────────────

st.markdown("---")
exp_col, note_col = st.columns([2, 4])
with exp_col:
    excel_bytes = build_excel(df)
    st.download_button(
        "⬇️  Download Excel (.xlsx)",
        data=excel_bytes,
        file_name="FCHK_Pass_Danger_Index.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with note_col:
    st.markdown(
        "<div style='font-size:.8rem;color:#888;padding-top:8px;'>"
        "Excel contains: <b>All Players</b> sheet + one sheet per position group, "
        "colour-coded PDI and dimension scores with all raw passing metrics."
        "</div>",
        unsafe_allow_html=True,
    )
