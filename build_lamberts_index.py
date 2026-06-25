"""
build_lamberts_index.py
────────────────────────────────────────────────────────────────────
Lamberts Index — Full Player Intelligence Report (All Positions)
Positions: GK · CB · FB · DM · CM · W · FW
Output:    data/Lamberts_Index_Full_Report.xlsx

Scores:
  LI Score      (0–100)  Position-specific SQS percentile
  Lamberts Index         LI Score Rank − Market Value Rank
  DIS / ADI / PADS / PDS  Defensive sub-indexes (def positions only)

Sheets produced:
  README · Priority List · Elite Picks · Lamberts Analysis ·
  Top 5 · All Players · GK · CB · FB · DM · CM · W · FW ·
  League Analysis · Age Bands · Physical Profiles ·
  Expiring 2026 · Budget Planner · Squad

Usage:
  python build_lamberts_index.py
  python build_lamberts_index.py --leagues Czech Slovakia --min-minutes 500
  python build_lamberts_index.py --output data/My_Report.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config ─────────────────────────────────────────────────────────────────────

ROOT        = Path(__file__).parent
WYSCOUT_DIR = ROOT / "Wyscout Files"

SKIP_FILES = {
    "FCHK Model V3 - Loaded Leagues", "FCHK Model V3 - Model Input",
    "FCHK Model V3 - Player Scores",  "FCHK Model V3 - Player Styles",
    "FCHK Model V3 - Recruitment Scores", "FCHK Model V3 - Smart Club Closeness",
    "FCHK Model V3 - Summary", "FCHK Model V3 Scores", "FCHK Scouting Report",
    "Leagues Overview", "Wyscout Anomaly Report", "Wyscout Full Scouting Report",
    "FCHK Model V3 Scores",
}

DEFAULT_MIN_MINUTES = 400
DEFAULT_MAX_AGE     = 30
DEFAULT_BUDGET      = 1_000_000

ALL_POSITIONS       = {"GK", "CB", "FB", "DM", "CM", "W", "FW"}
DEFENSIVE_POSITIONS = {"GK", "CB", "FB", "DM"}

POS_MAP: dict[str, str] = {
    "CF": "FW", "SS": "FW",
    "LW": "W",  "RW": "W",  "LWF": "W", "RWF": "W", "WF": "W",
    "LAMF": "W", "RAMF": "W",
    "AMF": "CM",
    "CMF": "CM", "LCM": "CM", "RCM": "CM", "LCMF": "CM", "RCMF": "CM",
    "DMF": "DM", "LDM": "DM", "RDM": "DM", "LDMF": "DM", "RDMF": "DM",
    "LB": "FB",  "RB": "FB",  "LWB": "FB", "RWB": "FB",
    "CB": "CB",  "LCB": "CB", "RCB": "CB",
    "GK": "GK",
}

SQS_BLUEPRINTS: dict[str, list[tuple[str, float]]] = {
    "GK": [
        ("Save rate, %",                      4.0),
        ("Prevented goals per 90",            3.0),
        ("Exits per 90",                      2.0),
        ("Aerial duels per 90",               1.5),
        ("Accurate passes, %",                1.5),
        ("Accurate long passes, %",           1.0),
    ],
    "CB": [
        ("Successful defensive actions per 90", 3.0),
        ("Defensive duels won, %",              2.5),
        ("Aerial duels won, %",                 2.0),
        ("Interceptions per 90",                2.0),
        ("PAdj Interceptions",                  1.5),
        ("Shots blocked per 90",                1.0),
        ("Accurate passes, %",                  1.0),
        ("Progressive passes per 90",           1.0),
    ],
    "FB": [
        ("Crosses per 90",                      2.0),
        ("Accurate crosses, %",                 1.5),
        ("xA per 90",                           2.0),
        ("Assists per 90",                      1.5),
        ("Progressive runs per 90",             1.5),
        ("Dribbles per 90",                     1.0),
        ("Successful defensive actions per 90", 2.0),
        ("Defensive duels won, %",              2.0),
        ("Aerial duels won, %",                 1.0),
        ("Progressive passes per 90",           1.5),
    ],
    "DM": [
        ("Successful defensive actions per 90", 3.0),
        ("Defensive duels won, %",              2.5),
        ("Interceptions per 90",                2.5),
        ("PAdj Interceptions",                  2.0),
        ("Aerial duels won, %",                 1.5),
        ("Passes per 90",                       1.5),
        ("Accurate passes, %",                  1.5),
        ("Progressive passes per 90",           1.5),
    ],
    "CM": [
        ("Passes per 90",                       2.0),
        ("Accurate passes, %",                  2.0),
        ("Progressive passes per 90",           2.5),
        ("Key passes per 90",                   2.5),
        ("xA per 90",                           2.0),
        ("Assists per 90",                      1.5),
        ("Progressive runs per 90",             1.5),
        ("Successful defensive actions per 90", 1.5),
        ("Goals per 90",                        1.5),
    ],
    "W": [
        ("Goals per 90",                        2.5),
        ("xG per 90",                           2.0),
        ("Assists per 90",                      2.0),
        ("xA per 90",                           2.0),
        ("Dribbles per 90",                     2.0),
        ("Successful dribbles, %",              1.5),
        ("Progressive runs per 90",             2.0),
        ("Touches in box per 90",               2.0),
        ("Key passes per 90",                   1.5),
    ],
    "FW": [
        ("Goals per 90",                        4.0),
        ("xG per 90",                           3.0),
        ("Non-penalty goals per 90",            2.5),
        ("Shots per 90",                        1.5),
        ("Shots on target, %",                  1.5),
        ("Goal conversion, %",                  1.5),
        ("Touches in box per 90",               1.5),
        ("Aerial duels won, %",                 1.5),
        ("Dribbles per 90",                     1.0),
        ("xA per 90",                           1.0),
    ],
}

HRADEC_SQUAD: dict[str, float] = {
    "GK": 61.5,
    "CB": 33.3,
    "FB": 81.8,
    "DM": 36.8,
    "CM": 53.7,
    "W":  27.6,
    "FW": 38.3,
}

POS_LABELS = {
    "GK": "GK — GOALKEEPER",
    "CB": "CB — CENTRE-BACK",
    "FB": "FB — FULL-BACK",
    "DM": "DM — DEFENSIVE MID",
    "CM": "CM — CENTRAL MID",
    "W":  "W — WINGER",
    "FW": "FW — FORWARD",
}

# ── Colors ─────────────────────────────────────────────────────────────────────

C = {
    "navy":   "0D1B2A",  "gold":   "C9A84C",  "steel":  "1B4F72",
    "top5":   "6D3460",  "elite":  "1A5276",  "high":   "1E8449",   "value":  "117A65",
    "fair":   "626567",  "over":   "922B21",   "white":  "FFFFFF",
    "ice":    "EBF5FB",  "mint":   "E8F8F5",   "pearl":  "F8F9FA",
    "header": "154360",  "sub":    "1A5276",   "dark":   "0D1B2A",
    "upgrade":"1A5276",  "rota":   "7D6608",   "depth":  "117A65",
    "li_pos": "1E8449",  "li_neg": "922B21",   "gold2":  "F0E68C",
    "teal":   "0B6E4F",  "purple": "6C3483",   "orange": "A04000",
}

# Per-position accent colors for sheet headers
POS_COLOR = {
    "GK": "154360", "CB": "0B5345", "FB": "145A32",
    "DM": "1A5276", "CM": "6C3483", "W":  "784212", "FW": "922B21",
}

# ── Wyscout column names (source) ──────────────────────────────────────────────

W = {
    # Defence
    "def_act":    "Successful defensive actions per 90",
    "def_duel":   "Defensive duels per 90",
    "def_duel_w": "Defensive duels won, %",
    "aerial":     "Aerial duels per 90",
    "aerial_w":   "Aerial duels won, %",
    "slides":     "Sliding tackles per 90",
    "padj_sl":    "PAdj Sliding tackles",
    "blocks":     "Shots blocked per 90",
    "intercept":  "Interceptions per 90",
    "padj_int":   "PAdj Interceptions",
    "fouls":      "Fouls per 90",
    "yellow":     "Yellow cards per 90",
    "red":        "Red cards per 90",
    "duels":      "Duels per 90",
    "duels_w":    "Duels won, %",
    # GK
    "save_pct":   "Save rate, %",
    "xg_against": "xG against per 90",
    "prevent_gls":"Prevented goals per 90",
    "conceded":   "Conceded goals per 90",
    "clean":      "Clean sheets",
    "exits":      "Exits per 90",
    "back_recv":  "Back passes received as GK per 90",
    "gk_aerial":  "Aerial duels per 90.1",
    # Attack
    "goals":      "Goals per 90",
    "xg":         "xG per 90",
    "np_goals":   "Non-penalty goals per 90",
    "head_gls":   "Head goals per 90",
    "shots":      "Shots per 90",
    "shot_tgt":   "Shots on target, %",
    "goal_conv":  "Goal conversion, %",
    "assists":    "Assists per 90",
    "xa":         "xA per 90",
    "key_pass":   "Key passes per 90",
    "shot_ass":   "Shot assists per 90",
    "touch_box":  "Touches in box per 90",
    "off_duel":   "Offensive duels per 90",
    "off_duel_w": "Offensive duels won, %",
    "foul_suff":  "Fouls suffered per 90",
    "att_act":    "Successful attacking actions per 90",
    # Passing
    "passes":     "Passes per 90",
    "acc_pass":   "Accurate passes, %",
    "prog_pass":  "Progressive passes per 90",
    "acc_prog":   "Accurate progressive passes, %",
    "fwd_pass":   "Forward passes per 90",
    "acc_fwd":    "Accurate forward passes, %",
    "back_pass":  "Back passes per 90",
    "lat_pass":   "Lateral passes per 90",
    "long_pass":  "Long passes per 90",
    "acc_long":   "Accurate long passes, %",
    "smart_pass": "Smart passes per 90",
    "acc_smart":  "Accurate smart passes, %",
    "pass_f3":    "Passes to final third per 90",
    "acc_pf3":    "Accurate passes to final third, %",
    "pass_box":   "Passes to penalty area per 90",
    "acc_pbox":   "Accurate passes to penalty area, %",
    "thru_pass":  "Through passes per 90",
    "acc_thru":   "Accurate through passes, %",
    "deep_compl": "Deep completions per 90",
    "sec_ass":    "Second assists per 90",
    "thr_ass":    "Third assists per 90",
    "avg_plen":   "Average pass length, m",
    "avg_lplen":  "Average long pass length, m",
    # Carrying
    "dribbles":   "Dribbles per 90",
    "succ_drib":  "Successful dribbles, %",
    "prog_run":   "Progressive runs per 90",
    "accel":      "Accelerations per 90",
    "recv_pass":  "Received passes per 90",
    "recv_long":  "Received long passes per 90",
    # Wide
    "crosses":    "Crosses per 90",
    "acc_cross":  "Accurate crosses, %",
    "lcross":     "Crosses from left flank per 90",
    "acc_lcross": "Accurate crosses from left flank, %",
    "rcross":     "Crosses from right flank per 90",
    "acc_rcross": "Accurate crosses from right flank, %",
    "cross_box":  "Crosses to goalie box per 90",
    "deep_cross": "Deep completed crosses per 90",
}

# ── Output column blueprints per position ──────────────────────────────────────

# Shared identity block
_ID = ["Player", "Team", "League", "Birth Country", "Pos", "Full Position",
       "Age", "Height (cm)", "Weight (kg)", "Foot", "On Loan",
       "Contract", "Exp?", "Mkt Val (€)", "Model Val (€)", "Val Ratio"]
_SCORE = ["Tier", "LI Score", "Lamberts Index"]
_DEF   = ["DIS", "ADI", "PADS", "PDS"]
_SITE  = ["Status", "vs Hradec", "Minutes", "Matches", "Min/Match"]

POS_COLS: dict[str, list[str]] = {
    "ALL": _ID + _SCORE + _DEF + _SITE + [
        # Defensive
        "Def Actions/90", "Def Duels/90", "Def Duel Won %",
        "Aerial/90", "Aerial Won %",
        "Intercept/90", "PAdj Intercept",
        "Slides/90", "PAdj Slides", "Blocks/90",
        "Fouls/90", "Yellow/90", "Red/90",
        "Duels/90", "Duels Won %",
        # GK only
        "Save %", "xG Against/90", "Prevent Gls/90", "Conceded/90", "Clean Sheets", "Exits/90",
        # Attack
        "Goals/90", "xG/90", "NP Goals/90", "Head Gls/90",
        "Shots/90", "Shot On Tgt %", "Goal Conv %",
        "Assists/90", "xA/90", "Key Pass/90", "Shot Assists/90",
        "Touch Box/90", "Off Duels/90", "Off Duel Won %", "Fouls Suf/90",
        # Passing
        "Passes/90", "Acc Pass %",
        "Prog Pass/90", "Acc Prog Pass %",
        "Fwd Pass/90", "Acc Fwd Pass %",
        "Long Pass/90", "Acc Long Pass %",
        "Smart Pass/90", "Acc Smart Pass %",
        "Pass F3rd/90", "Acc Pass F3rd %",
        "Pass Box/90", "Acc Pass Box %",
        "Thru Pass/90", "Acc Thru Pass %",
        "Deep Compl/90", "2nd Assist/90", "3rd Assist/90",
        "Avg Pass Len (m)", "Avg Long Pass (m)",
        # Carrying
        "Dribbles/90", "Succ Drib %",
        "Prog Runs/90", "Accelerations/90",
        "Recv Pass/90", "Recv Long Pass/90",
        # Wide
        "Crosses/90", "Acc Cross %",
        "L Cross/90", "Acc L Cross %",
        "R Cross/90", "Acc R Cross %",
        "Cross to Box/90", "Deep Cross/90",
    ],
    "GK": _ID + _SCORE + _SITE + [
        "Save %", "xG Against/90", "Prevent Gls/90", "Conceded/90",
        "Clean Sheets", "Exits/90",
        "Aerial/90", "Aerial Won %",
        "Back Pass Recv/90",
        "Passes/90", "Acc Pass %",
        "Long Pass/90", "Acc Long Pass %",
        "Avg Pass Len (m)", "Avg Long Pass (m)",
        "Def Actions/90", "Def Duels/90", "Def Duel Won %",
        "Fouls/90", "Yellow/90",
        "Duels/90", "Duels Won %",
    ],
    "CB": _ID + _SCORE + _DEF + _SITE + [
        "Def Actions/90", "Def Duels/90", "Def Duel Won %",
        "Aerial/90", "Aerial Won %",
        "Intercept/90", "PAdj Intercept",
        "Slides/90", "PAdj Slides", "Blocks/90",
        "Fouls/90", "Yellow/90", "Red/90",
        "Duels/90", "Duels Won %",
        "Head Gls/90",
        "Prog Pass/90", "Acc Prog Pass %",
        "Acc Pass %", "Fwd Pass/90", "Acc Fwd Pass %",
        "Long Pass/90", "Acc Long Pass %",
        "Pass F3rd/90", "Acc Pass F3rd %",
        "Deep Compl/90",
        "Avg Pass Len (m)", "Avg Long Pass (m)",
        "Prog Runs/90", "Recv Pass/90", "Recv Long Pass/90",
    ],
    "FB": _ID + _SCORE + ["DIS", "ADI", "PDS"] + _SITE + [
        "Def Actions/90", "Def Duels/90", "Def Duel Won %",
        "Aerial/90", "Aerial Won %",
        "Intercept/90", "PAdj Intercept",
        "Fouls/90", "Yellow/90",
        "Crosses/90", "Acc Cross %",
        "L Cross/90", "Acc L Cross %",
        "R Cross/90", "Acc R Cross %",
        "Cross to Box/90", "Deep Cross/90",
        "xA/90", "Assists/90", "Key Pass/90",
        "Touch Box/90", "Shot Assists/90",
        "Prog Pass/90", "Acc Prog Pass %",
        "Acc Pass %", "Fwd Pass/90", "Pass F3rd/90", "Acc Pass F3rd %",
        "Dribbles/90", "Succ Drib %",
        "Prog Runs/90", "Accelerations/90",
        "Fouls Suf/90", "Off Duels/90", "Off Duel Won %",
        "Recv Pass/90",
    ],
    "DM": _ID + _SCORE + _DEF + _SITE + [
        "Def Actions/90", "Def Duels/90", "Def Duel Won %",
        "Intercept/90", "PAdj Intercept",
        "Slides/90", "PAdj Slides", "Blocks/90",
        "Aerial/90", "Aerial Won %",
        "Fouls/90", "Yellow/90",
        "Duels/90", "Duels Won %",
        "Passes/90", "Acc Pass %",
        "Prog Pass/90", "Acc Prog Pass %",
        "Fwd Pass/90", "Acc Fwd Pass %",
        "Long Pass/90", "Acc Long Pass %",
        "Pass F3rd/90", "Acc Pass F3rd %",
        "Smart Pass/90", "Acc Smart Pass %",
        "Thru Pass/90", "Acc Thru Pass %",
        "Key Pass/90", "xA/90", "Assists/90",
        "Goals/90", "xG/90",
        "Avg Pass Len (m)", "Avg Long Pass (m)",
        "Recv Pass/90",
    ],
    "CM": _ID + _SCORE + _SITE + [
        "Passes/90", "Acc Pass %",
        "Prog Pass/90", "Acc Prog Pass %",
        "Fwd Pass/90", "Acc Fwd Pass %",
        "Long Pass/90", "Acc Long Pass %",
        "Smart Pass/90", "Acc Smart Pass %",
        "Pass F3rd/90", "Acc Pass F3rd %",
        "Pass Box/90", "Acc Pass Box %",
        "Thru Pass/90", "Acc Thru Pass %",
        "Key Pass/90", "xA/90", "Assists/90",
        "Shot Assists/90", "2nd Assist/90", "3rd Assist/90",
        "Deep Compl/90",
        "Goals/90", "xG/90", "NP Goals/90",
        "Shots/90", "Shot On Tgt %",
        "Touch Box/90", "Prog Runs/90", "Dribbles/90", "Succ Drib %",
        "Accelerations/90", "Fouls Suf/90",
        "Def Actions/90", "Intercept/90", "Fouls/90",
        "Avg Pass Len (m)",
        "Recv Pass/90",
    ],
    "W": _ID + _SCORE + _SITE + [
        "Goals/90", "xG/90", "NP Goals/90",
        "Shots/90", "Shot On Tgt %", "Goal Conv %",
        "Assists/90", "xA/90", "Key Pass/90", "Shot Assists/90",
        "2nd Assist/90", "3rd Assist/90",
        "Crosses/90", "Acc Cross %",
        "L Cross/90", "Acc L Cross %",
        "R Cross/90", "Acc R Cross %",
        "Cross to Box/90", "Deep Cross/90",
        "Dribbles/90", "Succ Drib %",
        "Prog Runs/90", "Accelerations/90",
        "Touch Box/90", "Fouls Suf/90",
        "Off Duels/90", "Off Duel Won %",
        "Fwd Pass/90", "Acc Fwd Pass %",
        "Prog Pass/90", "Pass F3rd/90",
        "Def Actions/90",
    ],
    "FW": _ID + _SCORE + _SITE + [
        "Goals/90", "xG/90", "NP Goals/90", "Head Gls/90",
        "Shots/90", "Shot On Tgt %", "Goal Conv %",
        "Assists/90", "xA/90", "Key Pass/90",
        "Shot Assists/90", "2nd Assist/90", "3rd Assist/90",
        "Touch Box/90", "Pass Box/90", "Acc Pass Box %",
        "Dribbles/90", "Succ Drib %",
        "Prog Runs/90", "Accelerations/90",
        "Aerial/90", "Aerial Won %",
        "Fouls Suf/90", "Off Duels/90", "Off Duel Won %",
        "Def Actions/90",
    ],
}

# Raw stat alias → Wyscout column name (for ALL sheet and position sheets)
STAT_ALIAS: dict[str, str] = {
    "Def Actions/90":     "Successful defensive actions per 90",
    "Def Duels/90":       "Defensive duels per 90",
    "Def Duel Won %":     "Defensive duels won, %",
    "Aerial/90":          "Aerial duels per 90",
    "Aerial Won %":       "Aerial duels won, %",
    "Slides/90":          "Sliding tackles per 90",
    "PAdj Slides":        "PAdj Sliding tackles",
    "Blocks/90":          "Shots blocked per 90",
    "Intercept/90":       "Interceptions per 90",
    "PAdj Intercept":     "PAdj Interceptions",
    "Fouls/90":           "Fouls per 90",
    "Yellow/90":          "Yellow cards per 90",
    "Red/90":             "Red cards per 90",
    "Duels/90":           "Duels per 90",
    "Duels Won %":        "Duels won, %",
    "Save %":             "Save rate, %",
    "xG Against/90":      "xG against per 90",
    "Prevent Gls/90":     "Prevented goals per 90",
    "Conceded/90":        "Conceded goals per 90",
    "Clean Sheets":       "Clean sheets",
    "Exits/90":           "Exits per 90",
    "Back Pass Recv/90":  "Back passes received as GK per 90",
    "Goals/90":           "Goals per 90",
    "xG/90":              "xG per 90",
    "NP Goals/90":        "Non-penalty goals per 90",
    "Head Gls/90":        "Head goals per 90",
    "Shots/90":           "Shots per 90",
    "Shot On Tgt %":      "Shots on target, %",
    "Goal Conv %":        "Goal conversion, %",
    "Assists/90":         "Assists per 90",
    "xA/90":              "xA per 90",
    "Key Pass/90":        "Key passes per 90",
    "Shot Assists/90":    "Shot assists per 90",
    "Touch Box/90":       "Touches in box per 90",
    "Off Duels/90":       "Offensive duels per 90",
    "Off Duel Won %":     "Offensive duels won, %",
    "Fouls Suf/90":       "Fouls suffered per 90",
    "Passes/90":          "Passes per 90",
    "Acc Pass %":         "Accurate passes, %",
    "Prog Pass/90":       "Progressive passes per 90",
    "Acc Prog Pass %":    "Accurate progressive passes, %",
    "Fwd Pass/90":        "Forward passes per 90",
    "Acc Fwd Pass %":     "Accurate forward passes, %",
    "Back Pass/90":       "Back passes per 90",
    "Lat Pass/90":        "Lateral passes per 90",
    "Long Pass/90":       "Long passes per 90",
    "Acc Long Pass %":    "Accurate long passes, %",
    "Smart Pass/90":      "Smart passes per 90",
    "Acc Smart Pass %":   "Accurate smart passes, %",
    "Pass F3rd/90":       "Passes to final third per 90",
    "Acc Pass F3rd %":    "Accurate passes to final third, %",
    "Pass Box/90":        "Passes to penalty area per 90",
    "Acc Pass Box %":     "Accurate passes to penalty area, %",
    "Thru Pass/90":       "Through passes per 90",
    "Acc Thru Pass %":    "Accurate through passes, %",
    "Deep Compl/90":      "Deep completions per 90",
    "2nd Assist/90":      "Second assists per 90",
    "3rd Assist/90":      "Third assists per 90",
    "Avg Pass Len (m)":   "Average pass length, m",
    "Avg Long Pass (m)":  "Average long pass length, m",
    "Dribbles/90":        "Dribbles per 90",
    "Succ Drib %":        "Successful dribbles, %",
    "Prog Runs/90":       "Progressive runs per 90",
    "Accelerations/90":   "Accelerations per 90",
    "Recv Pass/90":       "Received passes per 90",
    "Recv Long Pass/90":  "Received long passes per 90",
    "Crosses/90":         "Crosses per 90",
    "Acc Cross %":        "Accurate crosses, %",
    "L Cross/90":         "Crosses from left flank per 90",
    "Acc L Cross %":      "Accurate crosses from left flank, %",
    "R Cross/90":         "Crosses from right flank per 90",
    "Acc R Cross %":      "Accurate crosses from right flank, %",
    "Cross to Box/90":    "Crosses to goalie box per 90",
    "Deep Cross/90":      "Deep completed crosses per 90",
}

# ── Data loading ───────────────────────────────────────────────────────────────

def load_leagues(leagues: list[str] | None, min_minutes: int) -> pd.DataFrame:
    if leagues is None:
        paths = sorted(p for p in WYSCOUT_DIR.glob("*.xlsx") if p.stem not in SKIP_FILES)
    else:
        paths = [WYSCOUT_DIR / f"{lg}.xlsx" for lg in leagues]

    frames: list[pd.DataFrame] = []
    for path in paths:
        if not path.exists():
            print(f"  [warn] {path} not found — skipping")
            continue
        try:
            df = pd.read_excel(path)
        except Exception as e:
            print(f"  [warn] Could not read {path.name}: {e}")
            continue
        df = df.copy()
        df["_League"] = path.stem
        frames.append(df)
        print(f"  Loaded {path.stem}: {len(df)} rows")

    if not frames:
        raise RuntimeError(f"No Wyscout files found in {WYSCOUT_DIR}")

    raw = pd.concat(frames, ignore_index=True)

    mins_col = next((c for c in ["Minutes played", "MinutesPlayed", "Minutes"] if c in raw.columns), None)
    raw["_minutes"] = pd.to_numeric(raw[mins_col], errors="coerce").fillna(0) if mins_col else 0.0
    raw = raw[raw["_minutes"] >= min_minutes].copy()
    print(f"  → {len(raw)} players after {min_minutes}+ min filter")
    return raw.reset_index(drop=True)


def map_position(pos_str: str) -> str:
    if not isinstance(pos_str, str):
        return "Other"
    first = pos_str.split(",")[0].strip()
    return POS_MAP.get(first, "Other")


def add_position_group(df: pd.DataFrame) -> pd.DataFrame:
    pos_col = next((c for c in ["Position", "Pos"] if c in df.columns), None)
    if pos_col:
        df["_pos_group"]     = df[pos_col].apply(map_position)
        df["_full_position"] = df[pos_col].fillna("Unknown")
    else:
        df["_pos_group"]     = "Other"
        df["_full_position"] = "Unknown"
    return df


# ── Index computation ──────────────────────────────────────────────────────────

def _pct_rank(series: pd.Series, ascending: bool = True) -> pd.Series:
    return series.rank(pct=True, ascending=ascending) * 100


def _get(df: pd.DataFrame, col: str) -> pd.Series:
    if col in df.columns:
        return pd.to_numeric(df[col], errors="coerce").fillna(0)
    return pd.Series(0.0, index=df.index)


def _wt(components: list[tuple[pd.Series, float]]) -> pd.Series:
    total_w = sum(w for _, w in components)
    return sum(s * w for s, w in components) / total_w if total_w else pd.Series(0.0)


def compute_sqs(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_SQS_raw"] = np.nan
    df["_LI_score"] = np.nan

    for pos, blueprint in SQS_BLUEPRINTS.items():
        mask = df["_pos_group"] == pos
        if not mask.any():
            continue
        grp = df.loc[mask].copy()
        comps = [(_pct_rank(_get(grp, col)), w) for col, w in blueprint]
        df.loc[mask, "_SQS_raw"] = _wt(comps).values

    for pos in SQS_BLUEPRINTS:
        mask = df["_pos_group"] == pos
        if not mask.any():
            continue
        df.loc[mask, "_LI_score"] = (
            df.loc[mask, "_SQS_raw"].rank(pct=True) * 100
        ).round(2)

    return df


def compute_defensive_indexes(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in ["_DIS", "_ADI", "_PADS", "_PDS"]:
        df[col] = np.nan

    for pos in DEFENSIVE_POSITIONS:
        mask = df["_pos_group"] == pos
        if not mask.any():
            continue
        grp = df.loc[mask].copy()

        dis = _wt([
            (_pct_rank(_get(grp, "Successful defensive actions per 90")), 3.0),
            (_pct_rank(_get(grp, "Defensive duels won, %")),             2.5),
            (_pct_rank(_get(grp, "Interceptions per 90")),               2.0),
            (_pct_rank(_get(grp, "PAdj Interceptions")),                 2.0),
            (_pct_rank(_get(grp, "Aerial duels won, %")),                1.5),
            (_pct_rank(_get(grp, "Shots blocked per 90")),               1.0),
        ]).round(2)

        adi = _pct_rank(
            _get(grp, "Aerial duels per 90") *
            _get(grp, "Aerial duels won, %") / 100.0
        ).round(2)

        pads = _wt([
            (_pct_rank(_get(grp, "PAdj Interceptions")),     2.5),
            (_pct_rank(_get(grp, "PAdj Sliding tackles")),   2.5),
            (_pct_rank(_get(grp, "Shots blocked per 90")),   1.5),
            (_pct_rank(_get(grp, "Defensive duels per 90")), 1.0),
        ]).round(2)

        cost = (_get(grp, "Fouls per 90") * 0.6 +
                _get(grp, "Yellow cards per 90") * 0.3 +
                _get(grp, "Red cards per 90") * 0.1)
        pds = _pct_rank(cost, ascending=False).round(2)

        df.loc[mask, "_DIS"]  = dis.values
        df.loc[mask, "_ADI"]  = adi.values
        df.loc[mask, "_PADS"] = pads.values
        df.loc[mask, "_PDS"]  = pds.values

    return df


def compute_mv_rank(df: pd.DataFrame) -> pd.DataFrame:
    mv_col = next((c for c in ["Market value", "MarketValue"] if c in df.columns), None)
    mv = pd.to_numeric(df[mv_col], errors="coerce").fillna(0) if mv_col else pd.Series(0.0, index=df.index)
    df["_mkt_val"] = mv
    df["_mv_rank"]  = mv.rank(pct=True) * 100
    return df


def compute_lamberts_index(df: pd.DataFrame) -> pd.DataFrame:
    df["_LI"] = (df["_LI_score"] - df["_mv_rank"]).round(2)

    def tier(v: float) -> str:
        if v >= 30: return "ELITE VALUE"
        if v >= 20: return "HIGH VALUE"
        if v >= 10: return "VALUE"
        if v >= 0:  return "FAIR VALUE"
        return "OVERPRICED"

    df["_tier"] = df["_LI"].apply(tier)

    # Promote top 5 per position group to TOP 5 tier
    for pos_grp, grp_idx in df.groupby("_pos_group").groups.items():
        grp = df.loc[grp_idx].sort_values("_LI", ascending=False)
        top5_idx = grp.head(5).index
        df.loc[top5_idx, "_tier"] = "TOP 5"

    return df


def compute_vs_hradec(df: pd.DataFrame) -> pd.DataFrame:
    vs_col, status_col = [], []
    for _, row in df.iterrows():
        pos    = row.get("_pos_group", "Other")
        hradec = HRADEC_SQUAD.get(pos, 50.0)
        li     = float(row.get("_LI_score", 50.0) or 50.0)
        gap    = round(li - hradec, 1)
        vs_col.append(gap)
        if gap > 0:      status_col.append("CLEAR UPGRADE")
        elif gap > -10:  status_col.append("ROTATIONAL / COVER")
        else:            status_col.append("DEPTH")
    df["_vs_hradec"] = vs_col
    df["_status"]    = status_col
    return df


def age_band(age) -> str:
    try:
        a = int(age)
        if a <= 21: return "U21"
        if a <= 23: return "U23"
        if a <= 25: return "U25"
        if a <= 28: return "U28"
        return "U30+"
    except Exception:
        return "Unknown"


def model_value(mkt: float, li_score: float) -> float:
    return round(mkt * max(li_score / 50.0, 0.1), 0)


def val_ratio_str(mkt: float, model: float) -> str:
    return "N/A" if mkt <= 0 else f"{model / mkt:.1f}×"


# ── Master table ───────────────────────────────────────────────────────────────

def build_master(df: pd.DataFrame) -> pd.DataFrame:
    contract_col = next(
        (c for c in ["Contract expires", "ContractExpires"] if c in df.columns), None
    )
    matches_col  = next(
        (c for c in ["Matches played", "MatchesPlayed"] if c in df.columns), None
    )
    rows: list[dict] = []

    for _, r in df.iterrows():
        mkt   = float(r.get("_mkt_val", 0) or 0)
        li_s  = float(r.get("_LI_score", 0) or 0)
        mod_v = model_value(mkt, li_s)

        contract = r.get(contract_col) if contract_col else None
        if pd.notna(contract):
            try:   contract = pd.to_datetime(contract).strftime("%Y-%m-%d")
            except Exception: contract = str(contract)
        else:
            contract = None

        exp_year = None
        if contract:
            try:   exp_year = str(pd.to_datetime(contract).year)
            except Exception: pass

        def _f(col: str, default=0.0):
            v = r.get(col)
            try:   return round(float(v), 2) if pd.notna(v) else default
            except Exception: return default

        def _i(col: str, default=0):
            v = r.get(col)
            try:   return int(float(v)) if pd.notna(v) else default
            except Exception: return default

        mins   = _i("_minutes")
        matches = _i(matches_col) if matches_col else 0
        mpm    = round(mins / matches, 0) if matches > 0 else 0

        row: dict = {
            "Player":         r.get("Player", ""),
            "Team":           r.get("Team", ""),
            "League":         r.get("_League", ""),
            "Birth Country":  r.get("Birth country", ""),
            "Pos":            r.get("_pos_group", ""),
            "Full Position":  r.get("_full_position", ""),
            "Age":            _i("Age"),
            "Age Band":       age_band(r.get("Age")),
            "Height (cm)":    _i("Height"),
            "Weight (kg)":    _i("Weight"),
            "Foot":           r.get("Foot", ""),
            "On Loan":        r.get("On loan", ""),
            "Contract":       contract,
            "Exp?":           exp_year,
            "Mkt Val (€)":    int(mkt) if mkt > 0 else 0,
            "Model Val (€)":  int(mod_v),
            "Val Ratio":      val_ratio_str(mkt, mod_v),
            "Tier":           r.get("_tier", ""),
            "LI Score":       round(li_s, 2),
            "Lamberts Index": round(float(r.get("_LI", 0) or 0), 2),
            "DIS":            _f("_DIS"),
            "ADI":            _f("_ADI"),
            "PADS":           _f("_PADS"),
            "PDS":            _f("_PDS"),
            "Status":         r.get("_status", ""),
            "vs Hradec":      r.get("_vs_hradec", 0),
            "Minutes":        mins,
            "Matches":        matches,
            "Min/Match":      int(mpm),
        }
        # All stat aliases
        for alias, src in STAT_ALIAS.items():
            row[alias] = _f(src)

        rows.append(row)

    master = pd.DataFrame(rows)
    master = master.sort_values("Lamberts Index", ascending=False).reset_index(drop=True)
    return master


# ── openpyxl helpers ──────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border(color="CCCCCC") -> Border:
    thin = Side(style="thin", color=color)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        try:
            max_len = max(
                len(str(col_cells[0].value or "")),
                *(len(str(c.value or "")) for c in col_cells[1:15]),
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)
        except Exception:
            pass


# ── xlsxwriter helpers ────────────────────────────────────────────────────────

def _xf(wb_x, bold=False, font_color="#000000", bg_color=None,
         size=9, italic=False, align="center", wrap=False, border=True, num_fmt=None):
    props: dict = {
        "font_size": size, "bold": bold, "italic": italic,
        "font_color": font_color, "align": align, "valign": "vcenter",
    }
    if wrap:    props["text_wrap"] = True
    if bg_color: props["bg_color"] = bg_color; props["pattern"] = 1
    if border:  props.update({"border": 1, "border_color": "#CCCCCC"})
    if num_fmt: props["num_format"] = num_fmt
    return wb_x.add_format(props)


TIER_BG   = {
    "TOP 5":       f"#{C['top5']}",
    "ELITE VALUE": f"#{C['elite']}", "HIGH VALUE": f"#{C['high']}",
    "VALUE": f"#{C['value']}", "FAIR VALUE": f"#{C['fair']}",
    "OVERPRICED": f"#{C['over']}",
}
STATUS_BG = {
    "CLEAR UPGRADE": f"#{C['upgrade']}", "ROTATIONAL / COVER": f"#{C['rota']}",
    "DEPTH": f"#{C['depth']}",
}


def write_data_sheet(wb_x, sheet_name: str, title: str, subtitle: str,
                     df: pd.DataFrame, pos_accent: str | None = None) -> None:
    """Write a DataFrame to an xlsxwriter worksheet with full formatting + AutoFilter."""
    accent = pos_accent or C["dark"]

    if df.empty:
        ws = wb_x.add_worksheet(sheet_name)
        ws.write(0, 0, title)
        return

    ws = wb_x.add_worksheet(sheet_name)
    ws.set_zoom(85)
    ws.hide_gridlines(2)

    cols   = list(df.columns)
    n_cols = len(cols)

    # Row 1: title
    fmt_t = _xf(wb_x, bold=True, font_color="#FFFFFF", bg_color=f"#{accent}",
                 size=13, align="left", border=False)
    ws.merge_range(0, 0, 0, n_cols - 1, title, fmt_t)
    ws.set_row(0, 24)

    # Row 2: subtitle
    fmt_s = _xf(wb_x, italic=True, font_color=f"#{C['gold']}", bg_color=f"#{accent}",
                 size=9, align="left", border=False)
    ws.merge_range(1, 0, 1, n_cols - 1, subtitle, fmt_s)
    ws.set_row(1, 14)

    # Row 3: headers
    fmt_h = _xf(wb_x, bold=True, font_color="#FFFFFF",
                 bg_color=f"#{C['header']}", size=9, wrap=True)
    for j, col in enumerate(cols):
        ws.write(2, j, col, fmt_h)
    ws.set_row(2, 30)

    ws.freeze_panes(3, 0)
    ws.autofilter(2, 0, 2, n_cols - 1)

    # Format cache
    tier_idx   = cols.index("Tier")           if "Tier"           in cols else -1
    status_idx = cols.index("Status")         if "Status"         in cols else -1
    li_idx     = cols.index("Lamberts Index") if "Lamberts Index" in cols else -1
    lis_idx    = cols.index("LI Score")       if "LI Score"       in cols else -1

    fmt_e   = _xf(wb_x, bg_color=f"#{C['ice']}")
    fmt_o   = _xf(wb_x, bg_color=f"#{C['white']}")
    fmt_be  = _xf(wb_x, bold=True, bg_color=f"#{C['ice']}")
    fmt_bo  = _xf(wb_x, bold=True, bg_color=f"#{C['white']}")
    fmt_lpe = _xf(wb_x, bold=True, font_color=f"#{C['li_pos']}", bg_color=f"#{C['ice']}")
    fmt_lne = _xf(wb_x, bold=True, font_color=f"#{C['li_neg']}", bg_color=f"#{C['ice']}")
    fmt_lpo = _xf(wb_x, bold=True, font_color=f"#{C['li_pos']}", bg_color=f"#{C['white']}")
    fmt_lno = _xf(wb_x, bold=True, font_color=f"#{C['li_neg']}", bg_color=f"#{C['white']}")

    tier_fmts   = {k: _xf(wb_x, bold=True, font_color="#FFFFFF", bg_color=v)
                   for k, v in TIER_BG.items()}
    status_fmts = {k: _xf(wb_x, bold=True, font_color="#FFFFFF", bg_color=v)
                   for k, v in STATUS_BG.items()}

    for i, row_vals in enumerate(df.itertuples(index=False)):
        xrow = i + 3
        even = (i % 2 == 0)
        base = fmt_e if even else fmt_o
        bold = fmt_be if even else fmt_bo
        lp   = fmt_lpe if even else fmt_lpo
        ln   = fmt_lne if even else fmt_lno

        for j, val in enumerate(row_vals):
            if j == tier_idx:
                fmt = tier_fmts.get(str(val), base)
            elif j == status_idx:
                fmt = status_fmts.get(str(val), base)
            elif j == li_idx:
                try:   fmt = lp if float(val or 0) >= 0 else ln
                except Exception: fmt = base
            elif j == lis_idx:
                fmt = bold
            else:
                fmt = base

            if val is None or (isinstance(val, float) and np.isnan(val)):
                ws.write_blank(xrow, j, None, fmt)
            elif isinstance(val, (int, float)):
                ws.write_number(xrow, j, val, fmt)
            else:
                ws.write_string(xrow, j, str(val), fmt)

    # Column widths
    WIDE = {"Player", "Team", "League", "Birth Country", "Full Position"}
    NARR = {"Foot", "On Loan", "Exp?", "Age", "Pos", "Age Band",
            "Min/Match", "Matches", "Acc Pass %", "Yellow/90", "Red/90"}
    for j, col in enumerate(cols):
        if col in WIDE:
            ws.set_column(j, j, 22)
        elif col in NARR or "%" in col:
            ws.set_column(j, j, 10)
        elif col in {"Contract", "Mkt Val (€)", "Model Val (€)"}:
            ws.set_column(j, j, 14)
        else:
            ws.set_column(j, j, 11)


# ── README ────────────────────────────────────────────────────────────────────

def build_readme(ws, leagues: list[str], total: int, clear: int, budget: int) -> None:
    ws.title = "README"
    ws.sheet_view.showGridLines = False

    ws.append(["LAMBERTS INDEX — FULL PLAYER INTELLIGENCE REPORT  ·  FC Hradec Králové 2025–26"])
    ws.merge_cells("A1:T1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=15)
    ws["A1"].fill = _fill(C["dark"])
    ws.row_dimensions[1].height = 32

    n_l = f"{len(leagues)} leagues" if len(leagues) > 5 else " + ".join(leagues)
    ws.append([f"Waltzing Analytics  ·  Lamberts Index methodology  ·  {n_l}  ·  "
               f"GK · CB · FB · DM · CM · W · FW  ·  Budget ≤ €{budget:,}"])
    ws.merge_cells("A2:T2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=10)
    ws["A2"].fill = _fill(C["dark"])
    ws.row_dimensions[2].height = 18
    ws.append([None])

    def section(label):
        ws.append([None, label])
        ws[f"B{ws.max_row}"].font = Font(bold=True, size=12, color=C["dark"])
        ws.row_dimensions[ws.max_row].height = 22

    def table_hdr(*labels):
        ws.append([None] + list(labels))
        r = ws.max_row
        for idx, lbl in enumerate(labels, 2):
            c = ws.cell(r, idx)
            c.font = Font(bold=True, color=C["white"])
            c.fill = _fill(C["header"])
            c.alignment = Alignment(horizontal="left")

    def table_row(col1, col2="", col3=""):
        ws.append([None, col1, col2, col3])
        r = ws.max_row
        ws.cell(r, 2).font = Font(bold=True, color=C["dark"])

    section("WORKBOOK STRUCTURE")
    table_hdr("Sheet", "Contents", "Rows")
    structure = [
        ("README",              "This guide — methodology, index definitions", ""),
        ("Priority List",       f"All {clear} clear upgrades ranked by Lamberts Index", str(clear)),
        ("Elite Picks",         "Players with Lamberts Index ≥ 30 — strongest buy signals", ""),
        ("Lamberts Analysis",   "Tier + position + league breakdown summary", ""),
        ("Top 5",               "Best 5 per position by Lamberts Index", ""),
        ("All Players",         f"Full {total}-player database — all stats — use AutoFilter", str(total)),
        ("GK",                  "Goalkeeper targets — tailored GK stat columns", ""),
        ("CB",                  "Centre-back targets — defensive + build-up stats", ""),
        ("FB",                  "Full-back targets — defence + crossing + carrying", ""),
        ("DM",                  "Defensive mid targets — defence + passing stats", ""),
        ("CM",                  "Central mid targets — passing + creativity + work rate", ""),
        ("W",                   "Winger targets — goals + creating + dribbling", ""),
        ("FW",                  "Forward targets — scoring + movement + aerial", ""),
        ("League Analysis",     "Average LI Score and tier distribution per league", ""),
        ("Age Bands",           "Breakdown by age group: U21 · U23 · U25 · U28 · U30+", ""),
        ("Physical Profiles",   "Height, weight, foot, nationality for every player", ""),
        ("Expiring 2026",       "Contract expiry opportunities — free or discounted", ""),
        ("Budget Planner",      f"Best clear upgrades within €{budget:,}", ""),
        ("Squad",               "Hradec Králové squad quality benchmarks per position", ""),
    ]
    for sht, desc, rows in structure:
        table_row(sht, desc, rows)

    ws.append([None])
    section("INDEX DEFINITIONS")
    table_hdr("Index", "Formula / Explanation")

    defs = [
        ("LI Score (0–100)",
         "Position-specific SQS percentile rank. Uses a blueprint of 6–10 weighted metrics tailored to "
         "each position. 100 = best player in that position across all scouted leagues."),
        ("Lamberts Index",
         "LI Score Rank − Market Value Rank. Positive = performs above market price. "
         "TOP 5 (best per pos)  ·  ELITE ≥ 30  ·  HIGH ≥ 20  ·  VALUE ≥ 10  ·  FAIR 0–9  ·  OVERPRICED < 0"),
        ("DIS — Defensive Impact Score (def positions only)",
         "Weighted rank 0–100: Def actions/90 ×3 · Def duels won% ×2.5 · "
         "Interceptions/90 ×2 · PAdj Interceptions ×2 · Aerial won% ×1.5 · Blocks/90 ×1"),
        ("ADI — Aerial Dominance Index (def positions only)",
         "Aerial duels/90 × aerial won% / 100, then ranked 0–100. "
         "Combines volume and quality — a player who wins 80% of 5 aerials beats one who wins 50% of 6."),
        ("PADS — Pressure-Adjusted Defensive Score (def positions only)",
         "Uses Wyscout PAdj metrics to strip possession-context bias: "
         "PAdj Interceptions ×2.5 · PAdj Sliding tackles ×2.5 · Blocks/90 ×1.5 · Def duels/90 ×1"),
        ("PDS — Positional Discipline Score (def positions only)",
         "Inverted rank of foul cost (Fouls/90 ×0.6 + Yellow/90 ×0.3 + Red/90 ×0.1). "
         "High PDS = defends cleanly."),
        ("vs Hradec",
         "LI Score − Hradec starter benchmark at same position. "
         ">0 = CLEAR UPGRADE  ·  −10..0 = ROTATIONAL / COVER  ·  <−10 = DEPTH"),
        ("Model Val (€)",
         "Modelled market value = Mkt Val × max(LI Score / 50, 0.1). "
         "Players with LI Score > 50 are valued above market; < 50 below."),
    ]
    for term, desc in defs:
        ws.append([None, term, desc])
        r = ws.max_row
        ws.cell(r, 2).font = Font(bold=True)
        ws.cell(r, 2).fill = _fill(C["mint"])
        ws.cell(r, 3).font = Font(size=9)

    ws.append([None])
    section("BLUEPRINT METRICS BY POSITION")
    table_hdr("Position", "Metrics (highest weight first)")
    for pos, bp in SQS_BLUEPRINTS.items():
        metrics = " · ".join(f"{col} ×{w}" for col, w in bp)
        ws.append([None, pos, metrics])
        ws.cell(ws.max_row, 2).font = Font(bold=True)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 100
    ws.column_dimensions["D"].width = 10


# ── Lamberts Analysis sheet ───────────────────────────────────────────────────

def build_analysis(ws, master: pd.DataFrame) -> None:
    ws.title = "Lamberts Analysis"
    ws.sheet_view.showGridLines = False

    def hdr_row(text, row_n, cols="A:M", size=13):
        ws.append([text])
        ws.merge_cells(f"A{ws.max_row}:{cols.split(':')[1]}{ws.max_row}")
        ws[f"A{ws.max_row}"].font = Font(bold=True, color=C["white"], size=size)
        ws[f"A{ws.max_row}"].fill = _fill(C["dark"])
        ws.row_dimensions[ws.max_row].height = 22 if size < 14 else 28

    def sub_row(text):
        ws.append([text])
        ws.merge_cells(f"A{ws.max_row}:M{ws.max_row}")
        ws[f"A{ws.max_row}"].font = Font(italic=True, color=C["gold"], size=9)
        ws[f"A{ws.max_row}"].fill = _fill(C["dark"])

    hdr_row("LAMBERTS INDEX — FULL ANALYSIS BREAKDOWN", 1, "A:M", 14)
    sub_row("Lamberts Index = LI Score Rank − Market Value Rank  ·  TOP 5 (best per pos)  ·  ELITE ≥30  ·  HIGH ≥20  ·  VALUE ≥10  ·  FAIR 0–9  ·  OVER <0")
    ws.append([None])

    def tbl_hdr(labels):
        ws.append(labels)
        r = ws.max_row
        for idx in range(1, len(labels) + 1):
            c = ws.cell(r, idx)
            if c.value:
                c.font = Font(bold=True, color=C["white"])
                c.fill = _fill(C["header"])
                c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 18

    def section_title(text):
        ws.append([text])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=11, color=C["dark"])
        ws.row_dimensions[ws.max_row].height = 20

    li = master["Lamberts Index"]

    # ── Overall statistics ─────────────────────────────────────────────────────
    section_title("OVERALL STATISTICS")
    overall = [
        ("Total players",   f"{len(master):,}"),
        ("Clear Upgrades",  f"{(master['Status']=='CLEAR UPGRADE').sum():,}"),
        ("TOP 5",           f"{(master['Tier']=='TOP 5').sum():,}"),
        ("ELITE VALUE",     f"{(master['Tier']=='ELITE VALUE').sum():,}"),
        ("HIGH VALUE",      f"{(master['Tier']=='HIGH VALUE').sum():,}"),
        ("VALUE",           f"{(master['Tier']=='VALUE').sum():,}"),
        ("Mean LI",         f"{li.mean():.2f}"),
        ("Median LI",       f"{li.median():.2f}"),
        ("Max LI",          f"{li.max():.2f}"),
        ("Min LI",          f"{li.min():.2f}"),
        ("Std Dev",         f"{li.std():.2f}"),
    ]
    tbl_hdr(["Metric", "Value"])
    for k, v in overall:
        ws.append([k, v])
        ws[f"A{ws.max_row}"].font = Font(bold=True)
    ws.append([None])

    # ── Position breakdown ─────────────────────────────────────────────────────
    section_title("BREAKDOWN BY POSITION")
    tbl_hdr(["Position", "Players", "Clear Upgrades", "Top 5", "Elite", "High",
             "Value", "Fair", "Over", "Avg LI Score", "Avg LI", "Avg Mkt Val (€)"])
    tier_order = ["TOP 5", "ELITE VALUE", "HIGH VALUE", "VALUE", "FAIR VALUE", "OVERPRICED"]
    for pos in ["GK", "CB", "FB", "DM", "CM", "W", "FW"]:
        grp = master[master["Pos"] == pos]
        if grp.empty:
            continue
        tc = grp["Tier"].value_counts()
        ws.append([
            pos, len(grp),
            (grp["Status"] == "CLEAR UPGRADE").sum(),
            tc.get("TOP 5", 0), tc.get("ELITE VALUE", 0), tc.get("HIGH VALUE", 0),
            tc.get("VALUE", 0), tc.get("FAIR VALUE", 0), tc.get("OVERPRICED", 0),
            round(grp["LI Score"].mean(), 2),
            round(grp["Lamberts Index"].mean(), 2),
            int(grp["Mkt Val (€)"].mean()) if grp["Mkt Val (€)"].mean() > 0 else 0,
        ])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.append([None])

    # ── Tier distribution ──────────────────────────────────────────────────────
    section_title("TIER DISTRIBUTION")
    tbl_hdr(["Tier", "Count", "% Total", "Avg Age", "Avg Mkt Val (€)", "Avg LI"])
    for tier in tier_order:
        grp = master[master["Tier"] == tier]
        ws.append([
            tier,
            len(grp),
            f"{len(grp)/len(master)*100:.1f}%",
            round(pd.to_numeric(grp["Age"], errors="coerce").mean(), 1),
            int(grp["Mkt Val (€)"].mean()) if not grp.empty and grp["Mkt Val (€)"].mean() > 0 else 0,
            round(grp["Lamberts Index"].mean(), 2) if not grp.empty else 0,
        ])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.append([None])

    # ── Age analysis ───────────────────────────────────────────────────────────
    section_title("AGE BAND ANALYSIS")
    tbl_hdr(["Age Band", "Players", "Avg LI Score", "Avg LI", "Clear Upgrades", "Elite", "Avg Mkt Val (€)"])
    for band in ["U21", "U23", "U25", "U28", "U30+"]:
        grp = master[master["Age Band"] == band]
        ws.append([
            band, len(grp),
            round(grp["LI Score"].mean(), 2) if not grp.empty else 0,
            round(grp["Lamberts Index"].mean(), 2) if not grp.empty else 0,
            (grp["Status"] == "CLEAR UPGRADE").sum(),
            (grp["Tier"] == "ELITE VALUE").sum(),
            int(grp["Mkt Val (€)"].mean()) if not grp.empty and grp["Mkt Val (€)"].mean() > 0 else 0,
        ])
        ws.cell(ws.max_row, 1).font = Font(bold=True)

    for col in "ABCDEFGHIJK":
        ws.column_dimensions[col].width = 18


# ── Top 5 per position ────────────────────────────────────────────────────────

def build_top5(ws, master: pd.DataFrame) -> None:
    ws.title = "Top 5"
    ws.sheet_view.showGridLines = False

    ws.append(["LAMBERTS INDEX — TOP 5 PER POSITION  ·  Ranked by Lamberts Index"])
    ws.merge_cells("A1:P1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=14)
    ws["A1"].fill = _fill(C["dark"])
    ws.row_dimensions[1].height = 26

    ws.append(["Highest Lamberts Index per position group across all scouted leagues"])
    ws.merge_cells("A2:P2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["dark"])

    for pos, label in POS_LABELS.items():
        grp = master[master["Pos"] == pos].head(5)
        if grp.empty:
            continue

        ws.append([f"  {label}"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=16)
        ws[f"A{ws.max_row}"].font = Font(bold=True, color=C["white"], size=11)
        ws[f"A{ws.max_row}"].fill = _fill(POS_COLOR.get(pos, C["dark"]))
        ws.row_dimensions[ws.max_row].height = 20

        hdrs = ["#", "Player", "Team", "League", "Age", "Birth Country",
                "Height", "Foot", "Contract", "Mkt Val (€)",
                "LI Score", "Lamberts Index", "DIS", "ADI", "PADS", "PDS", "Status"]
        ws.append(hdrs)
        hdr_r = ws.max_row
        for cell in ws[hdr_r]:
            if cell.value:
                cell.font = Font(bold=True, color=C["white"], size=9)
                cell.fill = _fill(C["header"])
                cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[hdr_r].height = 18

        for i, (_, r) in enumerate(grp.iterrows()):
            row_data = [
                i + 1, r["Player"], r["Team"], r["League"], r["Age"], r["Birth Country"],
                r["Height (cm)"], r["Foot"], r["Contract"], r["Mkt Val (€)"],
                r["LI Score"], r["Lamberts Index"],
                r.get("DIS", ""), r.get("ADI", ""), r.get("PADS", ""), r.get("PDS", ""),
                r["Status"],
            ]
            ws.append(row_data)
            dr = ws.max_row
            bg = C["mint"] if i % 2 == 0 else C["white"]
            for cell in ws[dr]:
                cell.font = Font(size=9)
                cell.fill = _fill(bg)
                cell.alignment = Alignment(horizontal="center")
            ws.cell(dr, 1).font = Font(bold=True, size=10)
            ws.cell(dr, 2).font = Font(bold=True, size=9)

        ws.append([None])

    _autofit(ws)


# ── League Analysis ───────────────────────────────────────────────────────────

def build_league_analysis(ws, master: pd.DataFrame) -> None:
    ws.title = "League Analysis"
    ws.sheet_view.showGridLines = False

    ws.append(["LAMBERTS INDEX — LEAGUE ANALYSIS"])
    ws.merge_cells("A1:N1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=14)
    ws["A1"].fill = _fill(C["dark"])
    ws.row_dimensions[1].height = 26

    ws.append(["Average LI Score, tier distribution and player counts per scouted league"])
    ws.merge_cells("A2:N2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["dark"])

    cols = ["League", "Players", "Clear Upgrades", "Elite",
            "High", "Value", "Avg LI Score", "Max LI Score",
            "Avg LI", "Max LI", "Avg Age", "Avg Mkt Val (€)",
            "GK", "CB", "FB", "DM", "CM", "W", "FW"]
    ws.append(cols)
    hdr = ws.max_row
    for cell in ws[hdr]:
        if cell.value:
            cell.font = Font(bold=True, color=C["white"], size=9)
            cell.fill = _fill(C["header"])
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hdr].height = 30

    leagues = master.groupby("League")

    data_rows = []
    for league, grp in leagues:
        tc = grp["Tier"].value_counts()
        pc = grp["Pos"].value_counts()
        data_rows.append({
            "League":        league,
            "Players":       len(grp),
            "Clear Upgrades":(grp["Status"] == "CLEAR UPGRADE").sum(),
            "Elite":          tc.get("ELITE VALUE", 0),
            "High":           tc.get("HIGH VALUE", 0),
            "Value":          tc.get("VALUE", 0),
            "Avg LI Score":   round(grp["LI Score"].mean(), 2),
            "Max LI Score":   round(grp["LI Score"].max(), 2),
            "Avg LI":         round(grp["Lamberts Index"].mean(), 2),
            "Max LI":         round(grp["Lamberts Index"].max(), 2),
            "Avg Age":        round(pd.to_numeric(grp["Age"], errors="coerce").mean(), 1),
            "Avg Mkt Val (€)":int(grp["Mkt Val (€)"].mean()) if grp["Mkt Val (€)"].mean() > 0 else 0,
            "GK": pc.get("GK", 0), "CB": pc.get("CB", 0), "FB": pc.get("FB", 0),
            "DM": pc.get("DM", 0), "CM": pc.get("CM", 0),
            "W": pc.get("W", 0),   "FW": pc.get("FW", 0),
        })

    # Sort by Avg LI Score descending
    data_rows.sort(key=lambda x: x["Avg LI Score"], reverse=True)

    for i, row in enumerate(data_rows):
        ws.append([row[c] for c in cols])
        dr = ws.max_row
        bg = C["mint"] if i % 2 == 0 else C["white"]
        for cell in ws[dr]:
            cell.font = Font(size=9)
            cell.fill = _fill(bg)
            cell.alignment = Alignment(horizontal="center")
        ws.cell(dr, 1).font = Font(bold=True, size=9)
        ws.cell(dr, 1).alignment = Alignment(horizontal="left")

    ws.freeze_panes = f"A{hdr + 1}"
    _autofit(ws)
    ws.column_dimensions["A"].width = 28


# ── Age Bands sheet ───────────────────────────────────────────────────────────

def build_age_bands(ws, master: pd.DataFrame) -> None:
    ws.title = "Age Bands"
    ws.sheet_view.showGridLines = False

    ws.append(["LAMBERTS INDEX — AGE BAND ANALYSIS"])
    ws.merge_cells("A1:N1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=14)
    ws["A1"].fill = _fill(C["dark"])
    ws.row_dimensions[1].height = 26

    ws.append(["Clear upgrade candidates segmented by age bracket across all positions"])
    ws.merge_cells("A2:N2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["dark"])

    bands = ["U21", "U23", "U25", "U28", "U30+"]
    band_colors = {"U21": "117A65", "U23": "1A5276", "U25": "7D6608",
                   "U28": "784212", "U30+": "626567"}

    for band in bands:
        grp = master[master["Age Band"] == band].copy()
        if grp.empty:
            continue

        color = band_colors.get(band, C["dark"])
        label = {
            "U21": "U21 — AGE ≤ 21 (NEXT GENERATION)",
            "U23": "U23 — AGE 22–23 (BREAKTHROUGH TALENT)",
            "U25": "U25 — AGE 24–25 (PRIME DEVELOPMENT)",
            "U28": "U28 — AGE 26–28 (PEAK YEARS)",
            "U30+": "U30+ — AGE 29–30 (EXPERIENCED)",
        }.get(band, band)

        ws.append([f"  {label}  ·  {len(grp)} players  ·  "
                   f"{(grp['Status']=='CLEAR UPGRADE').sum()} clear upgrades  ·  "
                   f"Avg LI: {grp['Lamberts Index'].mean():.2f}"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=14)
        ws[f"A{ws.max_row}"].font = Font(bold=True, color=C["white"], size=11)
        ws[f"A{ws.max_row}"].fill = _fill(color)
        ws.row_dimensions[ws.max_row].height = 20

        hdrs = ["Player", "Team", "League", "Pos", "Age", "Birth Country",
                "Foot", "Contract", "Mkt Val (€)", "LI Score", "Lamberts Index", "Tier", "Status", "vs Hradec"]
        ws.append(hdrs)
        hdr_r = ws.max_row
        for cell in ws[hdr_r]:
            if cell.value:
                cell.font = Font(bold=True, color=C["white"], size=9)
                cell.fill = _fill(C["header"])
                cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[hdr_r].height = 18

        top_grp = grp.sort_values("Lamberts Index", ascending=False).head(50)
        for i, (_, r) in enumerate(top_grp.iterrows()):
            ws.append([
                r["Player"], r["Team"], r["League"], r["Pos"], r["Age"],
                r["Birth Country"], r["Foot"], r["Contract"],
                r["Mkt Val (€)"], r["LI Score"], r["Lamberts Index"],
                r["Tier"], r["Status"], r["vs Hradec"],
            ])
            dr = ws.max_row
            bg = C["mint"] if i % 2 == 0 else C["white"]
            for cell in ws[dr]:
                cell.font = Font(size=9)
                cell.fill = _fill(bg)
                cell.alignment = Alignment(horizontal="center")

        ws.append([None])

    _autofit(ws)


# ── Physical Profiles ─────────────────────────────────────────────────────────

def build_physical_profiles(ws, master: pd.DataFrame) -> None:
    ws.title = "Physical Profiles"
    ws.sheet_view.showGridLines = False

    ws.append(["LAMBERTS INDEX — PHYSICAL PROFILES"])
    ws.merge_cells("A1:P1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=14)
    ws["A1"].fill = _fill(C["dark"])
    ws.row_dimensions[1].height = 26

    ws.append(["Height · Weight · Preferred Foot · Nationality for all scouted players"])
    ws.merge_cells("A2:P2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["dark"])

    cols = ["Player", "Team", "League", "Pos", "Age", "Birth Country",
            "Height (cm)", "Weight (kg)", "Foot", "On Loan",
            "Contract", "Mkt Val (€)", "LI Score", "Lamberts Index", "Tier", "Status"]
    ws.append(cols)
    hdr = ws.max_row
    for cell in ws[hdr]:
        if cell.value:
            cell.font = Font(bold=True, color=C["white"], size=9)
            cell.fill = _fill(C["header"])
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hdr].height = 28

    df = master[cols].copy()
    for i, row_vals in enumerate(df.itertuples(index=False)):
        ws.append(list(row_vals))
        dr = ws.max_row
        bg = C["mint"] if i % 2 == 0 else C["white"]
        tier_val = str(row_vals[cols.index("Tier")])
        for cell in ws[dr]:
            cell.font = Font(size=9)
            cell.fill = _fill(bg)
            cell.alignment = Alignment(horizontal="center")
        ws.cell(dr, 1).font = Font(bold=True, size=9)
        ws.cell(dr, 1).alignment = Alignment(horizontal="left")
        # Tier colour
        tc = ws.cell(dr, cols.index("Tier") + 1)
        thex = {"TOP 5": C["top5"], "ELITE VALUE": C["elite"], "HIGH VALUE": C["high"],
                "VALUE": C["value"], "FAIR VALUE": C["fair"],
                "OVERPRICED": C["over"], "TOP 5": C["top5"]}.get(tier_val)
        if thex:
            tc.fill = _fill(thex)
            tc.font = Font(bold=True, color=C["white"], size=9)

    ws.freeze_panes = f"A{hdr + 1}"
    _autofit(ws)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22


# ── Budget Planner ────────────────────────────────────────────────────────────

def build_budget_planner(ws, master: pd.DataFrame, budget: int) -> None:
    ws.title = "Budget Planner"
    ws.sheet_view.showGridLines = False

    ws.append([f"LAMBERTS INDEX — BUDGET PLANNER  ·  FC Hradec Králové  ·  Cap: €{budget:,}"])
    ws.merge_cells("A1:R1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=13)
    ws["A1"].fill = _fill(C["dark"])
    ws.row_dimensions[1].height = 24

    ws.append(["ELITE VALUE + HIGH VALUE clear upgrades — sorted cheapest first — running total shown"])
    ws.merge_cells("A2:R2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["dark"])

    ws.append([None, None, None, None, None, "Total Budget →", f"€{budget:,}"])
    ws[f"F{ws.max_row}"].font = Font(bold=True)
    ws[f"G{ws.max_row}"].font = Font(bold=True, color=C["dark"])
    ws[f"G{ws.max_row}"].fill = _fill(C["gold"])

    eligible = master[
        master["Tier"].isin(["ELITE VALUE", "HIGH VALUE"]) &
        (master["Status"] == "CLEAR UPGRADE")
    ].sort_values("Mkt Val (€)").reset_index(drop=True)

    cols = ["#", "Player", "Pos", "Age", "Team", "League", "Birth Country",
            "Foot", "Height (cm)", "Weight (kg)", "Contract",
            "Mkt Val (€)", "Model Val (€)", "Val Ratio",
            "LI Score", "Lamberts Index", "vs Hradec", "Running Total (€)"]

    ws.append(cols)
    hdr = ws.max_row
    for cell in ws[hdr]:
        if cell.value:
            cell.font = Font(bold=True, color=C["white"], size=9)
            cell.fill = _fill(C["header"])
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hdr].height = 30

    running = 0
    for i, (_, r) in enumerate(eligible.iterrows(), 1):
        mv = int(r["Mkt Val (€)"])
        running += mv
        over = running > budget
        ws.append([
            i, r["Player"], r["Pos"], r["Age"],
            r["Team"], r["League"], r["Birth Country"],
            r["Foot"], r["Height (cm)"], r["Weight (kg)"], r["Contract"],
            mv, int(r["Model Val (€)"]), r["Val Ratio"],
            r["LI Score"], r["Lamberts Index"], r["vs Hradec"], running,
        ])
        dr = ws.max_row
        row_bg = "FDE8E8" if over else (C["mint"] if i % 2 == 0 else C["white"])
        for cell in ws[dr]:
            cell.font = Font(size=9, color="922B21" if over else "000000")
            cell.fill = _fill(row_bg)
            cell.alignment = Alignment(horizontal="center")
        if over:
            ws.cell(dr, 18).font = Font(bold=True, size=9, color="922B21")

    ws.freeze_panes = f"A{hdr + 1}"
    _autofit(ws)
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 22


# ── Expiring contracts ────────────────────────────────────────────────────────

def build_expiring(wb_x, master: pd.DataFrame) -> None:
    df = master[master["Exp?"].astype(str) == "2026"].copy()
    cols = [c for c in POS_COLS["ALL"] if c in df.columns]
    df = df[cols]
    write_data_sheet(
        wb_x, "Expiring 2026",
        f"LAMBERTS INDEX — EXPIRING CONTRACTS 2026  ·  {len(df)} Players",
        "Free agent / pre-contract opportunities from January 2026  ·  Sorted by Lamberts Index",
        df,
    )


# ── Squad ─────────────────────────────────────────────────────────────────────

def build_squad(ws) -> None:
    ws.title = "Squad"
    ws.sheet_view.showGridLines = False

    ws.append(["FC HRADEC KRÁLOVÉ — SQUAD QUALITY BENCHMARKS  ·  All Positions"])
    ws.merge_cells("A1:L1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=14)
    ws["A1"].fill = _fill(C["dark"])
    ws.row_dimensions[1].height = 26

    ws.append(["Czech top-flight LI Score percentile for each starter  ·  Basis for vs Hradec calculation"])
    ws.merge_cells("A2:L2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["dark"])

    ws.append([None])
    ws.append(["SQUAD BENCHMARK RATINGS"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)

    ws.append(["Position", "LI Score Benchmark", "Hradec Reference", "Status in Czech 1.liga"])
    hdr = ws.max_row
    for cell in ws[hdr]:
        if cell.value:
            cell.font = Font(bold=True, color=C["white"])
            cell.fill = _fill(C["header"])
            cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[hdr].height = 20

    squad_data = {
        "GK": ("Adam Zadrazil",   "Starter — above average",      61.5),
        "CB": ("Daniel Horak",    "Weakest starter",               33.3),
        "FB": ("Martin Suchomel", "Starter — league-high level",   81.8),
        "DM": ("Jakub Elbel",     "Starter — below average",       36.8),
        "CM": ("Median starter",  "Median mid quality",             53.7),
        "W":  ("Weakest winger",  "Significant weakness",           27.6),
        "FW": ("Reference FW",    "Average striking level",         38.3),
    }
    for pos, (name, note, score) in squad_data.items():
        ws.append([pos, score, name, note])
        r = ws.max_row
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2).alignment = Alignment(horizontal="center")
        # Color-code quality
        color = C["high"] if score >= 60 else C["rota"] if score >= 40 else C["over"]
        ws.cell(r, 2).fill = _fill(color)
        ws.cell(r, 2).font = Font(bold=True, color=C["white"])

    ws.append([None])
    ws.append(["HOW vs HRADEC IS CALCULATED"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)
    ws.append(["vs Hradec = Target LI Score − Hradec benchmark at same position"])
    ws.append(["> 0     = CLEAR UPGRADE  ·  player is better than current Hradec starter"])
    ws.append(["−10..0  = ROTATIONAL / COVER  ·  close to or just below starter level"])
    ws.append(["< −10   = DEPTH  ·  clearly below current starter, squad-depth role only"])

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 45


# ── Main ───────────────────────────────────────────────────────────────────────

def run(leagues: list[str] | None, min_minutes: int, max_age: int,
        budget: int, output: Path) -> None:
    print(f"\n{'='*68}")
    print("  LAMBERTS INDEX — Full Player Intelligence Report")
    print(f"  Positions: GK · CB · FB · DM · CM · W · FW")
    print(f"  Leagues: {'ALL' if leagues is None else leagues}")
    print(f"  Min minutes: {min_minutes}  |  Max age: {max_age}")
    print(f"{'='*68}\n")

    print("Loading Wyscout files…")
    raw = load_leagues(leagues, min_minutes)

    if "Age" in raw.columns:
        raw = raw[pd.to_numeric(raw["Age"], errors="coerce").fillna(99) <= max_age]
        print(f"  → {len(raw)} players after age ≤ {max_age}")

    raw = add_position_group(raw)
    raw = raw[raw["_pos_group"].isin(ALL_POSITIONS)].copy()
    print(f"  → {len(raw)} players across all positions")

    print("Computing Lamberts Index scores…")
    raw = compute_sqs(raw)
    raw = compute_defensive_indexes(raw)
    raw = compute_mv_rank(raw)
    raw = compute_lamberts_index(raw)
    raw = compute_vs_hradec(raw)

    print("Building master table…")
    master = build_master(raw)

    total = len(master)
    clear = int((master["Status"] == "CLEAR UPGRADE").sum())
    elite = int((master["Tier"] == "ELITE VALUE").sum())
    print(f"  → {total:,} players | {clear:,} clear upgrades | {elite:,} ELITE VALUE")

    league_list = leagues if leagues else sorted(master["League"].unique().tolist())

    # Force .xlsx
    if output.suffix.lower() not in {".xlsx", ".xls"}:
        output = output.with_suffix(".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)
    xls_path = output.with_suffix(".tmp.xlsx")

    def _cols(pos: str) -> list[str]:
        """Get columns for a position sheet, filtered to what master has."""
        return [c for c in POS_COLS[pos] if c in master.columns]

    # ── Phase 1: xlsxwriter — large data sheets ────────────────────────────────
    print(f"\nWriting workbook → {output}")
    print("  Building data sheets (xlsxwriter)…")

    wb_x = xlsxwriter.Workbook(str(xls_path), {"constant_memory": False})

    # Priority List
    prio_df = master[_cols("ALL")][master["Status"] == "CLEAR UPGRADE"].copy()
    print(f"    Priority List ({len(prio_df):,} rows)…")
    write_data_sheet(wb_x, "Priority List",
        f"LAMBERTS INDEX — PRIORITY LIST  ·  {len(prio_df):,} Clear Upgrades",
        "All outperform Hradec starters at same position  ·  Sorted by Lamberts Index",
        prio_df)

    # Elite Picks
    elite_df = master[_cols("ALL")][master["Tier"] == "ELITE VALUE"].copy()
    print(f"    Elite Picks ({len(elite_df):,} rows)…")
    write_data_sheet(wb_x, "Elite Picks",
        f"LAMBERTS INDEX — ELITE VALUE  ·  {len(elite_df):,} Players with LI ≥ 30",
        "Strongest buy signals across all positions  ·  Sorted by Lamberts Index",
        elite_df)

    # All Players
    all_df = master[_cols("ALL")].copy()
    print(f"    All Players ({total:,} rows)…")
    write_data_sheet(wb_x, "All Players",
        f"LAMBERTS INDEX — ALL PLAYERS  ·  {total:,} Candidates  ·  {len(league_list)} Leagues",
        "Full database with all raw stats — use column AutoFilter dropdowns to narrow results",
        all_df)

    # Position sheets
    for pos in ["GK", "CB", "FB", "DM", "CM", "W", "FW"]:
        cols = _cols(pos)
        grp  = master[cols][master["Pos"] == pos].copy()
        print(f"    {pos} ({len(grp):,} rows)…")
        write_data_sheet(wb_x, pos,
            f"LAMBERTS INDEX — {POS_LABELS[pos]}  ·  {len(grp):,} Candidates",
            f"Position-tailored columns  ·  Sorted by Lamberts Index",
            grp, pos_accent=POS_COLOR.get(pos))

    # Expiring 2026
    build_expiring(wb_x, master)

    wb_x.close()
    print("  xlsxwriter sheets done.")

    # ── Phase 2: openpyxl — summary/structural sheets ──────────────────────────
    print("  Building summary sheets (openpyxl)…")
    wb = load_workbook(str(xls_path))

    # Insert at front
    ws_readme   = wb.create_sheet("README",           0)
    ws_analysis = wb.create_sheet("Lamberts Analysis", 3)
    ws_top5     = wb.create_sheet("Top 5",             4)
    ws_league   = wb.create_sheet("League Analysis")
    ws_age      = wb.create_sheet("Age Bands")
    ws_phys     = wb.create_sheet("Physical Profiles")
    ws_budget   = wb.create_sheet("Budget Planner")
    ws_squad    = wb.create_sheet("Squad")

    build_readme(ws_readme, league_list, total, clear, budget)
    build_analysis(ws_analysis, master)
    build_top5(ws_top5, master)
    build_league_analysis(ws_league, master)
    build_age_bands(ws_age, master)
    build_physical_profiles(ws_phys, master)
    build_budget_planner(ws_budget, master, budget)
    build_squad(ws_squad)

    if "Sheet1" in wb.sheetnames:
        del wb["Sheet1"]

    wb.save(str(output))
    xls_path.unlink(missing_ok=True)

    size_mb = output.stat().st_size / 1_048_576
    print(f"\nDone.  {size_mb:.1f} MB → {output.resolve()}")
    print(f"Sheets ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Lamberts Index — Full Player Intelligence Report from Wyscout data"
    )
    parser.add_argument("--leagues",     nargs="+", default=None,
                        help="League names (without .xlsx). Omit for ALL leagues.")
    parser.add_argument("--min-minutes", type=int, default=DEFAULT_MIN_MINUTES)
    parser.add_argument("--max-age",     type=int, default=DEFAULT_MAX_AGE)
    parser.add_argument("--budget",      type=int, default=DEFAULT_BUDGET)
    parser.add_argument("--output",      type=Path,
                        default=ROOT / "data" / "Lamberts_Index_Full_Report.xlsx")
    args = parser.parse_args()
    run(
        leagues=args.leagues,
        min_minutes=args.min_minutes,
        max_age=args.max_age,
        budget=args.budget,
        output=args.output,
    )
