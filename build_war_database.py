"""
build_war_database.py — WAR + Baseball-style scouting ratings for all wide attackers

WAR Methodology
---------------
Offensive rate  = xG/90 + xA/90×0.85 + ShotAssists/90×0.28
Carry rate      = ProgRuns/90×0.025  + (Dribbles/90 × Drib%/100)×0.015
Replacement     = 15th-percentile total rate, all DB wide attackers (≥300 min)
WAR             = (rate − replacement) × (minutes/90) ÷ 3.0 goals/win

Baseball-Style Grades (20-80 Scale)
-------------------------------------
20-80 maps to z-scores via norm.ppf:  grade = 50 + z×10, clipped to [20,80]
  80 = 99.9th pctile (elite)   60 = 84th (plus)    50 = average
  70 = 97.7th (plus-plus)      55 = 69th (above)   40 = 16th (below avg)

Tools graded: Finishing · Dribbling · Athleticism · Crossing · Creativity · Defending
OVR  = weighted average of tools (current grade)
FV   = OVR adjusted for age ceiling (Future Value)
Scout = letter grade from OVR (A / B / C / D / F with +/-)

Plus-Stats (100 = league average, like OPS+/wRC+)
--------------------------------------------------
xG+    = player xG/90 ÷ DB-mean xG/90 × 100
xA+    = player xA/90 ÷ DB-mean xA/90 × 100
xP+    = player (xG+xA)/90 ÷ DB-mean × 100    [wRC+ analogue]
GCONV  = Goals/xG × 100  (goal conversion vs expectation, 100 = neutral)
WAR+   = WAR_per_90 ÷ DB-mean WAR_per_90 × 100

Output: reports/WAR_All_Players.xlsx
  Sheet 1 — WAR Rankings      : all players sorted by WAR
  Sheet 2 — Scouting Grades   : tool grades + OVR/FV/Scout/Plus-stats
  Sheet 3 — Methodology       : parameter reference
"""
from __future__ import annotations
import warnings; warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from pathlib import Path
from scipy.stats import norm

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    HAS_OPX = True
except ImportError:
    HAS_OPX = False

OUT_DIR = Path("reports")
OUT_DIR.mkdir(exist_ok=True)

# ── Constants ──────────────────────────────────────────────────────────────────
WIDE_ATK_POS    = {"LAMF", "RAMF", "LW", "RW", "LWF", "RWF", "AMF", "LWB", "RWB"}
MIN_MINS        = 300
GOALS_PER_WIN   = 3.0
REPL_PERCENTILE = 15

NON_METRIC = {
    "Player", "Team", "Team within selected timeframe", "Position",
    "Age", "Market value", "Contract expires", "Birth country",
    "Passport country", "Foot", "Height", "Weight", "On loan",
}

# Tool weights for OVR (must sum to 1.0)
TOOL_WEIGHTS = {
    "G_Finishing":   0.20,
    "G_Dribbling":   0.25,
    "G_Athleticism": 0.20,
    "G_Crossing":    0.10,
    "G_Creativity":  0.20,
    "G_Defending":   0.05,
}


# ── Data loading ───────────────────────────────────────────────────────────────

def _numeric(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        if col not in NON_METRIC:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def _gc(df: pd.DataFrame, col: str, default: float = 0.0) -> pd.Series:
    return df[col].fillna(default) if col in df.columns else pd.Series(default, index=df.index)


def load_pool() -> pd.DataFrame:
    print("  Loading Wyscout files …")
    parts, skipped = [], []
    for f in sorted(Path("data/Wyscout DB").glob("*.xlsx")):
        if f.stem in ("Czech U17", "Czech U19"):
            continue
        try:
            d = _numeric(pd.read_excel(f))
            d["_league"] = f.stem
            d["_pos1"]   = d["Position"].astype(str).str.split(",").str[0].str.strip()
            parts.append(d)
        except Exception:
            skipped.append(f.stem)

    if skipped:
        print(f"  Skipped {len(skipped)} files (read errors)")

    df_all = pd.concat(parts, ignore_index=True)
    mask   = (df_all["_pos1"].isin(WIDE_ATK_POS) &
              (df_all["Minutes played"].fillna(0) >= MIN_MINS))
    pool   = df_all[mask].copy().reset_index(drop=True)
    print(f"  {len(pool):,} wide attackers across {df_all['_league'].nunique()} leagues")
    return pool


# ── WAR calculation ────────────────────────────────────────────────────────────

def compute_war_rates(df: pd.DataFrame):
    xg  = _gc(df, "xG per 90")
    xa  = _gc(df, "xA per 90")
    sa  = _gc(df, "Shot assists per 90")
    pr  = _gc(df, "Progressive runs per 90")
    d   = _gc(df, "Dribbles per 90")
    dp  = _gc(df, "Successful dribbles, %", 50.0)
    off   = xg + xa * 0.85 + sa * 0.28
    carry = pr * 0.025 + (d * dp / 100.0) * 0.015
    return off, carry, off + carry


# ── Baseball-style grade engine ────────────────────────────────────────────────

def pct_to_grade(pct: float | pd.Series) -> int | pd.Series:
    """Percentile (0–100) → 20-80 scouting grade via normal z-score."""
    p = np.clip(pct / 100, 0.001, 0.999)
    grade = 50 + norm.ppf(p) * 10
    if isinstance(grade, np.ndarray):
        return np.clip(grade, 20, 80).round().astype(int)
    return int(np.clip(grade, 20, 80).round())


def grade_to_letter(ovr: float) -> str:
    """20-80 OVR → letter grade with +/-."""
    if ovr >= 75: return "A+"
    if ovr >= 70: return "A"
    if ovr >= 65: return "A-"
    if ovr >= 62: return "B+"
    if ovr >= 57: return "B"
    if ovr >= 53: return "B-"
    if ovr >= 51: return "C+"
    if ovr >= 48: return "C"
    if ovr >= 44: return "C-"
    if ovr >= 40: return "D+"
    if ovr >= 35: return "D"
    return "F"


def fv_age_bonus(age) -> int:
    """Age adjustment to OVR for future value (ceiling projection)."""
    if pd.isna(age): return 0
    a = float(age)
    if a <= 18: return 8
    if a <= 20: return 6
    if a <= 22: return 4
    if a <= 24: return 2
    if a <= 27: return 0
    if a <= 29: return -2
    if a <= 31: return -4
    return -7


def fv_label(fv: int) -> str:
    if fv >= 75: return "Elite / Franchise"
    if fv >= 70: return "Star / Top-5 talent"
    if fv >= 65: return "Quality regular"
    if fv >= 60: return "Solid starter"
    if fv >= 55: return "Above-average regular"
    if fv >= 50: return "Average professional"
    if fv >= 45: return "Fringe starter"
    if fv >= 40: return "Depth / rotation"
    if fv >= 35: return "Developmental"
    return "Below professional"


def compute_tool_raws(df: pd.DataFrame) -> pd.DataFrame:
    """Compute raw (un-graded) tool metrics for every row."""
    xg  = _gc(df, "xG per 90")
    g   = _gc(df, "Goals per 90")
    sh  = _gc(df, "Shots per 90")
    bx  = _gc(df, "Touches in box per 90")
    d   = _gc(df, "Dribbles per 90")
    dp  = _gc(df, "Successful dribbles, %", 50.0)
    pr  = _gc(df, "Progressive runs per 90")
    ac  = _gc(df, "Accelerations per 90")
    cr  = _gc(df, "Crosses per 90")
    ca  = _gc(df, "Accurate crosses, %", 50.0)
    xa  = _gc(df, "xA per 90")
    sa  = _gc(df, "Shot assists per 90")
    kp  = _gc(df, "Key passes per 90")
    da  = _gc(df, "Successful defensive actions per 90")
    ic  = _gc(df, "Interceptions per 90")

    return pd.DataFrame({
        "_t_finishing":   xg * 0.55 + g * 0.25 + sh * 0.10 + bx * 0.10,
        "_t_dribbling":   d * (dp / 100.0),          # effective dribbles/90
        "_t_athleticism": pr + ac * 0.6,
        "_t_crossing":    cr * (ca / 100.0),          # accurate crosses/90
        "_t_creativity":  xa * 1.8 + sa * 0.55 + kp * 0.30,
        "_t_defending":   da + ic,
    }, index=df.index)


def compute_scouting_grades(pool: pd.DataFrame) -> pd.DataFrame:
    """
    Compute 20-80 tool grades, OVR, FV, letter grade, and plus-stats
    for all rows in pool (should already be deduplicated).
    """
    raws = compute_tool_raws(pool)

    # Percentile of each tool within the full pool
    grade_cols = {}
    for raw_col, grade_col in [
        ("_t_finishing",   "G_Finishing"),
        ("_t_dribbling",   "G_Dribbling"),
        ("_t_athleticism", "G_Athleticism"),
        ("_t_crossing",    "G_Crossing"),
        ("_t_creativity",  "G_Creativity"),
        ("_t_defending",   "G_Defending"),
    ]:
        vals = raws[raw_col].values
        pcts = pd.Series(vals).apply(
            lambda v: float(np.mean(vals < v) * 100) if pd.notna(v) else 50.0
        )
        grade_cols[grade_col] = pcts_to_grade_series(pcts)

    grades = pd.DataFrame(grade_cols, index=pool.index)

    # OVR = weighted average of tool grades
    ovr = sum(grades[col] * w for col, w in TOOL_WEIGHTS.items())
    grades["OVR"] = ovr.round().astype(int).clip(20, 80)

    # FV = OVR + age ceiling bonus
    ages = _gc(pool, "Age", 26.0)
    grades["FV"] = (ovr + ages.apply(fv_age_bonus)).round().astype(int).clip(20, 80)

    # Letter grades
    grades["Scout"]    = grades["OVR"].apply(grade_to_letter)
    grades["FV_Label"] = grades["FV"].apply(fv_label)

    # ── Plus-stats (100 = DB average) ────────────────────────────────────────
    xg_vals  = _gc(pool, "xG per 90")
    xa_vals  = _gc(pool, "xA per 90")
    g_vals   = _gc(pool, "Goals per 90")
    xp_vals  = xg_vals + xa_vals

    db_xg  = float(xg_vals.mean()) or 1e-9
    db_xa  = float(xa_vals.mean()) or 1e-9
    db_xp  = float(xp_vals.mean()) or 1e-9

    grades["xG_plus"] = (xg_vals / db_xg * 100).round(1)
    grades["xA_plus"] = (xa_vals / db_xa * 100).round(1)
    grades["xP_plus"] = (xp_vals / db_xp * 100).round(1)   # wRC+ analogue

    # GCONV: goal conversion vs expectation (100 = converting at xG rate)
    with np.errstate(divide="ignore", invalid="ignore"):
        gconv = np.where(xg_vals > 0.01, g_vals / xg_vals * 100, np.nan)
    grades["GCONV"] = pd.Series(gconv, index=pool.index).round(1)

    # WAR+ (WAR per 90 vs DB average)
    off, carry, total_rate = compute_war_rates(pool)
    repl = float(np.percentile(total_rate.dropna(), REPL_PERCENTILE))
    war90 = (total_rate - repl) / GOALS_PER_WIN
    db_war90_mean = float(war90.mean()) or 1e-9
    # shift so 0-WAR90 = 100, better = above 100
    grades["WAR_plus"] = ((war90 / db_war90_mean) * 100).round(1) \
                          if db_war90_mean > 0 else pd.Series(100.0, index=pool.index)

    return grades.reset_index(drop=True)


def pcts_to_grade_series(pcts: pd.Series) -> pd.Series:
    """Vectorised percentile → 20-80 conversion."""
    p = pcts.clip(0.1, 99.9) / 100
    z = p.apply(lambda x: norm.ppf(x))
    return (50 + z * 10).clip(20, 80).round().astype(int)


# ── Build everything ───────────────────────────────────────────────────────────

def build_all(pool: pd.DataFrame):
    # ── WAR ──────────────────────────────────────────────────────────────────
    off, carry, total = compute_war_rates(pool)
    repl       = float(np.percentile(total.dropna(), REPL_PERCENTILE))
    repl_off   = float(np.percentile(off.dropna(),   REPL_PERCENTILE))
    repl_carry = float(np.percentile(carry.dropna(), REPL_PERCENTILE))
    print(f"  Replacement rate = {repl:.4f}  "
          f"(off {repl_off:.4f}  carry {repl_carry:.4f})")

    minutes = _gc(pool, "Minutes played", 450.0)
    n90     = minutes / 90.0
    war     = (total - repl) * n90 / GOALS_PER_WIN

    res = pd.DataFrame({
        "Player":          pool["Player"],
        "Team":            pool.get("Team within selected timeframe",
                                    pool.get("Team", "")),
        "League":          pool["_league"],
        "Position":        pool["Position"],
        "Age":             _gc(pool, "Age", np.nan).replace(0, np.nan),
        "Minutes":         minutes.round(0).astype(int),
        "90s":             n90.round(2),
        "WAR":             war.round(3),
        "WAR_per_90":      ((total - repl) / GOALS_PER_WIN).round(4),
        "Offensive_rate":  off.round(4),
        "Carry_rate":      carry.round(4),
        "Total_rate":      total.round(4),
        "Rate_vs_repl":    (total - repl).round(4),
    })

    # DB rank & percentile
    war_arr = war.values
    res["DB_rank"]       = res["WAR"].rank(ascending=False, method="min").astype(int)
    res["DB_percentile"] = pd.Series(war_arr).apply(
        lambda w: round(float(np.mean(war_arr < w) * 100), 1) if pd.notna(w) else np.nan
    ).values

    # League rank & percentile
    res["Lg_rank"] = (res.groupby("League")["WAR"]
                        .rank(ascending=False, method="min").astype(int))
    res["Lg_n"]    = res.groupby("League")["WAR"].transform("count").astype(int)
    res["Lg_percentile"] = res.groupby("League")["WAR"].transform(
        lambda x: x.apply(lambda w: round(float((x < w).mean() * 100), 1)
                          if pd.notna(w) else np.nan))

    # ── Deduplicate (keep max-minutes per player+team) ────────────────────────
    before = len(res)
    dedup  = (res.sort_values("Minutes", ascending=False)
                 .drop_duplicates(subset=["Player", "Team"], keep="first"))
    keep_idx = dedup.index.tolist()          # original positions in res/pool
    res  = dedup.reset_index(drop=True)
    pool = pool.loc[keep_idx].reset_index(drop=True)
    removed = before - len(res)
    if removed:
        print(f"  Deduplicated: removed {removed} duplicate player-team rows")

    # ── Sort by WAR descending ────────────────────────────────────────────────
    order = res["WAR"].argsort()[::-1].values
    res  = res.iloc[order].reset_index(drop=True)
    pool = pool.iloc[order].reset_index(drop=True)
    res.insert(0, "Rank", range(1, len(res) + 1))

    print(f"\n  Top 10 by WAR:")
    for _, r in res.head(10).iterrows():
        print(f"    {int(r['Rank']):>4}.  {str(r['Player']):<28}  "
              f"{str(r['League']):<20}  WAR={r['WAR']:+.3f}")

    # ── Scouting grades ───────────────────────────────────────────────────────
    print("  Computing scouting grades …")
    grades = compute_scouting_grades(pool)

    # Merge player IDs back for the grades sheet
    grades_out = pd.concat([
        res[["Rank", "Player", "Team", "League", "Position", "Age",
             "Minutes", "WAR", "WAR_per_90"]].reset_index(drop=True),
        grades.reset_index(drop=True),
    ], axis=1)

    print(f"\n  Sample grades (top 5 by OVR):")
    for _, r in grades_out.nlargest(5, "OVR").iterrows():
        print(f"    {str(r['Player']):<28}  OVR={r['OVR']}  FV={r['FV']}  "
              f"({r['Scout']})  xP+={r['xP_plus']:.0f}")

    return res, grades_out, repl


# ── Excel formatting ───────────────────────────────────────────────────────────

def _col_width(df: pd.DataFrame, col: str) -> int:
    hdr = len(str(col)) + 2
    try:
        data = int(df[col].astype(str).str.len().max()) + 2
    except Exception:
        data = 10
    return max(hdr, min(data, 40))


def _style_sheet(wb, sheet_name: str, df: pd.DataFrame,
                 color_cols: list[tuple[str, str]] | None = None,
                 left_cols: list[int] | None = None):
    """Apply standard header + alternating rows + colour scales to a sheet."""
    if not HAS_OPX:
        return
    ws = wb[sheet_name]

    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1D4ED8")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c_align   = Alignment(horizontal="center", vertical="center")
    l_align   = Alignment(horizontal="left",   vertical="center")
    light     = PatternFill("solid", fgColor="F7F9FC")

    for cell in ws[1]:
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        for cell in row:
            cell.alignment = c_align
            if row_idx % 2 == 0:
                cell.fill = light
        for ci in (left_cols or [2, 3, 4, 5]):
            if ci <= len(row):
                row[ci - 1].alignment = l_align

    cols = list(df.columns)
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _col_width(df, col)

    for col_name, scale in (color_cols or []):
        if col_name not in cols:
            continue
        ci  = cols.index(col_name) + 1
        ltr = get_column_letter(ci)
        rng = f"{ltr}2:{ltr}{ws.max_row}"
        if scale == "rg":    # red → green
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="percentile", start_value=5,  start_color="DC2626",
                mid_type="percentile",   mid_value=50,   mid_color="D97706",
                end_type="percentile",   end_value=95,   end_color="16A34A",
            ))
        elif scale == "gr":  # green → red (e.g. rank: low = good)
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="percentile", start_value=5,  start_color="16A34A",
                mid_type="percentile",   mid_value=50,   mid_color="D97706",
                end_type="percentile",   end_value=95,   end_color="DC2626",
            ))


def save_excel(res: pd.DataFrame, grades_out: pd.DataFrame,
               repl: float, path: Path):

    method_rows = [
        ("WAR — Offensive rate",  "xG/90  +  xA/90×0.85  +  ShotAssists/90×0.28"),
        ("WAR — Carry rate",      "ProgRuns/90×0.025  +  (Dribbles/90×Drib%/100)×0.015"),
        ("WAR — Replacement",     f"{repl:.4f}  (15th-pctile total rate, all DB wide attackers ≥{MIN_MINS} min)"),
        ("WAR — Goals/win",       f"{GOALS_PER_WIN:.1f}"),
        ("WAR — Formula",         "(rate − replacement) × (minutes/90) ÷ goals_per_win"),
        ("", ""),
        ("20-80 Scale",           "Converts DB percentile via z-score: grade = 50 + norm.ppf(pct/100)×10, clipped [20,80]"),
        ("G_Finishing",           "xG×0.55 + Goals×0.25 + Shots×0.10 + BoxTouches×0.10  →  percentile → grade"),
        ("G_Dribbling",           "Dribbles/90 × DribSuccess%  (effective dribbles)  →  pctile → grade"),
        ("G_Athleticism",         "ProgRuns/90 + Accelerations/90×0.6  →  pctile → grade"),
        ("G_Crossing",            "Crosses/90 × CrossAcc%  (accurate crosses)  →  pctile → grade"),
        ("G_Creativity",          "xA×1.8 + ShotAssists×0.55 + KeyPasses×0.30  →  pctile → grade"),
        ("G_Defending",           "DefActions/90 + Interceptions/90  →  pctile → grade"),
        ("OVR",                   "Finish×20% + Dribble×25% + Athlete×20% + Cross×10% + Create×20% + Def×5%"),
        ("FV (Future Value)",     "OVR + age bonus: ≤18→+8  ≤20→+6  ≤22→+4  ≤24→+2  ≤27→0  ≤29→−2  ≤31→−4  32+→−7"),
        ("Scout",                 "Letter grade from OVR: A+≥75  A≥70  A-≥65  B+≥62  B≥57  B-≥53  C+≥51  C≥48  C-≥44  D+≥40  D≥35  F<35"),
        ("", ""),
        ("xG+",                   "player xG/90 ÷ DB-mean xG/90 × 100  (100 = average, like OPS+)"),
        ("xA+",                   "player xA/90 ÷ DB-mean xA/90 × 100"),
        ("xP+",                   "player (xG+xA)/90 ÷ DB-mean (xG+xA)/90 × 100  (wRC+ analogue)"),
        ("GCONV",                 "Goals/90 ÷ xG/90 × 100  (goal conversion vs expectation, 100 = neutral)"),
        ("WAR+",                  "WAR_per_90 ÷ DB-mean WAR_per_90 × 100"),
        ("Positions",             ", ".join(sorted(WIDE_ATK_POS))),
        ("Min. minutes",          f"≥{MIN_MINS}"),
    ]

    method_df = pd.DataFrame(method_rows, columns=["Parameter", "Definition"])

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        res.to_excel(writer, index=False, sheet_name="WAR Rankings")
        grades_out.to_excel(writer, index=False, sheet_name="Scouting Grades")
        method_df.to_excel(writer, index=False, sheet_name="Methodology")

        if HAS_OPX:
            war_cols  = [("WAR", "rg"), ("WAR_per_90", "rg"), ("DB_percentile", "rg"),
                         ("DB_rank", "gr"), ("Lg_rank", "gr")]
            _style_sheet(writer.book, "WAR Rankings", res, war_cols)

            grade_cols = [
                ("OVR", "rg"), ("FV", "rg"),
                ("G_Finishing", "rg"), ("G_Dribbling", "rg"), ("G_Athleticism", "rg"),
                ("G_Crossing", "rg"), ("G_Creativity", "rg"), ("G_Defending", "rg"),
                ("xG_plus", "rg"), ("xA_plus", "rg"), ("xP_plus", "rg"),
                ("GCONV", "rg"), ("WAR_plus", "rg"),
            ]
            _style_sheet(writer.book, "Scouting Grades", grades_out, grade_cols)

            ws3 = writer.book["Methodology"]
            hdr = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            fill = PatternFill("solid", fgColor="1D4ED8")
            for cell in ws3[1]:
                cell.font = hdr; cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws3.columns:
                w = max(len(str(c.value or "")) for c in col) + 4
                ws3.column_dimensions[col[0].column_letter].width = min(w, 90)

    print(f"\n  Saved → {path}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("Building WAR + Scouting Grades database …")
    pool                  = load_pool()
    res, grades_out, repl = build_all(pool)

    out = OUT_DIR / "WAR_All_Players.xlsx"
    save_excel(res, grades_out, repl, out)
    print(f"\n  {len(res):,} players  ·  {res['League'].nunique()} leagues")
    print(f"  Sheets: WAR Rankings  |  Scouting Grades  |  Methodology")


if __name__ == "__main__":
    main()
