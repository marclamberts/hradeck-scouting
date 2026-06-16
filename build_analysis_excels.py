"""
Generates two Excel files:
1. League Tiers.xlsx  — all leagues organised by tier
2. Team Ratings.xlsx  — team-level aggregated ratings with within-league and
                        cross-tier rankings
"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUT_DIR = Path("League Analysis")
OUT_DIR.mkdir(exist_ok=True)

# ── colour palette ─────────────────────────────────────────────────────────────
TIER_COLOURS = {
    1: "1A3A6B",   # deep navy   – Elite
    2: "1565C0",   # rich blue   – Top
    3: "2E7D32",   # forest green– Strong
    4: "E65100",   # deep orange – Developing
    5: "6A1E55",   # deep purple – Lower
    6: "4E342E",   # dark brown  – Youth/Grassroots
}
TIER_LIGHT = {
    1: "DDEEFF",
    2: "E3F2FD",
    3: "E8F5E9",
    4: "FFF3E0",
    5: "F3E5F5",
    6: "EFEBE9",
}
HEADER_BG  = "2C3E50"   # dark slate for column headers
WHITE      = "FFFFFF"
LIGHT_GREY = "F5F5F5"

def thin_border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

def header_style(ws, row, col_count, bg=HEADER_BG):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# FILE 1 – League Tiers
# ══════════════════════════════════════════════════════════════════════════════
def build_league_tiers():
    leagues = pd.read_excel("data/Leagues Overview.xlsx")
    leagues = leagues.sort_values(["Tier", "Country", "League Name"])

    tier_meta = (
        leagues[["Tier", "Tier Label"]]
        .drop_duplicates()
        .sort_values("Tier")
        .set_index("Tier")["Tier Label"]
        .to_dict()
    )

    out_path = OUT_DIR / "League Tiers.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # ── Summary sheet ──────────────────────────────────────────────────
        summary_rows = []
        for tier, label in tier_meta.items():
            sub = leagues[leagues["Tier"] == tier]
            summary_rows.append({
                "Tier": tier,
                "Tier Label": label,
                "League Count": len(sub),
                "Countries": sub["Country"].nunique(),
                "Total Players in DB": sub["Players in DB"].sum(),
            })
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        ws = writer.sheets["Summary"]
        ws.row_dimensions[1].height = 22
        header_style(ws, 1, len(summary_df.columns))
        for r in range(2, len(summary_df) + 2):
            tier_val = ws.cell(row=r, column=1).value
            bg = TIER_LIGHT.get(tier_val, WHITE)
            for c in range(1, len(summary_df.columns) + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(
                    bold=(c == 2),
                    color=TIER_COLOURS.get(tier_val, "000000"),
                    size=10,
                )
                cell.border = thin_border()
        set_col_widths(ws, [8, 20, 14, 12, 20])

        # ── All Leagues sheet ──────────────────────────────────────────────
        all_df = leagues[["Tier", "Tier Label", "League Name", "Country", "Division", "Players in DB"]].copy()
        all_df.to_excel(writer, sheet_name="All Leagues", index=False)

        ws = writer.sheets["All Leagues"]
        ws.row_dimensions[1].height = 22
        header_style(ws, 1, len(all_df.columns))
        for r in range(2, len(all_df) + 2):
            tier_val = ws.cell(row=r, column=1).value
            bg = TIER_LIGHT.get(tier_val, WHITE) if r % 2 == 0 else WHITE
            for c in range(1, len(all_df.columns) + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="left" if c in (3, 4) else "center",
                                           vertical="center")
                cell.font = Font(size=9)
                cell.border = thin_border()
        set_col_widths(ws, [6, 20, 38, 22, 12, 16])

        # ── One sheet per tier ─────────────────────────────────────────────
        for tier, label in tier_meta.items():
            sub = leagues[leagues["Tier"] == tier][
                ["League Name", "Country", "Division", "Players in DB", "Notes"]
            ].copy()
            safe_label = label.replace("/", "-")
            sheet_name = f"Tier {tier} - {safe_label}"[:31]
            sub.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            ws.row_dimensions[1].height = 22
            header_style(ws, 1, len(sub.columns), bg=TIER_COLOURS[tier])
            for r in range(2, len(sub) + 2):
                bg = TIER_LIGHT[tier] if r % 2 == 0 else WHITE
                for c in range(1, len(sub.columns) + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.alignment = Alignment(horizontal="left" if c == 1 else "center",
                                               vertical="center")
                    cell.font = Font(size=9)
                    cell.border = thin_border()
            set_col_widths(ws, [38, 22, 12, 16, 30])

    print(f"  Saved: {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# FILE 2 – Team Ratings
# ══════════════════════════════════════════════════════════════════════════════
TIER_WEIGHT = {
    "Tier 1": 1.00,   # Elite
    "Tier 2": 0.93,   # Top
    "Tier 3": 0.86,   # Strong
    "Tier 4": 0.79,   # Developing
    "Tier 5": 0.72,   # Lower
    "Tier 6": 0.65,   # Youth
}

SCORE_COLS = [
    "PlayerQualityScore",
    "ValueRecruitmentScore",
    "AttackingScore",
    "CreationScore",
    "DefendingScore",
    "PhysicalScore",
    "GoalsAddedScore",
    "xThreatScore",
    "ActionValueScore",
    "SuccessProbability",
]

def build_team_ratings():
    rec = pd.read_excel("data/FCHK Model V3 - Recruitment Scores.xlsx")

    # Minutes-weighted average for each score column
    rec["_w"] = rec["MinutesPlayed"].clip(lower=1)
    for col in SCORE_COLS:
        rec[f"_wt_{col}"] = rec[col] * rec["_w"]

    grp = rec.groupby(["TeamName", "LeagueLabel", "CountryLabel", "TierLabel"])
    agg = grp.agg(
        Players=("PlayerName" if "PlayerName" in rec.columns else "_w", "count"),
        TotalMinutes=("_w", "sum"),
        **{col: (f"_wt_{col}", "sum") for col in SCORE_COLS},
    ).reset_index()

    # Divide weighted sums by total minutes to get weighted averages
    for col in SCORE_COLS:
        agg[col] = agg[col] / agg["TotalMinutes"]

    # Derive tier number from label for sorting / colour
    agg["TierNum"] = agg["TierLabel"].str.extract(r"(\d+)").astype(float).astype("Int64")

    # Raw team score = average of all score columns
    agg["RawTeamScore"] = agg[SCORE_COLS].mean(axis=1)

    # Tier-adjusted score (enables cross-league comparison)
    agg["TierWeight"] = agg["TierLabel"].map(TIER_WEIGHT).fillna(0.65)
    agg["TierAdjustedScore"] = agg["RawTeamScore"] * agg["TierWeight"]

    # Within-league rank (by raw score – fairer for intra-league)
    agg["LeagueRank"] = (
        agg.groupby("LeagueLabel")["RawTeamScore"]
        .rank(method="min", ascending=False)
        .astype(int)
    )

    # Overall cross-tier rank (by tier-adjusted score)
    agg["OverallRank"] = (
        agg["TierAdjustedScore"]
        .rank(method="min", ascending=False)
        .astype(int)
    )

    agg = agg.sort_values(["TierNum", "LeagueLabel", "LeagueRank"])

    display_cols = [
        "OverallRank", "LeagueRank", "TeamName", "LeagueLabel", "CountryLabel",
        "TierLabel", "Players", "TotalMinutes",
        "RawTeamScore", "TierAdjustedScore",
        "PlayerQualityScore", "ValueRecruitmentScore",
        "AttackingScore", "CreationScore", "DefendingScore",
        "PhysicalScore", "GoalsAddedScore", "xThreatScore",
        "ActionValueScore", "SuccessProbability",
    ]
    out_df = agg[display_cols].copy()

    # Round floats
    float_cols = [c for c in display_cols if c not in
                  ("OverallRank","LeagueRank","TeamName","LeagueLabel",
                   "CountryLabel","TierLabel","Players")]
    out_df[float_cols] = out_df[float_cols].round(1)

    out_path = OUT_DIR / "Team Ratings.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # ── All Teams sheet ────────────────────────────────────────────────
        out_df.sort_values("OverallRank").to_excel(
            writer, sheet_name="All Teams (Overall Rank)", index=False
        )
        _style_team_sheet(writer, "All Teams (Overall Rank)", out_df, by="overall")

        # ── Per-tier sheets ────────────────────────────────────────────────
        for tier_num in sorted(out_df["TierLabel"].str.extract(r"(\d+)")[0].unique()):
            tier_label = f"Tier {tier_num}"
            sub = out_df[out_df["TierLabel"] == tier_label].sort_values(
                ["LeagueLabel", "LeagueRank"]
            )
            if sub.empty:
                continue
            # pull readable label from weight dict
            sheet_name = f"{tier_label}"[:31]
            sub.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_team_sheet(writer, sheet_name, sub, tier_num=int(tier_num))

        # ── Per-league sheets ──────────────────────────────────────────────
        for league in sorted(out_df["LeagueLabel"].unique()):
            sub = out_df[out_df["LeagueLabel"] == league].sort_values("LeagueRank")
            sname = league[:31]
            sub.to_excel(writer, sheet_name=sname, index=False)
            tier_num = int(sub["TierLabel"].iloc[0].split()[-1])
            _style_team_sheet(writer, sname, sub, tier_num=tier_num)

    print(f"  Saved: {out_path}")


def _style_team_sheet(writer, sheet_name, df, by="tier", tier_num=None):
    ws = writer.sheets[sheet_name]
    n_cols = len(df.columns)
    ws.row_dimensions[1].height = 32
    header_style(ws, 1, n_cols)

    for r in range(2, len(df) + 2):
        # Determine tier for this row
        tier_label_val = ws.cell(row=r, column=6).value    # TierLabel is col 6
        try:
            tnum = int(str(tier_label_val).split()[-1]) if tier_label_val else (tier_num or 1)
        except Exception:
            tnum = tier_num or 1

        row_bg = TIER_LIGHT.get(tnum, WHITE) if r % 2 == 0 else WHITE

        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=row_bg)
            cell.border = thin_border()
            # Score columns (9 onwards) get centred numeric alignment
            if c <= 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True, size=9, color=TIER_COLOURS.get(tnum, "000000"))
            elif c == 3:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(bold=True, size=9)
            elif c in (4, 5, 6):
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(size=9)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=9)

    # Column widths
    widths = [10, 10, 26, 30, 16, 10, 8, 14,
              14, 16,
              16, 18,
              14, 14, 14,
              14, 14, 12,
              14, 16]
    set_col_widths(ws, widths[:n_cols])

    # Freeze header row
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    print("Building League Tiers.xlsx ...")
    build_league_tiers()

    print("Building Team Ratings.xlsx ...")
    build_team_ratings()

    print("Done. Files saved to:", OUT_DIR.resolve())
