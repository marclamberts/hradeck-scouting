"""
attacker_outliers_per_tier.py
─────────────────────────────
Finds the most positive statistical outliers among attackers, broken down
by the Leagues Overview tier system (Tier 1 Elite → Tier 5 Lower).

Output: Player Profiles/Attacker Outliers by Tier.xlsx
"""
from __future__ import annotations
from pathlib import Path
import numpy as np
import pandas as pd
from scipy.stats import norm
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH   = Path("Player Profiles/Attacker Outliers by Tier.xlsx")
WYSCOUT_DIR = Path("data/Wyscout DB")
MIN_MINUTES = 400

# ── Wyscout stem → (Country label for join, Division number) ─────────────────
# Multi-part files collapse to a single stem
STEM_COUNTRY_DIV: dict[str, tuple[str, int]] = {
    # ── Tier 1 (Elite 5) ──────────────────────────────────────────────────
    "England":        ("England",      1),
    "Spain":          ("Spain",        1),
    "Germany":        ("Germany",      1),
    "Italy":          ("Italy",        1),
    "France":         ("France",       1),
    # ── Tier 2 (Top) ──────────────────────────────────────────────────────
    "Argentina":      ("Argentina",    1),
    "Brazil":         ("Brazil",       1),
    "Mexico":         ("Mexico",       1),
    "Netherlands":    ("Netherlands",  1),
    "Belgium":        ("Belgium",      1),
    "Turkiye":        ("Türkiye",      1),
    "Portugal":       ("Portugal",     1),
    "Russia":         ("Russia",       1),
    "Ukraine":        ("Ukraine",      1),
    "Scotland":       ("Scotland",     1),
    "Japan":          ("Japan",        1),
    "USA":            ("USA",          1),
    "Korea":          ("South Korea",  1),
    "Saudi":          ("Saudi Arabia", 1),
    "China":          ("China",        1),
    "Qatar":          ("Qatar",        1),
    "UAE":            ("UAE",          1),
    # ── Tier 3 (Strong) ───────────────────────────────────────────────────
    "England II":     ("England",      2),
    "Spain II":       ("Spain",        2),
    "Germany II":     ("Germany",      2),
    "Italy II":       ("Italy",        2),
    "France II":      ("France",       2),
    "Austria":        ("Austria",      1),
    "Switzerland":    ("Switzerland",  1),
    "Denmark":        ("Denmark",      1),
    "Sweden":         ("Sweden",       1),
    "Norway":         ("Norway",       1),
    "Poland":         ("Poland",       1),
    "Czech":          ("Czech Republic", 1),
    "Greece":         ("Greece",       1),
    "Croatia":        ("Croatia",      1),
    "Romania":        ("Romania",      1),
    "Serbia":         ("Serbia",       1),
    "Colombia":       ("Colombia",     1),
    "Chile":          ("Chile",        1),
    "Peru":           ("Peru",         1),
    "Uruguay":        ("Uruguay",      1),
    "Australia":      ("Australia",    1),
    "Egypt":          ("Egypt",        1),
    "Morocco":        ("Morocco",      1),
    "Nigeria":        ("Nigeria",      1),
    "South Africa":   ("South Africa", 1),
    "India":          ("India",        1),
    "Indonesia":      ("Indonesia",    1),
    "Malaysia":       ("Malaysia",     1),
    "Thailand":       ("Thailand",     1),
    "Canada":         ("Canada",       1),
    # ── Tier 4 (Developing) ───────────────────────────────────────────────
    "England III":    ("England",      3),
    "Spain III":      ("Spain",        3),
    "Germany III":    ("Germany",      3),
    "France III":     ("France",       3),
    "Austria II":     ("Austria",      2),
    "Switzerland II": ("Switzerland",  2),
    "Denmark II":     ("Denmark",      2),
    "Sweden II":      ("Sweden",       2),
    "Norway II":      ("Norway",       2),
    "Poland II":      ("Poland",       2),
    "Czech II":       ("Czech Republic", 2),
    "Greece II":      ("Greece",       2),
    "Russia II":      ("Russia",       2),
    "Ukraine II":     ("Ukraine",      2),
    "Scotland II":    ("Scotland",     2),
    "Japan II III":   ("Japan",        2),
    "USA II":         ("USA",          2),
    "Korea II":       ("South Korea",  2),
    "Saudi II":       ("Saudi Arabia", 2),
    "China II":       ("China",        2),
    "Portugal II":    ("Portugal",     2),
    "Belgium II":     ("Belgium",      2),
    "Turkiye II":     ("Türkiye",      2),
    "Netherlands II": ("Netherlands",  2),
    "Serbia II":      ("Serbia",       2),
    "Argentina II":   ("Argentina",    2),
    "Brazil II":      ("Brazil",       2),
    "Mexico II":      ("Mexico",       2),
    "Chile II":       ("Chile",        2),
    "Uruguay II":     ("Uruguay",      2),
    "Ecuador":        ("Ecuador",      1),
    "Ecuador II":     ("Ecuador",      2),
    "Albania":        ("Albania",      1),
    "Andorra":        ("Andorra",      1),
    "Armenia":        ("Armenia",      1),
    "Azerbaijan":     ("Azerbaijan",   1),
    "Bahrain":        ("Bahrain",      1),
    "Bangladesh":     ("Bangladesh",   1),
    "Bolivia":        ("Bolivia",      1),
    "Bosnia":         ("Bosnia",       1),
    "Bulgaria":       ("Bulgaria",     1),
    "Cambodia":       ("Cambodia",     1),
    "Costa Rica":     ("Costa Rica",   1),
    "Cyprus":         ("Cyprus",       1),
    "Finland":        ("Finland",      1),
    "Georgia":        ("Georgia",      1),
    "Guatemala":      ("Guatemala",    1),
    "Honduras":       ("Honduras",     1),
    "Hong Kong":      ("Hong Kong",    1),
    "Hungary":        ("Hungary",      1),
    "Iceland":        ("Iceland",      1),
    "Ireland":        ("Ireland",      1),
    "Jordan":         ("Jordan",       1),
    "Kazakhstan":     ("Kazakhstan",   1),
    "Kosovo":         ("Kosovo",       1),
    "Kyrgystan":      ("Kyrgyzstan",   1),
    "Latvia":         ("Latvia",       1),
    "Lithuania":      ("Lithuania",    1),
    "Malta":          ("Malta",        1),
    "Moldovia":       ("Moldova",      1),
    "Montenegro":     ("Montenegro",   1),
    "Nicaragua":      ("Nicaragua",    1),
    "Northern Ireland": ("Northern Ireland", 1),
    "Panama":         ("Panama",       1),
    "Paraguay":       ("Paraguay",     1),
    "Philippines":    ("Philippines",  1),
    "Singapore":      ("Singapore",    1),
    "Slovakia":       ("Slovakia",     1),
    "Slovenia":       ("Slovenia",     1),
    "Tunisia":        ("Tunisia",      1),
    "Vietnam":        ("Vietnam",      1),
    "Wales":          ("Wales",        1),
    "Uzbekistan":     ("Uzbekistan",   1),
    "Venezuela":      ("Venezuela",    1),
    "El Salvador":    ("El Salvador",  1),
    "Estonia":        ("Estonia",      1),       # not in overview → Tier 4
    "Faroe Islands":  ("Faroe Islands",1),       # not in overview → Tier 4
    # ── Tier 5 (Lower) ────────────────────────────────────────────────────
    "England IV":     ("England",      4),
    "England V":      ("England",      5),
    "Norway III":     ("Norway",       3),
    "Denmark III":    ("Denmark",      3),
    "Denmark IV":     ("Denmark",      4),
    "Poland III":     ("Poland",       3),
    "Portugal III":   ("Portugal",     3),
    "Netherlands III":("Netherlands",  3),
    "Sweden III":     ("Sweden",       3),
    "Korea III":      ("South Korea",  3),
    "Scotland III":   ("Scotland",     3),
    "Scotland IV":    ("Scotland",     4),
    "Cyprus II":      ("Cyprus",       2),
    "Hungary II":     ("Hungary",      2),
    "Ireland II":     ("Ireland",      2),
    "Slovakia II":    ("Slovakia",     2),
    "Slovenia II":    ("Slovenia",     2),
    "Finland II":     ("Finland",      2),
    "USA III":        ("USA",          3),
    # ── Multi-part (collapsed) ────────────────────────────────────────────
    "Italy III - Part I":   ("Italy", 3),
    "Italy III - Part II":  ("Italy", 3),
    "Italy III - Part III": ("Italy", 3),
    "Italy III - Part IV":  ("Italy", 3),
    "Germany 4 - Part I":   ("Germany", 4),
    "Germany 4 - Part II":  ("Germany", 4),
    "Germany 4 - Part III": ("Germany", 4),
    "Germany 4 - Part IV":  ("Germany", 4),
    "Australia II - Part I":   ("Australia", 2),
    "Australia II - Part II":  ("Australia", 2),
    "Australia II - Part III": ("Australia", 2),
    "Australia II - Part IV":  ("Australia", 2),
    "Australia II - Part V":   ("Australia", 2),
    "Australia II - Part VI":  ("Australia", 2),
    "Australia II - Part VII": ("Australia", 2),
    # ── Youth/Grassroots ──────────────────────────────────────────────────
    "Czech U17": ("Czech Republic", 99),
    "Czech U19": ("Czech Republic", 99),
}

POSITION_FAMILY: dict[str, str] = {
    "CF": "Central attacker", "SS": "Central attacker",
    "LW": "Wide attacker", "RW": "Wide attacker",
    "LWF": "Wide attacker", "RWF": "Wide attacker", "WF": "Wide attacker",
    "AMF": "Advanced midfielder", "LAMF": "Advanced midfielder", "RAMF": "Advanced midfielder",
    "CMF": "Deep midfielder", "LCM": "Deep midfielder", "RCM": "Deep midfielder",
    "LCMF": "Deep midfielder", "RCMF": "Deep midfielder",
    "DMF": "Deep midfielder", "LDM": "Deep midfielder", "RDM": "Deep midfielder",
    "LDMF": "Deep midfielder", "RDMF": "Deep midfielder",
    "LB": "Wide defender", "RB": "Wide defender", "LWB": "Wide defender", "RWB": "Wide defender",
    "CB": "Central defender", "LCB": "Central defender", "RCB": "Central defender",
    "GK": "Goalkeeper",
}
ATTACKER_FAMILIES = {"Central attacker", "Wide attacker"}

NON_METRIC = {
    "Player", "Team", "Team within selected timeframe", "Position",
    "Age", "Market value", "Contract expires", "Birth country",
    "Passport country", "Foot", "Height", "Weight", "On loan",
}

ATTACK_METRICS = [
    "Goals per 90", "xG per 90", "Non-penalty goals per 90",
    "Assists per 90", "xA per 90", "Shot assists per 90",
    "Shots per 90", "Shots on target, %", "Goal conversion, %",
    "Touches in box per 90", "Key passes per 90", "Smart passes per 90",
    "Progressive runs per 90", "Dribbles per 90", "Successful dribbles, %",
    "Aerial duels won, %", "Crosses per 90", "Accurate crosses, %",
    "Offensive duels won, %", "Fouls suffered per 90",
]

DISPLAY_COLS = [
    "Player", "Team", "League", "Country", "Tier", "TierLabel",
    "Position", "PositionFamily", "Age", "Minutes played",
    "Goals per 90", "xG per 90", "Assists per 90", "xA per 90",
    "Shots per 90", "Touches in box per 90", "Key passes per 90",
    "Dribbles per 90", "Successful dribbles, %",
    "Aerial duels won, %", "Crosses per 90",
    "_peak_z", "_mean_z", "_anomaly_breadth", "_anomaly_score", "_anomaly_type",
]

TIER_THEMES = {
    1: {"header": "1A237E", "light": "E8EAF6", "label": "Elite"},
    2: {"header": "1565C0", "light": "E3F2FD", "label": "Top"},
    3: {"header": "2E7D32", "light": "E8F5E9", "label": "Strong"},
    4: {"header": "E65100", "light": "FFF3E0", "label": "Developing"},
    5: {"header": "6A1E55", "light": "F3E5F5", "label": "Lower"},
    6: {"header": "4E342E", "light": "EFEBE9", "label": "Youth"},
}
WHITE = "FFFFFF"

def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_row(ws, row, ncols, bg):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin()
    ws.row_dimensions[row].height = 28

def _data_row(ws, r, ncols, bg, peak_col_idx=None, row_data=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(
            horizontal="left" if c <= 4 else "center", vertical="center"
        )
        cell.font = Font(size=9, bold=(c == 1))
        cell.border = _thin()
    # Colour peak_z cell
    if peak_col_idx and row_data is not None:
        try:
            pz = float(row_data[peak_col_idx - 1])
            colour = "E53935" if pz >= 3.0 else ("FF7043" if pz >= 2.5 else "FB8C00")
            ws.cell(row=r, column=peak_col_idx).font = Font(
                size=9, bold=True, color=colour
            )
        except Exception:
            pass

def _set_widths(ws, df):
    for i, col in enumerate(df.columns, 1):
        w = max(len(str(col)) * 1.05, 10.0)
        if col in ("Player", "Team", "League"):
            w = max(w, 22.0)
        elif col in ("_anomaly_type", "PositionFamily", "TierLabel"):
            w = max(w, 18.0)
        ws.column_dimensions[get_column_letter(i)].width = min(w, 28.0)

# ── Load & enrich ─────────────────────────────────────────────────────────────

def load_attackers() -> pd.DataFrame:
    leagues_ov = pd.read_excel("data/Leagues Overview.xlsx")

    # Build: (Country, Division) → (Tier, TierLabel)
    tier_map: dict[tuple[str, int], tuple[int, str]] = {}
    for _, row in leagues_ov.iterrows():
        div = row["Division"]
        if isinstance(div, int) or (isinstance(div, str) and str(div).isdigit()):
            tier_map[(str(row["Country"]).strip(), int(div))] = (
                int(row["Tier"]), str(row["Tier Label"])
            )

    # Default fallback: small nation top leagues → Tier 4
    def _lookup_tier(stem: str) -> tuple[int, str]:
        if stem not in STEM_COUNTRY_DIV:
            return (4, "Developing")
        country, div = STEM_COUNTRY_DIV[stem]
        if div == 99:
            return (6, "Youth/Grassroots")
        result = tier_map.get((country, div))
        if result:
            return result
        # fallback for countries not in overview (Estonia, Faroe Islands, etc.)
        return (4, "Developing")

    files = sorted(WYSCOUT_DIR.glob("*.xlsx"))
    print(f"  Loading {len(files)} files …")
    frames = []
    for path in files:
        tier, tier_label = _lookup_tier(path.stem)
        if tier == 6:
            continue  # skip youth
        try:
            df = pd.read_excel(path)
            df.columns = [str(c).strip() for c in df.columns]
            if "Position" in df.columns:
                df["Position"] = (
                    df["Position"].astype(str).str.split(",").str[0].str.strip()
                )
            df["League"]    = path.stem
            df["Tier"]      = tier
            df["TierLabel"] = tier_label
            # Country from the mapping
            country = STEM_COUNTRY_DIV.get(path.stem, ("?", 0))[0]
            df["Country"]   = country
            frames.append(df)
        except Exception as e:
            print(f"    [skip] {path.stem}: {e}")

    df = pd.concat(frames, ignore_index=True)

    # Coerce numerics
    for col in df.columns:
        if col not in NON_METRIC | {"League", "Tier", "TierLabel", "Country"}:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Minutes filter
    df = df.loc[df.get("Minutes played", pd.Series(0, index=df.index)).fillna(0) >= MIN_MINUTES]

    # Keep only attackers
    df["PositionFamily"] = df["Position"].map(POSITION_FAMILY).fillna("Other")
    df = df.loc[df["PositionFamily"].isin(ATTACKER_FAMILIES)].copy()
    print(f"  → {len(df):,} attackers across {df['Tier'].nunique()} tiers")
    return df.reset_index(drop=True)


def compute_tier_anomalies(df: pd.DataFrame) -> pd.DataFrame:
    """Z-score within each tier and return anomaly-flagged rows."""
    metrics = [m for m in ATTACK_METRICS if m in df.columns]
    all_frames = []

    for tier in sorted(df["Tier"].unique()):
        sub = df.loc[df["Tier"] == tier].copy()
        if len(sub) < 10:
            continue

        X   = sub[metrics].fillna(0).values.astype(float)
        mu  = X.mean(axis=0)
        sig = X.std(axis=0)
        sig = np.where(sig == 0, 1e-9, sig)
        Z   = (X - mu) / sig

        for i, m in enumerate(metrics):
            sub[f"_z_{m}"] = Z[:, i]

        z_cols = [f"_z_{m}" for m in metrics]
        sub["_peak_z"]          = sub[z_cols].max(axis=1)
        sub["_mean_z"]          = sub[z_cols].mean(axis=1).round(3)
        sub["_anomaly_breadth"] = (sub[z_cols] >= 1.6).sum(axis=1)
        sub["_anomaly_score"]   = (
            0.45 * sub["_peak_z"].clip(lower=0)
            + 0.35 * sub["_anomaly_breadth"]
            + 0.20 * sub["_mean_z"].clip(lower=0)
        )

        def _classify(row) -> str:
            pz = float(row["_peak_z"])
            br = int(row["_anomaly_breadth"])
            age = float(pd.to_numeric(row.get("Age"), errors="coerce") or 99)
            if br >= 6:     return "Multi-dimensional"
            if pz >= 3.5:   return "Exceptional"
            if pz >= 2.5 and br <= 2: return "Specialist Elite"
            if pz >= 1.8 and age <= 21: return "Age-adjusted Gem"
            if br >= 3:     return "Consistent Overperformer"
            return "Emerging Talent"

        sub["_anomaly_type"] = sub.apply(_classify, axis=1)
        all_frames.append(sub)

    return pd.concat(all_frames, ignore_index=True) if all_frames else pd.DataFrame()


def build_excel(df_anomalies: pd.DataFrame):
    avail = [c for c in DISPLAY_COLS if c in df_anomalies.columns]

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:

        # ── Summary sheet ─────────────────────────────────────────────────────
        summary_rows = []
        for tier in sorted(df_anomalies["Tier"].unique()):
            sub = df_anomalies.loc[df_anomalies["Tier"] == tier]
            label = sub["TierLabel"].iloc[0] if len(sub) else ""
            top3  = sub.nlargest(3, "_anomaly_score")[["Player", "Team", "League", "_peak_z", "_anomaly_type"]]
            for rank, (_, row) in enumerate(top3.iterrows(), 1):
                summary_rows.append({
                    "Tier": tier,
                    "Tier Label": label,
                    "Total Outliers": len(sub),
                    "Rank in Tier": rank,
                    "Player": row["Player"],
                    "Team": row["Team"],
                    "League": row["League"],
                    "Peak Z": round(float(row["_peak_z"]), 2),
                    "Type": row["_anomaly_type"],
                })
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]
        _header_row(ws, 1, len(summary_df.columns), "1A1A2E")
        for r_idx, (_, row) in enumerate(summary_df.iterrows(), 2):
            tier = int(row["Tier"]) if pd.notna(row.get("Tier")) else 4
            theme = TIER_THEMES.get(tier, TIER_THEMES[4])
            bg = theme["light"] if r_idx % 2 == 0 else WHITE
            _data_row(ws, r_idx, len(summary_df.columns), bg)
        _set_widths(ws, summary_df)
        ws.freeze_panes = "A2"

        # ── Per-tier sheets ───────────────────────────────────────────────────
        for tier in sorted(df_anomalies["Tier"].unique()):
            sub = df_anomalies.loc[df_anomalies["Tier"] == tier].copy()
            if sub.empty:
                continue
            tier_label = sub["TierLabel"].iloc[0]
            sheet_name = f"Tier {tier} - {tier_label}"[:31]
            theme = TIER_THEMES.get(tier, TIER_THEMES[4])

            disp = sub[avail].sort_values("_anomaly_score", ascending=False).reset_index(drop=True)
            # Round floats
            num_cols = [c for c in disp.columns if disp[c].dtype.kind == "f"]
            disp[num_cols] = disp[num_cols].round(3)
            disp.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            _header_row(ws, 1, len(disp.columns), theme["header"])

            peak_idx = disp.columns.tolist().index("_peak_z") + 1 if "_peak_z" in disp.columns else None
            for r_idx, row_vals in enumerate(disp.itertuples(index=False), 2):
                bg = theme["light"] if r_idx % 2 == 0 else WHITE
                vals = list(row_vals)
                _data_row(ws, r_idx, len(disp.columns), bg, peak_idx, vals)
            _set_widths(ws, disp)
            ws.freeze_panes = "A2"

        # ── Combined all-tier sheet (sorted by anomaly score) ─────────────────
        all_disp = df_anomalies[avail].sort_values(
            ["_peak_z"], ascending=False
        ).reset_index(drop=True)
        num_cols = [c for c in all_disp.columns if all_disp[c].dtype.kind == "f"]
        all_disp[num_cols] = all_disp[num_cols].round(3)
        all_disp.to_excel(writer, sheet_name="All Tiers", index=False)
        ws = writer.sheets["All Tiers"]
        _header_row(ws, 1, len(all_disp.columns), "2C3E50")
        peak_idx = all_disp.columns.tolist().index("_peak_z") + 1 if "_peak_z" in all_disp.columns else None
        for r_idx, row_vals in enumerate(all_disp.itertuples(index=False), 2):
            tier_val = int(row_vals[all_disp.columns.tolist().index("Tier")]) if "Tier" in all_disp.columns else 4
            theme = TIER_THEMES.get(tier_val, TIER_THEMES[4])
            bg = theme["light"] if r_idx % 2 == 0 else WHITE
            _data_row(ws, r_idx, len(all_disp.columns), bg, peak_idx, list(row_vals))
        _set_widths(ws, all_disp)
        ws.freeze_panes = "A2"

    print(f"  Saved → {OUT_PATH}")


def main():
    print("Loading attacker data …")
    df = load_attackers()

    print("Computing tier-level z-score anomalies …")
    df_anom = compute_tier_anomalies(df)
    # Keep only outliers (peak_z ≥ 1.8)
    df_anom = df_anom.loc[df_anom["_peak_z"] >= 1.8].copy()
    print(f"  → {len(df_anom):,} outlier rows across all tiers")

    print("Building Excel …")
    build_excel(df_anom)

    # ── Print top 5 per tier to terminal ─────────────────────────────────────
    print("\n" + "═" * 80)
    print("  TOP ATTACKER OUTLIERS BY TIER")
    print("═" * 80)
    for tier in sorted(df_anom["Tier"].unique()):
        sub = df_anom.loc[df_anom["Tier"] == tier]
        label = sub["TierLabel"].iloc[0]
        top = sub.nlargest(5, "_anomaly_score")[
            ["Player", "Team", "League", "Position", "Age",
             "Goals per 90", "xG per 90", "_peak_z", "_anomaly_type"]
        ].reset_index(drop=True)
        top.index += 1
        print(f"\n  Tier {tier} — {label}  ({len(sub)} outliers total)")
        print(top.round(3).to_string())
    print("\n" + "═" * 80)


if __name__ == "__main__":
    main()
