"""
build_lamberts_total.py
───────────────────────
Build a Lamberts Index Model workbook from raw Wyscout files.

Mirrors the structure of Lamberts_Index_Model__Scouting_CZ__SVK.xlsx:
  README · Priority List · Elite Picks · Lamberts Analysis ·
  Top 5 per Position · All Targets · GK · CB · FB · DM · CM · W · FW ·
  Expiring 2026 · Budget Planner · Squad

Usage:
  python build_lamberts_total.py
  python build_lamberts_total.py --leagues "Czech II" Slovakia "Slovakia II" --min-minutes 500
  python build_lamberts_total.py --output data/My_Model.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers as xl_numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Config ─────────────────────────────────────────────────────────────────────

ROOT = Path(__file__).parent
WYSCOUT_DIR = ROOT / "Wyscout Files"

SKIP_FILES = {"FCHK Model V3 - Loaded Leagues", "FCHK Model V3 - Model Input",
               "FCHK Model V3 - Player Scores", "FCHK Model V3 - Player Styles",
               "FCHK Model V3 - Recruitment Scores", "FCHK Model V3 - Smart Club Closeness",
               "FCHK Model V3 - Summary", "FCHK Model V3 Scores", "FCHK Scouting Report",
               "Leagues Overview", "Wyscout Anomaly Report", "Wyscout Full Scouting Report",
               "FCHK Model V3 Scores"}

DEFAULT_LEAGUES = None  # None = all leagues in WYSCOUT_DIR
DEFAULT_MIN_MINUTES = 400
DEFAULT_MAX_AGE = 30
DEFAULT_BUDGET = 1_000_000

# Position mapping (first token of Wyscout position string → group)
POS_MAP: dict[str, str] = {
    "CF": "FW", "SS": "FW",
    "LW": "W",  "RW": "W",  "LWF": "W",  "RWF": "W",  "WF": "W",
    "LAMF": "W", "RAMF": "W",
    "AMF": "CM",
    "CMF": "CM", "LCM": "CM", "RCM": "CM", "LCMF": "CM", "RCMF": "CM",
    "DMF": "DM", "LDM": "DM", "RDM": "DM", "LDMF": "DM", "RDMF": "DM",
    "LB": "FB",  "RB": "FB",  "LWB": "FB", "RWB": "FB",
    "CB": "CB",  "LCB": "CB", "RCB": "CB",
    "GK": "GK",
}

# Per-position SQS metrics (weights)
SQS_BLUEPRINTS: dict[str, list[tuple[str, float]]] = {
    "GK": [
        ("Save rate, %", 4.0),
        ("Prevented goals per 90", 3.0),
        ("Exits per 90", 2.0),
        ("Aerial duels per 90", 1.5),
        ("Accurate passes, %", 1.5),
        ("Accurate long passes, %", 1.0),
    ],
    "CB": [
        ("Successful defensive actions per 90", 3.0),
        ("Defensive duels won, %", 2.5),
        ("Aerial duels won, %", 2.0),
        ("Interceptions per 90", 2.0),
        ("PAdj Interceptions", 1.5),
        ("Shots blocked per 90", 1.0),
        ("Accurate passes, %", 1.0),
        ("Progressive passes per 90", 1.0),
    ],
    "FB": [
        ("Crosses per 90", 2.0),
        ("Accurate crosses, %", 1.5),
        ("xA per 90", 2.0),
        ("Assists per 90", 1.5),
        ("Progressive runs per 90", 1.5),
        ("Dribbles per 90", 1.0),
        ("Successful defensive actions per 90", 2.0),
        ("Defensive duels won, %", 2.0),
        ("Aerial duels won, %", 1.0),
        ("Progressive passes per 90", 1.5),
    ],
    "DM": [
        ("Successful defensive actions per 90", 3.0),
        ("Defensive duels won, %", 2.5),
        ("Interceptions per 90", 2.5),
        ("PAdj Interceptions", 2.0),
        ("Aerial duels won, %", 1.5),
        ("Passes per 90", 1.5),
        ("Accurate passes, %", 1.5),
        ("Progressive passes per 90", 1.5),
    ],
    "CM": [
        ("Passes per 90", 2.0),
        ("Accurate passes, %", 2.0),
        ("Progressive passes per 90", 2.5),
        ("Key passes per 90", 2.5),
        ("xA per 90", 2.0),
        ("Assists per 90", 1.5),
        ("Progressive runs per 90", 1.5),
        ("Successful defensive actions per 90", 1.5),
        ("Goals per 90", 1.5),
    ],
    "W": [
        ("Goals per 90", 2.5),
        ("xG per 90", 2.0),
        ("Assists per 90", 2.0),
        ("xA per 90", 2.0),
        ("Dribbles per 90", 2.0),
        ("Successful dribbles, %", 1.5),
        ("Progressive runs per 90", 2.0),
        ("Touches in box per 90", 2.0),
        ("Key passes per 90", 1.5),
    ],
    "FW": [
        ("Goals per 90", 4.0),
        ("xG per 90", 3.0),
        ("Non-penalty goals per 90", 2.5),
        ("Shots per 90", 1.5),
        ("Shots on target, %", 1.5),
        ("Goal conversion, %", 1.5),
        ("Touches in box per 90", 1.5),
        ("Aerial duels won, %", 1.5),
        ("Dribbles per 90", 1.0),
        ("xA per 90", 1.0),
    ],
}

# Hradec Králové squad quality (Czech top-flight percentile ranks)
HRADEC_SQUAD: dict[str, float] = {
    "GK": 61.5,   # starter Quality %
    "CB": 33.3,   # weakest starter
    "FB": 81.8,
    "DM": 36.8,
    "CM": 53.7,   # median starter
    "W":  27.6,   # weakest starter
    "FW": 38.3,
}

# ── Colors ─────────────────────────────────────────────────────────────────────
C = {
    "navy":    "0D1B2A",
    "gold":    "C9A84C",
    "elite":   "1A5276",
    "high":    "1E8449",
    "value":   "117A65",
    "fair":    "626567",
    "over":    "922B21",
    "white":   "FFFFFF",
    "light":   "EBF5FB",
    "header":  "154360",
    "upgrade": "1A5276",
    "depth":   "117A65",
    "rota":    "7D6608",
}


# ── Data loading ───────────────────────────────────────────────────────────────

def load_leagues(leagues: list[str] | None, min_minutes: int) -> pd.DataFrame:
    if leagues is None:
        paths = sorted(p for p in WYSCOUT_DIR.glob("*.xlsx")
                       if p.stem not in SKIP_FILES)
    else:
        paths = [WYSCOUT_DIR / f"{lg}.xlsx" for lg in leagues]

    frames: list[pd.DataFrame] = []
    for path in paths:
        if not path.exists():
            print(f"  [warn] {path} not found — skipping")
            continue
        lg = path.stem
        try:
            df = pd.read_excel(path)
        except Exception as e:
            print(f"  [warn] Could not read {path.name}: {e}")
            continue
        df = df.copy()
        df["_League"] = lg
        frames.append(df)
        print(f"  Loaded {lg}: {len(df)} players")

    if not frames:
        raise RuntimeError(f"No Wyscout files found in {WYSCOUT_DIR}")

    raw = pd.concat(frames, ignore_index=True)

    # Standardise minutes column
    mins_col = next((c for c in ["Minutes played", "MinutesPlayed", "Minutes"] if c in raw.columns), None)
    if mins_col:
        raw["_minutes"] = pd.to_numeric(raw[mins_col], errors="coerce").fillna(0)
    else:
        raw["_minutes"] = 0

    raw = raw[raw["_minutes"] >= min_minutes].copy()
    print(f"  → {len(raw)} players after {min_minutes}+ minute filter")
    return raw.reset_index(drop=True)


# ── Position normalisation ─────────────────────────────────────────────────────

def map_position(pos_str: str) -> str:
    if not isinstance(pos_str, str):
        return "Other"
    first = pos_str.split(",")[0].strip()
    return POS_MAP.get(first, "Other")


def add_position_group(df: pd.DataFrame) -> pd.DataFrame:
    pos_col = next((c for c in ["Position", "Pos"] if c in df.columns), None)
    if pos_col:
        df["_pos_group"] = df[pos_col].apply(map_position)
        df["_full_position"] = df[pos_col].fillna("Unknown")
    else:
        df["_pos_group"] = "Other"
        df["_full_position"] = "Unknown"
    return df


# ── SQS Score ─────────────────────────────────────────────────────────────────

def compute_sqs(df: pd.DataFrame) -> pd.DataFrame:
    """Compute per-position SQS Rank (0–100 percentile within the pool)."""
    df = df.copy()
    df["_raw_sqs"] = np.nan

    for pos, blueprint in SQS_BLUEPRINTS.items():
        mask = df["_pos_group"] == pos
        if mask.sum() == 0:
            continue
        grp = df.loc[mask].copy()
        score = pd.Series(0.0, index=grp.index)
        total_w = 0.0
        for metric, w in blueprint:
            if metric in grp.columns:
                vals = pd.to_numeric(grp[metric], errors="coerce").fillna(0)
                pct = vals.rank(pct=True) * 100
                score += w * pct
                total_w += w
        if total_w > 0:
            score /= total_w
        df.loc[mask, "_raw_sqs"] = score.values

    # SQS rank = percentile within the same position group (0–100)
    df["_sqs_rank"] = np.nan
    for pos in SQS_BLUEPRINTS:
        mask = df["_pos_group"] == pos
        if mask.sum() == 0:
            continue
        df.loc[mask, "_sqs_rank"] = (
            df.loc[mask, "_raw_sqs"].rank(pct=True) * 100
        ).round(2)

    return df


# ── Market Value ───────────────────────────────────────────────────────────────

def compute_mv_rank(df: pd.DataFrame) -> pd.DataFrame:
    mv_col = next((c for c in ["Market value", "MarketValue"] if c in df.columns), None)
    if mv_col:
        mv = pd.to_numeric(df[mv_col], errors="coerce").fillna(0)
    else:
        mv = pd.Series(0.0, index=df.index)

    df["_mkt_val"] = mv
    df["_mv_rank"] = mv.rank(pct=True) * 100
    return df


def model_value(mkt: float, sqs: float) -> float:
    """
    Model value: market value × (SQS / 50).
    If SQS > 50, player is undervalued → model val > market val.
    If SQS < 50, player is overvalued → model val < market val.
    """
    ratio = max(sqs / 50.0, 0.1)
    return round(mkt * ratio, 0)


def val_ratio_str(mkt: float, model: float) -> str:
    if mkt <= 0:
        return "N/A"
    ratio = model / mkt
    return f"{ratio:.1f}×"


# ── Lamberts Index ─────────────────────────────────────────────────────────────

def compute_lamberts(df: pd.DataFrame) -> pd.DataFrame:
    df["_lamberts"] = (df["_sqs_rank"] - df["_mv_rank"]).round(2)

    def tier(li: float) -> str:
        if li >= 30:
            return "ELITE VALUE"
        if li >= 20:
            return "HIGH VALUE"
        if li >= 10:
            return "VALUE"
        if li >= 0:
            return "FAIR VALUE"
        return "OVERPRICED"

    df["_tier"] = df["_lamberts"].apply(tier)
    return df


# ── vs Hradec & Status ─────────────────────────────────────────────────────────

def compute_vs_hradec(df: pd.DataFrame) -> pd.DataFrame:
    """
    vs Hradec = target SQS rank − weakest Hradec starter quality at same pos.
    Status:
      CLEAR UPGRADE  → vs Hradec > 0
      ROTATIONAL / COVER → −10 < vs Hradec ≤ 0
      DEPTH          → vs Hradec ≤ −10
    """
    vs_col: list[float] = []
    status_col: list[str] = []

    for _, row in df.iterrows():
        pos = row.get("_pos_group", "Other")
        hradec_quality = HRADEC_SQUAD.get(pos, 50.0)
        sqs = row.get("_sqs_rank", 50.0)
        if pd.isna(sqs):
            sqs = 50.0
        gap = round(sqs - hradec_quality, 1)
        vs_col.append(gap)

        if gap > 0:
            status_col.append("CLEAR UPGRADE")
        elif gap > -10:
            status_col.append("ROTATIONAL / COVER")
        else:
            status_col.append("DEPTH")

    df["_vs_hradec"] = vs_col
    df["_status"] = status_col
    return df


# ── Build master table ─────────────────────────────────────────────────────────

STAT_MAP = {
    "Goals/90":     "Goals per 90",
    "xG/90":        "xG per 90",
    "Assists/90":   "Assists per 90",
    "xA/90":        "xA per 90",
    "Prog Pass/90": "Progressive passes per 90",
    "Prog Run/90":  "Progressive runs per 90",
    "Box Touch/90": "Touches in box per 90",
    "Dribbles/90":  "Dribbles per 90",
    "Drib Succ %":  "Successful dribbles, %",
    "Def Duel %":   "Defensive duels won, %",
    "Aerial %":     "Aerial duels won, %",
    "Interceptions":"Interceptions per 90",
    "Key Pass/90":  "Key passes per 90",
    "Save %":       "Save rate, %",
    "Prev Goals/90":"Prevented goals per 90",
}


def build_master(df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict] = []
    contract_col = next((c for c in ["Contract expires", "ContractExpires"] if c in df.columns), None)

    for _, r in df.iterrows():
        mkt = float(r.get("_mkt_val", 0) or 0)
        sqs = float(r.get("_sqs_rank", 0) or 0)
        mod_val = model_value(mkt, sqs)
        contract = r.get(contract_col, None) if contract_col else None
        if pd.notna(contract):
            try:
                contract = pd.to_datetime(contract).strftime("%Y-%m-%d")
            except Exception:
                contract = str(contract)
        else:
            contract = None

        exp_year = None
        if contract:
            try:
                exp_year = str(pd.to_datetime(contract).year)
            except Exception:
                pass

        row: dict = {
            "Player":          r.get("Player", ""),
            "Team":            r.get("Team", ""),
            "League":          r.get("_League", ""),
            "Pos":             r.get("_pos_group", ""),
            "Full Position":   r.get("_full_position", ""),
            "Age":             int(r.get("Age", 0)) if pd.notna(r.get("Age")) else "",
            "Contract":        contract,
            "Exp?":            exp_year,
            "Mkt Val (€)":     int(mkt) if mkt > 0 else 0,
            "Model Val (€)":   int(mod_val),
            "Val Ratio":       val_ratio_str(mkt, mod_val),
            "Tier":            r.get("_tier", ""),
            "SQS Rank":        round(float(r.get("_sqs_rank", 0) or 0), 2),
            "Lamberts Index":  round(float(r.get("_lamberts", 0) or 0), 2),
            "Status":          r.get("_status", ""),
            "vs Hradec":       r.get("_vs_hradec", 0),
            "Minutes":         int(r.get("_minutes", 0)),
        }

        for out_col, in_col in STAT_MAP.items():
            val = r.get(in_col, 0)
            try:
                row[out_col] = round(float(val), 2) if pd.notna(val) else 0
            except Exception:
                row[out_col] = 0

        rows.append(row)

    master = pd.DataFrame(rows)
    master = master.sort_values("Lamberts Index", ascending=False).reset_index(drop=True)
    return master


OUTPUT_COLS = [
    "Player", "Team", "League", "Pos", "Full Position", "Age", "Contract", "Exp?",
    "Mkt Val (€)", "Model Val (€)", "Val Ratio", "Tier", "SQS Rank", "Lamberts Index",
    "Status", "vs Hradec", "Minutes",
    "Goals/90", "xG/90", "Assists/90", "xA/90",
    "Prog Pass/90", "Prog Run/90", "Box Touch/90",
    "Dribbles/90", "Drib Succ %", "Def Duel %", "Aerial %",
    "Interceptions", "Key Pass/90", "Save %", "Prev Goals/90",
]


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size)


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        try:
            max_len = max(
                len(str(col_cells[0].value or "")),
                *(len(str(c.value or "")) for c in col_cells[1:10]),
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)
        except Exception:
            pass


def write_data_sheet(ws, title: str, subtitle: str, df: pd.DataFrame) -> None:
    """Write title row, subtitle row, then data with styled header."""
    # Title
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 10))
    ws["A1"].font = Font(bold=True, color=C["white"], size=13)
    ws["A1"].fill = _fill(C["navy"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Subtitle
    ws.append([subtitle])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(df.columns), 10))
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["navy"])
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    if df.empty:
        return

    # Header row
    ws.append(list(df.columns))
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font = Font(bold=True, color=C["white"], size=9)
        cell.fill = _fill(C["header"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[hdr_row].height = 28

    # Data rows
    tier_colors = {
        "ELITE VALUE": C["elite"],
        "HIGH VALUE":  C["high"],
        "VALUE":       C["value"],
        "FAIR VALUE":  C["fair"],
        "OVERPRICED":  C["over"],
    }
    status_colors = {
        "CLEAR UPGRADE":      C["upgrade"],
        "ROTATIONAL / COVER": C["rota"],
        "DEPTH":              C["depth"],
    }

    tier_col_idx = list(df.columns).index("Tier") + 1 if "Tier" in df.columns else None
    status_col_idx = list(df.columns).index("Status") + 1 if "Status" in df.columns else None
    li_col_idx = list(df.columns).index("Lamberts Index") + 1 if "Lamberts Index" in df.columns else None

    for i, row_vals in enumerate(df.itertuples(index=False), start=1):
        ws.append(list(row_vals))
        data_row = ws.max_row
        bg = C["light"] if i % 2 == 0 else C["white"]
        for cell in ws[data_row]:
            cell.font = Font(size=9)
            cell.fill = _fill(bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()

        # Colour tier cell
        if tier_col_idx:
            tc = ws.cell(data_row, tier_col_idx)
            tier_val = str(tc.value or "")
            hex_c = tier_colors.get(tier_val)
            if hex_c:
                tc.fill = _fill(hex_c)
                tc.font = Font(bold=True, color=C["white"], size=9)

        # Colour status cell
        if status_col_idx:
            sc = ws.cell(data_row, status_col_idx)
            status_val = str(sc.value or "")
            hex_c = status_colors.get(status_val)
            if hex_c:
                sc.fill = _fill(hex_c)
                sc.font = Font(bold=True, color=C["white"], size=9)

        # Bold LI
        if li_col_idx:
            lc = ws.cell(data_row, li_col_idx)
            lc.font = Font(bold=True, size=9)

    ws.freeze_panes = f"A{hdr_row + 1}"
    _autofit(ws)


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_readme(ws, leagues: list[str], total: int, clear: int, budget: int) -> None:
    ws.title = "README"
    ws.sheet_view.showGridLines = False

    data = [
        ("FC HRADEC KRÁLOVÉ — LAMBERTS INDEX RECRUITMENT MODEL 2025–26", None),
        (f"Waltzing Analytics  ·  Jamestown / Marc Lamberts methodology  ·  "
         f"{len(leagues) if leagues else 'All'} leagues  ·  Budget ≤ €{budget:,}  ·  Age ≤ 30", None),
        (None, None),
        (None, "WORKBOOK STRUCTURE"),
        (None, "Sheet"),
        (None, "README"),
        (None, "Priority List"),
        (None, "Elite Picks"),
        (None, "Lamberts Analysis"),
        (None, "Top 5 per Position"),
        (None, "All Targets"),
        (None, "GK / CB / FB / DM / CM / W / FW"),
        (None, "Expiring 2026"),
        (None, "Budget Planner"),
        (None, "Squad"),
    ]

    desc_map = {
        "README":             "This guide",
        "Priority List":      f"All {clear} clear upgrades ranked by Lamberts Index",
        "Elite Picks":        "Players with Lamberts Index ≥ 30",
        "Lamberts Analysis":  "Tier breakdown and league statistics",
        "Top 5 per Position": "Best 5 per position by Lamberts Index",
        "All Targets":        f"Full {total} candidate database with filters",
        "GK / CB / FB / DM / CM / W / FW": "Position-specific scouting boards",
        "Expiring 2026":      "Players whose contracts expire in 2026 (free/cheap)",
        "Budget Planner":     f"Build squad within €{budget:,} budget",
        "Squad":              "Hradec Králové current quality benchmarks",
    }

    ws.append(["FC HRADEC KRÁLOVÉ — LAMBERTS INDEX RECRUITMENT MODEL 2025–26"])
    ws.merge_cells("A1:T1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=15)
    ws["A1"].fill = _fill(C["navy"])
    ws.row_dimensions[1].height = 28

    league_label = f"{len(leagues)} leagues" if len(leagues) > 5 else " + ".join(leagues)
    ws.append([f"Waltzing Analytics  ·  Jamestown / Marc Lamberts methodology  ·  "
               f"{league_label}  ·  Budget ≤ €{budget:,}  ·  Age ≤ 30"])
    ws.merge_cells("A2:T2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=10)
    ws["A2"].fill = _fill(C["navy"])
    ws.row_dimensions[2].height = 18

    ws.append([None])
    ws.append([None, "WORKBOOK STRUCTURE"])
    ws["B4"].font = Font(bold=True, size=11, color=C["navy"])
    ws.row_dimensions[4].height = 20

    ws.append([None, "Sheet", "Contents"])
    for col_letter in "BC":
        cell = ws[f"{col_letter}5"]
        cell.font = Font(bold=True, color=C["white"])
        cell.fill = _fill(C["header"])
        cell.alignment = Alignment(horizontal="left")

    for sheet_name, desc in desc_map.items():
        ws.append([None, sheet_name, desc])
        row = ws.max_row
        ws[f"B{row}"].font = Font(bold=True, color=C["navy"])

    ws.append([None])
    ws.append([None, "HOW THE LAMBERTS INDEX WORKS"])
    ws[f"B{ws.max_row}"].font = Font(bold=True, size=11, color=C["navy"])

    explanations = [
        ("SQS Rank", "Squad Quality Score — position-specific percentile rank of statistical output (0–100)"),
        ("Market Value Rank", "Transfermarkt market value percentile rank within the recruitment pool (0–100)"),
        ("Lamberts Index (LI)", "SQS Rank − Market Value Rank. Positive = undervalued. Negative = overpriced."),
        ("ELITE VALUE",  "LI ≥ 30 — buy signal: performs far above what the market charges"),
        ("HIGH VALUE",   "LI ≥ 20"),
        ("VALUE",        "LI ≥ 10"),
        ("FAIR VALUE",   "LI 0–9 — price reflects performance"),
        ("OVERPRICED",   "LI < 0 — market overcharges relative to output"),
        ("Model Val",    "Market value × (SQS / 50). Fair price the model assigns based on output."),
        ("vs Hradec",    "Target SQS − weakest Hradec starter at that position. Positive = clear upgrade."),
        ("CLEAR UPGRADE","vs Hradec > 0 — statistically better than current starter"),
        ("ROTATIONAL",   "−10 < vs Hradec ≤ 0 — squad cover / rotation option"),
        ("DEPTH",        "vs Hradec ≤ −10 — developmental or depth signing"),
    ]
    for term, desc in explanations:
        ws.append([None, term, desc])
        row = ws.max_row
        ws[f"B{row}"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 80


def build_lamberts_analysis(ws, master: pd.DataFrame) -> None:
    ws.title = "Lamberts Analysis"
    ws.sheet_view.showGridLines = False

    ws.append(["LAMBERTS INDEX ANALYSIS — COMPLETE BREAKDOWN"])
    ws.merge_cells("A1:N1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=13)
    ws["A1"].fill = _fill(C["navy"])
    ws.row_dimensions[1].height = 22

    ws.append(["LI = SQS Percentile Rank − Market Value Percentile Rank  ·  "
               "ELITE ≥30  ·  HIGH ≥20  ·  VALUE ≥10  ·  FAIR 0–9  ·  OVER <0"])
    ws.merge_cells("A2:N2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["navy"])
    ws.row_dimensions[2].height = 14

    ws.append([None])

    # Overall stats
    ws.append(["OVERALL LI STATISTICS"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)

    li = master["Lamberts Index"]
    overall = [
        ("Total Players", len(master)),
        ("Mean LI", round(li.mean(), 2)),
        ("Median LI", round(li.median(), 2)),
        ("Max LI", round(li.max(), 2)),
        ("Min LI", round(li.min(), 2)),
        ("Std Dev", round(li.std(), 2)),
    ]
    ws.append(["Metric", "Value", None,
               "Tier Breakdown", "Count", "%", None,
               "League", "Count", "Avg LI"])
    hdr = ws.max_row
    for cell in ws[hdr]:
        if cell.value:
            cell.font = Font(bold=True, color=C["white"])
            cell.fill = _fill(C["header"])

    tier_counts = master["Tier"].value_counts()
    league_stats = master.groupby("League")["Lamberts Index"].agg(["count", "mean"]).round(2)
    tier_order = ["ELITE VALUE", "HIGH VALUE", "VALUE", "FAIR VALUE", "OVERPRICED"]

    for i, (metric, val) in enumerate(overall):
        row = ws.max_row + 1
        ws.append([metric, val])

        # Tier column (D-F)
        tier_name = tier_order[i] if i < len(tier_order) else ""
        cnt = int(tier_counts.get(tier_name, 0))
        pct = round(cnt / len(master) * 100, 1) if len(master) > 0 else 0
        ws.cell(row, 4).value = tier_name
        ws.cell(row, 5).value = cnt
        ws.cell(row, 6).value = f"{pct}%"

        # League column (H-J)
        if i < len(league_stats):
            lname = league_stats.index[i]
            ws.cell(row, 8).value = lname
            ws.cell(row, 9).value = int(league_stats.loc[lname, "count"])
            ws.cell(row, 10).value = league_stats.loc[lname, "mean"]

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 12


def build_top5(ws, master: pd.DataFrame) -> None:
    ws.title = "Top 5 per Position"
    ws.sheet_view.showGridLines = False

    ws.append(["TOP 5 PLAYERS PER POSITION — Ranked by Lamberts Index"])
    ws.merge_cells("A1:M1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=13)
    ws["A1"].fill = _fill(C["navy"])
    ws.row_dimensions[1].height = 22

    ws.append(["Highest Lamberts Index player at each position group across all recruitment leagues"])
    ws.merge_cells("A2:M2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["navy"])

    medals = ["🥇", "🥈", "🥉", "4.", "5."]
    pos_labels = {"GK": "GK — GOALKEEPER", "CB": "CB — CENTRE-BACK", "FB": "FB — FULL-BACK",
                  "DM": "DM — DEFENSIVE MID", "CM": "CM — CENTRAL MID", "W": "W — WINGER", "FW": "FW — FORWARD"}

    for pos, label in pos_labels.items():
        grp = master[master["Pos"] == pos].head(5)
        if grp.empty:
            continue

        ws.append([f"  {label}"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=13)
        ws[f"A{ws.max_row}"].font = Font(bold=True, color=C["white"], size=10)
        ws[f"A{ws.max_row}"].fill = _fill(C["navy"])
        ws.row_dimensions[ws.max_row].height = 18

        ws.append(["#", "Player", "Team", "League", "Age", "Contract",
                   "Mkt Val (€)", "SQS Rank", "Lamberts LI", "Val Ratio",
                   "vs Hradec", "Status", "Tier"])
        hdr = ws.max_row
        for cell in ws[hdr]:
            cell.font = Font(bold=True, color=C["white"], size=9)
            cell.fill = _fill(C["header"])
            cell.alignment = Alignment(horizontal="center")

        for i, (_, r) in enumerate(grp.iterrows()):
            status_short = {
                "CLEAR UPGRADE": "▲ Clear",
                "ROTATIONAL / COVER": "↔ Rota",
                "DEPTH": "▼ Depth",
            }.get(r["Status"], r["Status"])
            ws.append([
                medals[i], r["Player"], r["Team"], r["League"],
                r["Age"], r["Contract"], r["Mkt Val (€)"],
                r["SQS Rank"], r["Lamberts Index"], r["Val Ratio"],
                r["vs Hradec"], status_short, r["Tier"],
            ])
            dr = ws.max_row
            bg = C["light"] if i % 2 == 0 else C["white"]
            for cell in ws[dr]:
                cell.font = Font(size=9)
                cell.fill = _fill(bg)
                cell.alignment = Alignment(horizontal="center")
            ws.cell(dr, 1).font = Font(bold=True, size=10)

        ws.append([None])

    _autofit(ws)


def build_squad(ws) -> None:
    ws.title = "Squad"
    ws.sheet_view.showGridLines = False

    ws.append(["FC HRADEC KRÁLOVÉ — CURRENT SQUAD QUALITY BENCHMARKS"])
    ws.merge_cells("A1:L1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=13)
    ws["A1"].fill = _fill(C["navy"])
    ws.row_dimensions[1].height = 22

    ws.append(["Czech top-flight league percentile ranks  ·  Higher = better  ·  Basis for vs Hradec gap calculation"])
    ws.merge_cells("A2:L2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["navy"])

    ws.append([None])

    ws.append(["SQUAD QUALITY BY POSITION"])
    ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)

    ws.append(["Position", "Hradec Quality %", "Note"])
    hdr = ws.max_row
    for cell in ws[hdr]:
        if cell.value:
            cell.font = Font(bold=True, color=C["white"])
            cell.fill = _fill(C["header"])
            cell.alignment = Alignment(horizontal="center")

    pos_notes = {
        "GK": "Adam Zadrazil — starter",
        "CB": "Daniel Horak — weakest starter",
        "FB": "Martin Suchomel — starter",
        "DM": "Jakub Elbel — starter",
        "CM": "Daniel Trubac — median starter",
        "W":  "Samuel Dancak — weakest starter",
        "FW": "Mick van Buren — starter",
    }
    for pos, quality in HRADEC_SQUAD.items():
        ws.append([pos, quality, pos_notes.get(pos, "")])
        row = ws.max_row
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40


def build_budget_planner(ws, master: pd.DataFrame, budget: int) -> None:
    ws.title = "Budget Planner"
    ws.sheet_view.showGridLines = False

    ws.append([f"BUDGET PLANNER — FC Hradec Králové  ·  Budget Cap: €{budget:,}"])
    ws.merge_cells("A1:O1")
    ws["A1"].font = Font(bold=True, color=C["white"], size=13)
    ws["A1"].fill = _fill(C["navy"])
    ws.row_dimensions[1].height = 22

    ws.append(["ELITE VALUE + HIGH VALUE clear upgrades sorted by market value (cheapest first). "
               "Build your squad within budget."])
    ws.merge_cells("A2:O2")
    ws["A2"].font = Font(italic=True, color=C["gold"], size=9)
    ws["A2"].fill = _fill(C["navy"])

    ws.append([None, None, None, None, "Total Budget", budget])

    ws.append([None])

    eligible = master[
        (master["Tier"].isin(["ELITE VALUE", "HIGH VALUE"])) &
        (master["Status"] == "CLEAR UPGRADE")
    ].sort_values("Mkt Val (€)").reset_index(drop=True)

    cols = ["#", "Player", "Pos", "Team", "League", "Age", "Contract",
            "Mkt Val (€)", "Model Val (€)", "Val Ratio", "SQS Rank",
            "Lamberts Index", "vs Hradec", "Tier", "Running Total (€)"]

    ws.append(cols)
    hdr = ws.max_row
    for cell in ws[hdr]:
        cell.font = Font(bold=True, color=C["white"], size=9)
        cell.fill = _fill(C["header"])
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hdr].height = 28

    running = 0
    for i, (_, r) in enumerate(eligible.iterrows(), 1):
        mv = int(r["Mkt Val (€)"])
        running += mv
        over_budget = running > budget
        ws.append([
            i, r["Player"], r["Pos"], r["Team"], r["League"],
            r["Age"], r["Contract"],
            mv, int(r["Model Val (€)"]), r["Val Ratio"],
            r["SQS Rank"], r["Lamberts Index"],
            r["vs Hradec"], r["Tier"], running,
        ])
        dr = ws.max_row
        for cell in ws[dr]:
            cell.font = Font(size=9, color="922B21" if over_budget else "000000")
            cell.fill = _fill("FDE8E8" if over_budget else (C["light"] if i % 2 == 0 else C["white"]))
            cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = f"A{hdr + 1}"
    _autofit(ws)


# ── Main ───────────────────────────────────────────────────────────────────────

def run(leagues: list[str], min_minutes: int, max_age: int, budget: int, output: Path) -> None:
    print(f"\n{'='*60}")
    print(f"  Lamberts Index Model Builder")
    print(f"  Leagues: {leagues}")
    print(f"  Min minutes: {min_minutes}  |  Max age: {max_age}")
    print(f"{'='*60}\n")

    # 1. Load data
    print("Loading Wyscout files…")
    raw = load_leagues(leagues, min_minutes)

    # 2. Filter age
    if "Age" in raw.columns:
        raw = raw[pd.to_numeric(raw["Age"], errors="coerce").fillna(99) <= max_age]
        print(f"  → {len(raw)} players after age ≤ {max_age} filter")

    # 3. Position groups
    raw = add_position_group(raw)
    raw = raw[raw["_pos_group"] != "Other"].copy()
    print(f"  → {len(raw)} players with known position")

    # 4. SQS + MV rank + Lamberts
    print("Computing SQS, market value rank, and Lamberts Index…")
    raw = compute_sqs(raw)
    raw = compute_mv_rank(raw)
    raw = compute_lamberts(raw)
    raw = compute_vs_hradec(raw)

    # 5. Build master table
    print("Building master table…")
    master = build_master(raw)
    master_out = master[[c for c in OUTPUT_COLS if c in master.columns]]

    print(f"  → {len(master)} total candidates")
    clear_upgrades = master[master["Status"] == "CLEAR UPGRADE"]
    print(f"  → {len(clear_upgrades)} clear upgrades")
    elite = master[master["Tier"] == "ELITE VALUE"]
    print(f"  → {len(elite)} ELITE VALUE players")

    # 6. Write workbook
    print(f"\nWriting workbook → {output}")
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # README
    print("  Writing README…")
    ws_readme = wb.create_sheet("README")
    league_list = leagues if leagues else sorted(master["League"].unique().tolist())
    build_readme(ws_readme, league_list, len(master), len(clear_upgrades), budget)

    # Priority List (clear upgrades only)
    print("  Writing Priority List…")
    ws_prio = wb.create_sheet("Priority List")
    prio_df = master_out[master_out["Status"] == "CLEAR UPGRADE"].copy()
    write_data_sheet(
        ws_prio,
        f"PRIORITY LIST — All {len(prio_df)} Clear Upgrades — Ranked by Lamberts Index",
        "All are clear statistical upgrades on current Hradec starters  ·  "
        f"Budget ≤ €{budget//1_000_000}M",
        prio_df,
    )

    # Elite Picks
    print("  Writing Elite Picks…")
    ws_elite = wb.create_sheet("Elite Picks")
    elite_df = master_out[master_out["Tier"] == "ELITE VALUE"].copy()
    write_data_sheet(
        ws_elite,
        f"ELITE VALUE — {len(elite_df)} Players with Lamberts Index ≥ 30",
        "Strongest buy signals  ·  Sorted by Lamberts Index",
        elite_df,
    )

    # Lamberts Analysis
    print("  Writing Lamberts Analysis…")
    ws_analysis = wb.create_sheet("Lamberts Analysis")
    build_lamberts_analysis(ws_analysis, master)

    # Top 5 per Position
    print("  Writing Top 5 per Position…")
    ws_top5 = wb.create_sheet("Top 5 per Position")
    build_top5(ws_top5, master_out)

    # All Targets
    print("  Writing All Targets…")
    ws_all = wb.create_sheet("All Targets")
    write_data_sheet(
        ws_all,
        f"ALL TARGETS — {len(master)} Candidates  ·  {' + '.join(leagues)}",
        "Use column filters to narrow by position, league, age, status, tier  ·  "
        "Ranked by Lamberts Index",
        master_out,
    )

    # Position sheets
    pos_config = [
        ("GK",  "GOALKEEPER TARGETS"),
        ("CB",  "CENTRE-BACK TARGETS"),
        ("FB",  "FULL-BACK TARGETS"),
        ("DM",  "DEFENSIVE MID TARGETS"),
        ("CM",  "CENTRAL MID TARGETS"),
        ("W",   "WINGER TARGETS"),
        ("FW",  "FORWARD TARGETS"),
    ]
    for pos, label in pos_config:
        grp = master_out[master_out["Pos"] == pos].copy()
        ws_pos = wb.create_sheet(pos)
        cnt = len(grp)
        print(f"  Writing {pos} ({cnt} players)…")
        write_data_sheet(
            ws_pos,
            f"{label}  ·  {cnt} candidates",
            "Sorted by Lamberts Index  ·  All candidates in recruitment universe",
            grp,
        )

    # Expiring 2026
    print("  Writing Expiring 2026…")
    ws_exp = wb.create_sheet("Expiring 2026")
    exp_df = master_out[master_out["Exp?"].astype(str) == "2026"].copy()
    write_data_sheet(
        ws_exp,
        f"EXPIRING CONTRACTS 2026 — {len(exp_df)} PLAYERS  ·  "
        "Free transfer or cut-price acquisition opportunity",
        "",
        exp_df,
    )

    # Budget Planner
    print("  Writing Budget Planner…")
    ws_budget = wb.create_sheet("Budget Planner")
    build_budget_planner(ws_budget, master, budget)

    # Squad
    print("  Writing Squad…")
    ws_squad = wb.create_sheet("Squad")
    build_squad(ws_squad)

    wb.save(output)
    size_mb = output.stat().st_size / 1_048_576
    print(f"\nDone. {size_mb:.1f} MB → {output.resolve()}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build Lamberts Index Model workbook from Wyscout data")
    parser.add_argument("--leagues", nargs="+", default=DEFAULT_LEAGUES,
                        help="League names (without .xlsx). Omit to load ALL leagues.")
    parser.add_argument("--min-minutes", type=int, default=DEFAULT_MIN_MINUTES)
    parser.add_argument("--max-age", type=int, default=DEFAULT_MAX_AGE)
    parser.add_argument("--budget", type=int, default=DEFAULT_BUDGET)
    parser.add_argument("--output", type=Path,
                        default=ROOT / "data" / "Lamberts_Index_Model_All_Leagues.xlsx")
    args = parser.parse_args()
    run(
        leagues=args.leagues,
        min_minutes=args.min_minutes,
        max_age=args.max_age,
        budget=args.budget,
        output=args.output,
    )
