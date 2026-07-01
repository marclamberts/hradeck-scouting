"""
build_nordic_baltic_ratings.py
────────────────────────────────
Data system for the Nordic-Baltic region (Sweden, Denmark, Norway, Finland,
Iceland, Latvia, Lithuania), built from the raw Wyscout league workbooks in
`data/Wyscout DB/`.

Produces two workbooks in `League Analysis/`:

1. Nordic-Baltic Team Ratings.xlsx
   Attacking and Defensive rating (0-100) for every club, aggregated from
   player per-90 output weighted by "Minutes played" and grouped by
   "Team within selected timeframe" (so mid-season transfers land on the
   right club). Ratings are percentiled within each league/division for a
   fair like-for-like read, plus a tier-adjusted version for cross-league
   comparison.

2. Nordic-Baltic League Rankings.xlsx
   Ranks the 15 leagues/divisions in the region against each other using a
   League Strength Index that blends the region's established tier
   weighting with data-driven signals (average valued-player market value),
   alongside descriptive scouting stats (age, squad size, style output).

Run:  python3 build_nordic_baltic_ratings.py
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import norm
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WYSCOUT_DB_DIR = Path(__file__).parent / "data" / "Wyscout DB"
OUT_DIR = Path(__file__).parent / "League Analysis"
OUT_DIR.mkdir(exist_ok=True)

# ── Region definition ───────────────────────────────────────────────────────
# file stem (in data/Wyscout DB) -> country / division / official league name
NORDIC_BALTIC_LEAGUES: dict[str, dict] = {
    "Sweden":      {"country": "Sweden",    "division": 1, "league_name": "Allsvenskan"},
    "Sweden II":   {"country": "Sweden",    "division": 2, "league_name": "Superettan"},
    "Sweden III":  {"country": "Sweden",    "division": 3, "league_name": "Division 2 (SWE)"},
    "Denmark":     {"country": "Denmark",   "division": 1, "league_name": "Superligaen"},
    "Denmark II":  {"country": "Denmark",   "division": 2, "league_name": "1st Division"},
    "Denmark III": {"country": "Denmark",   "division": 3, "league_name": "Danish 2nd Division"},
    "Denmark IV":  {"country": "Denmark",   "division": 4, "league_name": "Danish 3rd Division"},
    "Norway":      {"country": "Norway",    "division": 1, "league_name": "Eliteserien"},
    "Norway II":   {"country": "Norway",    "division": 2, "league_name": "1. divisjon"},
    "Norway III":  {"country": "Norway",    "division": 3, "league_name": "2. divisjon"},
    "Finland":     {"country": "Finland",   "division": 1, "league_name": "Veikkausliiga"},
    "Finland II":  {"country": "Finland",   "division": 2, "league_name": "Ykkonen"},
    "Iceland":     {"country": "Iceland",   "division": 1, "league_name": "Urvalsdeild"},
    "Latvia":      {"country": "Latvia",    "division": 1, "league_name": "Virsliga"},
    "Lithuania":   {"country": "Lithuania", "division": 1, "league_name": "A Lyga"},
}

TIER_WEIGHT = {1: 1.00, 2: 0.93, 3: 0.86, 4: 0.79, 5: 0.72, 6: 0.65}

# ── Rating blueprints: (metric, weight, direction) — direction -1 means "lower is better" ──
ATTACK_BLUEPRINT = [
    ("Goals per 90", 3.0, 1),
    ("xG per 90", 2.5, 1),
    ("Non-penalty goals per 90", 1.5, 1),
    ("Shots per 90", 1.0, 1),
    ("Shots on target, %", 1.0, 1),
    ("Goal conversion, %", 1.0, 1),
    ("Touches in box per 90", 1.5, 1),
    ("xA per 90", 2.0, 1),
    ("Key passes per 90", 1.5, 1),
    ("Assists per 90", 1.5, 1),
    ("Successful attacking actions per 90", 1.5, 1),
    ("Progressive runs per 90", 1.0, 1),
    ("Dribbles per 90", 0.75, 1),
    ("Successful dribbles, %", 0.75, 1),
    ("Deep completions per 90", 1.0, 1),
    ("Accurate crosses, %", 0.5, 1),
]

# Outfield defensive activity — aggregated across the whole squad
DEFENSE_OUTFIELD_BLUEPRINT = [
    ("Successful defensive actions per 90", 2.5, 1),
    ("Defensive duels per 90", 1.0, 1),
    ("Defensive duels won, %", 2.0, 1),
    ("Interceptions per 90", 2.0, 1),
    ("PAdj Interceptions", 1.5, 1),
    ("Aerial duels won, %", 1.5, 1),
    ("Shots blocked per 90", 1.0, 1),
    ("Sliding tackles per 90", 0.75, 1),
    ("PAdj Sliding tackles", 0.75, 1),
]

# Team defensive record — aggregated across goalkeeper minutes only
DEFENSE_RECORD_BLUEPRINT = [
    ("Conceded goals per 90", 3.0, -1),
    ("xG against per 90", 2.0, -1),
    ("Save rate, %", 1.5, 1),
    ("Prevented goals per 90", 2.0, 1),
]

DEFENSE_BLUEPRINT = DEFENSE_OUTFIELD_BLUEPRINT + DEFENSE_RECORD_BLUEPRINT

ALL_METRICS = sorted({m for m, *_ in ATTACK_BLUEPRINT + DEFENSE_BLUEPRINT})

TEAM_KEY = "Team within selected timeframe"


# ── Loading ──────────────────────────────────────────────────────────────────

def load_region_raw() -> pd.DataFrame:
    frames = []
    for stem, meta in NORDIC_BALTIC_LEAGUES.items():
        path = WYSCOUT_DB_DIR / f"{stem}.xlsx"
        if not path.exists():
            print(f"  ! Missing file, skipped: {path}")
            continue
        df = pd.read_excel(path)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.assign(
            LeagueFile=stem,
            Country=meta["country"],
            Division=meta["division"],
            LeagueName=meta["league_name"],
        )
        frames.append(df)

    raw = pd.concat(frames, ignore_index=True)
    raw["Position"] = raw["Position"].astype(str).str.split(",").str[0].str.strip()
    raw["Minutes played"] = pd.to_numeric(raw["Minutes played"], errors="coerce").fillna(0)
    raw["Market value"] = pd.to_numeric(raw["Market value"], errors="coerce").fillna(0)
    raw["Age"] = pd.to_numeric(raw["Age"], errors="coerce")
    for metric in ALL_METRICS:
        if metric in raw.columns:
            raw[metric] = pd.to_numeric(raw[metric], errors="coerce")
    raw = raw.loc[raw[TEAM_KEY].notna() & (raw["Minutes played"] > 0)]

    # Drop degenerate "team" entries with a single player and a handful of
    # minutes (trialist/placeholder rows in the source data, not real clubs).
    team_minutes = raw.groupby([TEAM_KEY, "LeagueFile"])["Minutes played"].transform("sum")
    raw = raw.loc[team_minutes >= 3000]

    return raw.reset_index(drop=True)


def _weighted_avg(group: pd.DataFrame, metric: str, weight_col: str = "Minutes played") -> float:
    if metric not in group.columns:
        return np.nan
    vals = group[metric]
    w = group[weight_col]
    mask = vals.notna() & (w > 0)
    if not mask.any():
        return np.nan
    return float((vals[mask] * w[mask]).sum() / w[mask].sum())


# ── Team-level aggregation ───────────────────────────────────────────────────

def build_team_metrics(raw: pd.DataFrame) -> pd.DataFrame:
    key_cols = [TEAM_KEY, "LeagueFile", "Country", "LeagueName", "Division"]
    rows = []
    for keys, grp in raw.groupby(key_cols, dropna=False):
        row = dict(zip(key_cols, keys))
        row["Players"] = grp["Player"].nunique()
        row["TotalMinutes"] = grp["Minutes played"].sum()
        for metric, _, _ in ATTACK_BLUEPRINT + DEFENSE_OUTFIELD_BLUEPRINT:
            row[metric] = _weighted_avg(grp, metric)

        gk_grp = grp.loc[grp["Position"] == "GK"]
        for metric, _, _ in DEFENSE_RECORD_BLUEPRINT:
            row[metric] = _weighted_avg(gk_grp, metric)
        row["GK Minutes"] = gk_grp["Minutes played"].sum()

        rows.append(row)

    return pd.DataFrame(rows)


def _composite_percentile(df: pd.DataFrame, blueprint: list[tuple[str, float, int]]) -> pd.Series:
    """Weighted z-score composite -> 0-100 percentile, computed within the given df."""
    available = [(m, w, d) for m, w, d in blueprint if m in df.columns and df[m].notna().any()]
    if not available:
        return pd.Series(50.0, index=df.index)
    total_w = sum(w for _, w, _ in available)
    z = pd.Series(0.0, index=df.index)
    for metric, weight, direction in available:
        col = df[metric]
        mu, sig = col.mean(), (col.std() or 1e-9)
        z += (weight / total_w) * direction * (col.fillna(mu) - mu) / sig
    return pd.Series(norm.cdf(z.values) * 100, index=df.index)


def rate_teams(team_metrics: pd.DataFrame) -> pd.DataFrame:
    df = team_metrics.copy()
    df["AttackingRating"] = np.nan
    df["DefensiveRating"] = np.nan

    for _, idx in df.groupby("LeagueFile").groups.items():
        sub = df.loc[idx]
        df.loc[idx, "AttackingRating"] = _composite_percentile(sub, ATTACK_BLUEPRINT).values
        df.loc[idx, "DefensiveRating"] = _composite_percentile(sub, DEFENSE_BLUEPRINT).values

    df["OverallRating"] = (df["AttackingRating"] + df["DefensiveRating"]) / 2

    tier_map = {stem: LEAGUE_TIER.get(stem, 5) for stem in NORDIC_BALTIC_LEAGUES}
    df["Tier"] = df["LeagueFile"].map(tier_map)
    df["TierLabel"] = df["Tier"].map(TIER_LABEL)
    df["TierWeight"] = df["Tier"].map(TIER_WEIGHT).fillna(0.65)
    df["TierAdjustedRating"] = df["OverallRating"] * df["TierWeight"]

    df["LeagueRank"] = df.groupby("LeagueFile")["OverallRating"].rank(method="min", ascending=False).astype(int)
    df["RegionRank"] = df["TierAdjustedRating"].rank(method="min", ascending=False).astype(int)

    round_cols = ["AttackingRating", "DefensiveRating", "OverallRating", "TierAdjustedRating"] + ALL_METRICS
    df[round_cols] = df[round_cols].round(2)

    return df.sort_values(["RegionRank"]).reset_index(drop=True)


# ── League-level ranking ─────────────────────────────────────────────────────

def rate_leagues(raw: pd.DataFrame, team_ratings: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for stem, meta in NORDIC_BALTIC_LEAGUES.items():
        sub = raw.loc[raw["LeagueFile"] == stem]
        if sub.empty:
            continue
        valued = sub.loc[sub["Market value"] > 0, "Market value"]
        teams = team_ratings.loc[team_ratings["LeagueFile"] == stem]
        tier = LEAGUE_TIER.get(stem, 5)
        rows.append({
            "LeagueFile": stem,
            "LeagueName": meta["league_name"],
            "Country": meta["country"],
            "Division": meta["division"],
            "Tier": tier,
            "TierLabel": TIER_LABEL.get(tier, "Lower"),
            "Teams": sub[TEAM_KEY].nunique(),
            "Players": sub["Player"].nunique(),
            "MedianAge": round(float(sub["Age"].median()), 1),
            "AvgMarketValue_EUR": round(float(valued.mean()), 0) if len(valued) else 0,
            "PctPlayersValued": round(100 * len(valued) / len(sub), 1),
            "AvgGoalsPer90": round(_weighted_avg(sub, "Goals per 90"), 2),
            "AvgxGPer90": round(_weighted_avg(sub, "xG per 90"), 2),
            "CompetitiveBalance": round(float(teams["OverallRating"].std()), 1) if len(teams) > 1 else 0.0,
        })
    df = pd.DataFrame(rows)

    df["TierScore"] = df["Tier"].map(TIER_WEIGHT).fillna(0.65) * 100
    max_val = df["AvgMarketValue_EUR"].max() or 1
    df["MarketValueScore"] = (df["AvgMarketValue_EUR"] / max_val) * 100

    # League Strength Index: mostly the region's established tier calibration,
    # fine-tuned by each league's own average valued-squad market value.
    df["LeagueStrengthIndex"] = (0.65 * df["TierScore"] + 0.35 * df["MarketValueScore"]).round(1)
    df["RegionRank"] = df["LeagueStrengthIndex"].rank(method="min", ascending=False).astype(int)

    return df.sort_values("RegionRank").reset_index(drop=True)


# ── Tier lookup from the existing Leagues Overview ──────────────────────────

def _load_league_tiers() -> tuple[dict[str, int], dict[int, str]]:
    overview = pd.read_excel(Path(__file__).parent / "data" / "Leagues Overview.xlsx")
    tier_by_stem: dict[str, int] = {}
    for stem, meta in NORDIC_BALTIC_LEAGUES.items():
        match = overview.loc[
            (overview["Country"] == meta["country"]) & (overview["Division"] == meta["division"])
        ]
        tier_by_stem[stem] = int(match["Tier"].iloc[0]) if not match.empty else 5
    tier_label = overview.drop_duplicates("Tier").set_index("Tier")["Tier Label"].to_dict()
    return tier_by_stem, tier_label


LEAGUE_TIER, TIER_LABEL = {}, {}


# ── Excel styling helpers ────────────────────────────────────────────────────

HEADER_BG = "2C3E50"
WHITE = "FFFFFF"
BAND = "F5F5F5"


def _thin_border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)


def _header_style(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


def _style_sheet(writer, sheet_name: str, df: pd.DataFrame, col_widths: list[float]):
    ws = writer.sheets[sheet_name]
    ws.row_dimensions[1].height = 30
    _header_style(ws, 1, len(df.columns))
    for r in range(2, len(df) + 2):
        bg = BAND if r % 2 == 0 else WHITE
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=9)
    for i, w in enumerate(col_widths[: len(df.columns)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ── Workbook builders ────────────────────────────────────────────────────────

TEAM_DISPLAY_COLS = [
    "RegionRank", "LeagueRank", TEAM_KEY, "Country", "LeagueName", "TierLabel",
    "Players", "TotalMinutes",
    "AttackingRating", "DefensiveRating", "OverallRating", "TierAdjustedRating",
    "Goals per 90", "xG per 90", "xA per 90", "Key passes per 90",
    "Successful defensive actions per 90", "Interceptions per 90",
    "Conceded goals per 90", "xG against per 90",
]
TEAM_COL_WIDTHS = [10, 10, 24, 12, 22, 14, 9, 12, 14, 14, 13, 16, 11, 10, 10, 13, 20, 15, 16, 14]


def build_team_ratings_workbook(team_ratings: pd.DataFrame):
    out_df = team_ratings[TEAM_DISPLAY_COLS].rename(columns={TEAM_KEY: "Team"})
    out_path = OUT_DIR / "Nordic-Baltic Team Ratings.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name="All Teams", index=False)
        _style_sheet(writer, "All Teams", out_df, TEAM_COL_WIDTHS)

        for stem in NORDIC_BALTIC_LEAGUES:
            sub = out_df.loc[team_ratings["LeagueFile"] == stem].sort_values("LeagueRank")
            if sub.empty:
                continue
            sname = stem[:31]
            sub.to_excel(writer, sheet_name=sname, index=False)
            _style_sheet(writer, sname, sub, TEAM_COL_WIDTHS)

    print(f"  Saved: {out_path}")


LEAGUE_DISPLAY_COLS = [
    "RegionRank", "LeagueName", "Country", "Division", "TierLabel",
    "Teams", "Players", "MedianAge", "AvgMarketValue_EUR", "PctPlayersValued",
    "AvgGoalsPer90", "AvgxGPer90", "CompetitiveBalance", "LeagueStrengthIndex",
]
LEAGUE_COL_WIDTHS = [10, 22, 12, 10, 14, 8, 9, 10, 16, 14, 12, 10, 16, 16]


def build_league_rankings_workbook(league_ratings: pd.DataFrame):
    out_df = league_ratings[LEAGUE_DISPLAY_COLS]
    out_path = OUT_DIR / "Nordic-Baltic League Rankings.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name="League Rankings", index=False)
        _style_sheet(writer, "League Rankings", out_df, LEAGUE_COL_WIDTHS)

        methodology = pd.DataFrame({
            "Note": [
                "LeagueStrengthIndex = 0.65 x TierScore + 0.35 x MarketValueScore.",
                "TierScore comes from the region's established Tier 1-6 scale in "
                "data/Leagues Overview.xlsx (100 for Tier 1 down to 65 for Tier 6).",
                "MarketValueScore is each league's average market value among "
                "players who carry a valuation (many lower-division players are "
                "unvalued and are excluded rather than treated as worthless), "
                "rescaled 0-100 against the strongest league in the region.",
                "CompetitiveBalance is the standard deviation of team "
                "OverallRating within the league - lower means tighter title/"
                "relegation races, higher means one or two teams dominate.",
                "AvgGoalsPer90 / AvgxGPer90 describe playing style/intensity, "
                "not quality - a weak league can still produce open, high-"
                "scoring games.",
            ]
        })
        methodology.to_excel(writer, sheet_name="Methodology", index=False)
        ws = writer.sheets["Methodology"]
        _header_style(ws, 1, 1)
        ws.column_dimensions["A"].width = 110
        for r in range(2, len(methodology) + 2):
            cell = ws.cell(row=r, column=1)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[r].height = 34

    print(f"  Saved: {out_path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    global LEAGUE_TIER, TIER_LABEL
    LEAGUE_TIER, TIER_LABEL = _load_league_tiers()

    print("Loading Nordic-Baltic Wyscout data ...")
    raw = load_region_raw()
    print(f"  {len(raw)} player-rows across {raw['LeagueFile'].nunique()} leagues")

    print("Aggregating team attacking/defensive ratings ...")
    team_metrics = build_team_metrics(raw)
    team_ratings = rate_teams(team_metrics)
    build_team_ratings_workbook(team_ratings)

    print("Ranking leagues against each other ...")
    league_ratings = rate_leagues(raw, team_ratings)
    build_league_rankings_workbook(league_ratings)

    print("Done. Files saved to:", OUT_DIR.resolve())


if __name__ == "__main__":
    main()
