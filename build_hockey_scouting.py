"""
build_hockey_scouting.py — Ice hockey scouting framework applied to wide attackers

Hockey Concept → Football Translation
--------------------------------------
Corsi For%     → Carry For%: progressive carrying actions vs duels faced
Fenwick For%   → Unblocked Carry%: dribbles won / (dribbles won + duels lost)
Shooting%      → Shot% and xShot% (goals/shots, xG/shots)
Primary Points → Direct creation (xA, assists)  vs secondary (shot assists, key passes)
PPt%           → Primary creation as % of total offensive output
Zone scores    → O-Zone (threat), N-Zone (transition), D-Zone (defending)
Game Score     → Per-90 impact metric (Ben Vollack adaptation)
PDO            → Goals/xG ratio (conversion luck indicator, 100 = neutral)
Line class     → 1st/2nd/3rd/4th line based on WAR_per_90 percentile
HERO Chart     → 5 attribute scores: Shooting, Carrying, Creating, Duelling, Defending

Output: reports/Hockey_Style_Scouting.xlsx
  Sheet 1 — Main Ratings         : all players + primary hockey-style KPIs
  Sheet 2 — Corsi/Fenwick        : possession and shot metrics
  Sheet 3 — Points Breakdown     : primary vs secondary creation
  Sheet 4 — Zone Performance     : O/N/D zone scores
  Sheet 5 — Line Classification  : tiered player groupings with profiles
  Sheet 6 — Methodology          : all formulas defined
"""
from __future__ import annotations
import warnings; warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from pathlib import Path
from scipy.stats import norm, percentileofscore

try:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.fills import PatternFill as PF
    HAS_OPX = True
except ImportError:
    HAS_OPX = False

OUT_DIR = Path("reports")
OUT_DIR.mkdir(exist_ok=True)

# ── Constants ──────────────────────────────────────────────────────────────────
WIDE_ATK_POS = {"LAMF", "RAMF", "LW", "RW", "LWF", "RWF", "AMF", "LWB", "RWB"}
MIN_MINS     = 300
GOALS_PER_WIN = 3.0
REPL_PCT      = 15

NON_METRIC = {
    "Player","Team","Team within selected timeframe","Position",
    "Age","Market value","Contract expires","Birth country",
    "Passport country","Foot","Height","Weight","On loan",
}

# Line classification thresholds (WAR_per_90 percentile)
LINE_THRESHOLDS = {
    "1st Line":  75,   # Franchise / elite
    "2nd Line":  50,   # Reliable starter
    "3rd Line":  25,   # Depth contributor
    "4th Line":   0,   # Development / fringe
}

LINE_LABELS = {
    "1st Line": "Cornerstone",
    "2nd Line": "Reliable Starter",
    "3rd Line": "Depth Contributor",
    "4th Line": "Development / Fringe",
}

LINE_COLOURS = {
    "1st Line": "16A34A",   # green
    "2nd Line": "2563EB",   # blue
    "3rd Line": "D97706",   # amber
    "4th Line": "DC2626",   # red
}


# ── Data loading ───────────────────────────────────────────────────────────────

def _numeric(df):
    for col in df.columns:
        if col not in NON_METRIC:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def _gc(df, col, default=0.0):
    return df[col].fillna(default) if col in df.columns else pd.Series(default, index=df.index)


def load_pool():
    print("  Loading Wyscout files …")
    parts = []
    for f in sorted(Path("data/Wyscout DB").glob("*.xlsx")):
        if f.stem in ("Czech U17", "Czech U19"):
            continue
        try:
            d = _numeric(pd.read_excel(f))
            d["_league"] = f.stem
            d["_pos1"]   = d["Position"].astype(str).str.split(",").str[0].str.strip()
            parts.append(d)
        except Exception:
            pass

    df_all = pd.concat(parts, ignore_index=True)
    mask   = (df_all["_pos1"].isin(WIDE_ATK_POS) &
              (df_all["Minutes played"].fillna(0) >= MIN_MINS))
    pool   = df_all[mask].copy().reset_index(drop=True)
    print(f"  {len(pool):,} players  ·  {df_all['_league'].nunique()} leagues")
    return pool


def deduplicate(pool):
    before = len(pool)
    pool = (pool.sort_values("Minutes played", ascending=False)
               .drop_duplicates(subset=["Player",
                   pool.get("Team within selected timeframe", pool.get("Team","")).name
                   if hasattr(pool.get("Team within selected timeframe", None), "name")
                   else "Team within selected timeframe"
                   if "Team within selected timeframe" in pool.columns else "Player"],
                   keep="first")
               .reset_index(drop=True))
    removed = before - len(pool)
    if removed:
        print(f"  Deduplicated: {removed} duplicates removed")
    return pool


# ── WAR (same formula as WAR database) ────────────────────────────────────────

def compute_war(pool):
    xg  = _gc(pool, "xG per 90")
    xa  = _gc(pool, "xA per 90")
    sa  = _gc(pool, "Shot assists per 90")
    pr  = _gc(pool, "Progressive runs per 90")
    d   = _gc(pool, "Dribbles per 90")
    dp  = _gc(pool, "Successful dribbles, %", 50.0)
    off   = xg + xa * 0.85 + sa * 0.28
    carry = pr * 0.025 + (d * dp / 100.0) * 0.015
    total = off + carry
    repl  = float(np.percentile(total.dropna(), REPL_PCT))
    mins  = _gc(pool, "Minutes played", 450.0)
    n90   = mins / 90.0
    war   = (total - repl) * n90 / GOALS_PER_WIN
    war90 = (total - repl) / GOALS_PER_WIN
    return war, war90, total, repl


# ── Corsi / Fenwick equivalents ────────────────────────────────────────────────

def compute_corsi_fenwick(pool):
    """
    CF% (Corsi For%) → Carry For%
        Numerator:   progressive runs + dribbles won (carrying possession forward)
        Denominator: numerator + offensive duels lost (possession giveaways)
        Interpretation: >50% = positive possession impact

    FF% (Fenwick For%) → Unblocked Carry%
        Numerator:   dribbles won only (clean, uncontested carries)
        Denominator: total dribble attempts
        Interpretation: >50% = winning more carries than losing

    xGF% → Expected Goals For% (quality of possession)
        xG / (xG + estimated xGA)
        We approximate xGA from defensive duels lost × 0.05 (exposure proxy)

    SCF% → Scoring Chance For% (box presence)
        BoxTouches / (BoxTouches + Aerial duels lost in box proxy)

    HDCF% → High Danger Carry For% (shots in box as % of all shots)
    """
    pr   = _gc(pool, "Progressive runs per 90")
    d    = _gc(pool, "Dribbles per 90")
    dp   = _gc(pool, "Successful dribbles, %", 50.0) / 100.0
    d_won = d * dp                     # dribbles won/90
    d_lost = d * (1 - dp)              # dribbles lost/90

    od_pct = _gc(pool, "Offensive duels won, %", 50.0) / 100.0
    od_tot = _gc(pool, "Offensive duels per 90", 0.0)   # may not exist
    od_won = od_tot * od_pct
    od_lost = od_tot * (1 - od_pct)

    # CF%: carrying actions / (carrying + possession losses)
    carry_fwd = pr + d_won
    poss_loss = d_lost + od_lost.clip(0)
    denom_cf  = carry_fwd + poss_loss
    cf_pct    = np.where(denom_cf > 0, carry_fwd / denom_cf * 100, 50.0)

    # FF%: dribble success rate (pure Fenwick)
    denom_ff  = d.clip(lower=0.01)
    ff_pct    = (d_won / denom_ff * 100).clip(0, 100)

    # xGF%: shot quality possession
    xg  = _gc(pool, "xG per 90")
    xg_against_proxy = od_lost * 0.05   # rough: each lost duel ≈ 0.05 xGA
    denom_xgf = xg + xg_against_proxy + 0.01
    xgf_pct   = (xg / denom_xgf * 100).clip(0, 100)

    # SCF%: scoring chance for%
    bx = _gc(pool, "Touches in box per 90")
    sh = _gc(pool, "Shots per 90")
    scf_pct = (bx / (bx + 2.0) * 100).clip(0, 100)   # box presence index

    # HDCF%: high danger (shots on target / all shots)
    sot_pct = _gc(pool, "Shots on target, %", 33.3)
    hdcf_pct = sot_pct.clip(0, 100)

    # PDO: actual shooting% / expected shooting% (luck indicator)
    g90   = _gc(pool, "Goals per 90")
    sh_90 = _gc(pool, "Shots per 90").clip(lower=0.01)
    shoot_pct   = (g90 / sh_90 * 100)
    xshoot_pct  = (xg  / sh_90 * 100)
    with np.errstate(divide="ignore", invalid="ignore"):
        pdo = np.where(xshoot_pct > 0.1, shoot_pct / xshoot_pct * 100, np.nan)

    return pd.DataFrame({
        "CF%":      pd.Series(cf_pct,    index=pool.index).round(1),
        "FF%":      ff_pct.round(1),
        "xGF%":     xgf_pct.round(1),
        "SCF%":     scf_pct.round(1),
        "HDCF%":    hdcf_pct.round(1),
        "Shoot%":   shoot_pct.round(2),
        "xShoot%":  xshoot_pct.round(2),
        "PDO":      pd.Series(pdo, index=pool.index).round(1),
    })


# ── Game Score (Ben Vollack adaptation) ───────────────────────────────────────

def compute_game_score(pool):
    """
    Hockey Game Score per 90 (adapted):
    GS/90 = 0.75×G + 0.65×A + 0.45×xA + 0.30×ShotAssists + 0.20×KeyPasses
           + 0.06×Crosses + 0.04×ProgRuns + 0.02×Dribbles - 0.10×Fouls
    Approximates a player's per-90 "game value" on a single scale.
    """
    g   = _gc(pool, "Goals per 90")
    a   = _gc(pool, "Assists per 90")
    xa  = _gc(pool, "xA per 90")
    sa  = _gc(pool, "Shot assists per 90")
    kp  = _gc(pool, "Key passes per 90")
    cr  = _gc(pool, "Crosses per 90")
    pr  = _gc(pool, "Progressive runs per 90")
    d   = _gc(pool, "Dribbles per 90")
    fo  = _gc(pool, "Fouls per 90")

    gs = (g  * 0.75 + a  * 0.65 + xa * 0.45 +
          sa * 0.30 + kp * 0.20 + cr * 0.06 +
          pr * 0.04 + d  * 0.02 - fo * 0.10)
    return gs.round(4)


# ── Points breakdown (primary vs secondary) ───────────────────────────────────

def compute_points(pool):
    """
    Hockey: primary assists (A1) > secondary assists (A2).
    Football adaptation:
      Primary    = xA/90 (direct chance creation, last action before shot)
      A1_equiv   = Assists/90 (actual primary contribution)
      Secondary  = ShotAssists/90 (two actions before goal)
      A2_equiv   = (ShotAssists - Assists)/90 clipped ≥0
      Creation   = KeyPasses/90 (further removed)
      P/G        = (Goals + Assists) per appearance (≈ per 36 min if avg 36 min/game)
      PPt%       = Primary / (Primary + Secondary + Creation) × 100
    """
    g  = _gc(pool, "Goals per 90")
    a  = _gc(pool, "Assists per 90")
    xa = _gc(pool, "xA per 90")
    sa = _gc(pool, "Shot assists per 90")
    kp = _gc(pool, "Key passes per 90")
    mins = _gc(pool, "Minutes played", 450.0)

    # Points per game (approx game = 90 min)
    games = (mins / 90.0).clip(lower=0.01)
    total_g  = g * mins / 90.0
    total_a  = a * mins / 90.0
    p_per_g  = (total_g + total_a) / games

    # Primary / secondary split
    primary   = xa                              # direct pre-assist
    secondary = (sa - xa).clip(lower=0)        # shot assists beyond xA
    deep      = kp * 0.4                       # key passes (deeper creation)
    denom     = (primary + secondary + deep + 0.001)
    ppt_pct   = (primary / denom * 100).clip(0, 100)
    a1_pct    = (a / (a + (sa - a).clip(0) + 0.001) * 100).clip(0, 100)

    return pd.DataFrame({
        "G_per_90":      g.round(3),
        "A_per_90":      a.round(3),
        "xA_per_90":     xa.round(3),
        "SA_per_90":     sa.round(3),
        "KP_per_90":     kp.round(3),
        "P/G":           p_per_g.round(3),
        "Primary%":      ppt_pct.round(1),
        "A1%":           a1_pct.round(1),
        "G+xA/90":       (g + xa).round(3),
        "xG+xA/90":      (_gc(pool,"xG per 90") + xa).round(3),
    })


# ── Zone performance ───────────────────────────────────────────────────────────

def compute_zones(pool):
    """
    O-Zone (Offensive Zone): threat generation in and around the box
    N-Zone (Neutral Zone):   transition and carrying through mid-third
    D-Zone (Defensive Zone): defensive contribution and recovery
    """
    # Raw O-Zone inputs
    xg  = _gc(pool, "xG per 90")
    g   = _gc(pool, "Goals per 90")
    sh  = _gc(pool, "Shots per 90")
    bx  = _gc(pool, "Touches in box per 90")
    xa  = _gc(pool, "xA per 90")
    cr  = _gc(pool, "Crosses per 90")
    ca  = _gc(pool, "Accurate crosses, %", 50.0)
    sa  = _gc(pool, "Shot assists per 90")
    o_raw = (xg * 2.0 + g * 1.5 + sh * 0.3 + bx * 0.5 +
             xa * 1.8 + cr * (ca / 100.0) * 0.8 + sa * 0.5)

    # Raw N-Zone inputs (transition and carrying)
    pr  = _gc(pool, "Progressive runs per 90")
    ac  = _gc(pool, "Accelerations per 90")
    d   = _gc(pool, "Dribbles per 90")
    dp  = _gc(pool, "Successful dribbles, %", 50.0)
    od  = _gc(pool, "Offensive duels won, %", 50.0)
    kp  = _gc(pool, "Key passes per 90")
    n_raw = (pr * 1.0 + ac * 0.6 + d * (dp / 100.0) * 0.8 +
             od * 0.015 + kp * 0.4)

    # Raw D-Zone inputs
    da  = _gc(pool, "Successful defensive actions per 90")
    ic  = _gc(pool, "Interceptions per 90")
    dd  = _gc(pool, "Defensive duels won, %", 50.0)
    ae  = _gc(pool, "Aerial duels won, %", 50.0)
    fo  = _gc(pool, "Fouls per 90")
    d_raw = (da * 1.2 + ic * 1.0 + dd * 0.015 + ae * 0.008 - fo * 0.1)

    # Convert raw scores to DB percentiles (0-100)
    def to_pct(series):
        vals = series.values
        return pd.Series(
            [round(float(np.mean(vals < v) * 100), 1) if pd.notna(v) else 50.0
             for v in vals],
            index=series.index,
        )

    o_pct = to_pct(o_raw)
    n_pct = to_pct(n_raw)
    d_pct = to_pct(d_raw)

    # Zone Balance: how consistently active across all three zones
    zone_mean = (o_pct + n_pct + d_pct) / 3
    zone_std  = pd.concat([o_pct, n_pct, d_pct], axis=1).std(axis=1)
    balance   = (100 - zone_std * 1.5).clip(0, 100).round(1)

    # Dominant Zone label
    def dominant(row):
        scores = {"O": row["OZone"], "N": row["NZone"], "D": row["DZone"]}
        return max(scores, key=scores.get) + "-Zone Specialist"

    df = pd.DataFrame({
        "OZone":      o_pct.round(1),
        "NZone":      n_pct.round(1),
        "DZone":      d_pct.round(1),
        "ZoneBalance":balance,
        "OvrlZone":   zone_mean.round(1),
    })
    df["DomZone"] = df.apply(dominant, axis=1)
    return df


# ── Line classification ────────────────────────────────────────────────────────

def classify_line(war90_series):
    pct_arr = war90_series.values
    lines   = []
    for w in pct_arr:
        if pd.isna(w):
            lines.append("4th Line")
            continue
        p = float(np.mean(pct_arr < w) * 100)
        if p >= LINE_THRESHOLDS["1st Line"]:
            lines.append("1st Line")
        elif p >= LINE_THRESHOLDS["2nd Line"]:
            lines.append("2nd Line")
        elif p >= LINE_THRESHOLDS["3rd Line"]:
            lines.append("3rd Line")
        else:
            lines.append("4th Line")
    return pd.Series(lines, index=war90_series.index)


def age_trajectory(age):
    if pd.isna(age): return "→ Unknown"
    a = float(age)
    if a <= 21: return "↑ Rising"
    if a <= 25: return "↑ Developing"
    if a <= 29: return "→ Peak"
    if a <= 32: return "↓ Declining"
    return "↓ Veteran"


# ── HERO Chart scores (5 attributes, hockey HERO visual adaptation) ────────────

def compute_hero(pool):
    """
    Hockey HERO chart: 5 attributes in a pentagon.
    Football adaptation:
      Shooting     → xG per shot quality + goals contribution
      Playmaking   → xA + primary creation
      Skating      → Athleticism: prog runs + accelerations (speed/motor)
      Physicality  → Duel dominance (off + def + aerial)
      Defence      → Defensive actions + interceptions
    Each scored 0-100 (DB percentile).
    """
    def to_pct(series):
        vals = series.values
        return pd.Series(
            [round(float(np.mean(vals < v) * 100), 1) if pd.notna(v) else 50.0
             for v in vals],
            index=series.index,
        )

    # Shooting (goal threat quality)
    xg  = _gc(pool, "xG per 90")
    sh  = _gc(pool, "Shots per 90").clip(lower=0.01)
    g   = _gc(pool, "Goals per 90")
    shooting_raw = xg * 0.6 + (xg / sh) * 2.0 + g * 0.4

    # Playmaking
    xa  = _gc(pool, "xA per 90")
    sa  = _gc(pool, "Shot assists per 90")
    kp  = _gc(pool, "Key passes per 90")
    playmaking_raw = xa * 2.0 + sa * 0.7 + kp * 0.4

    # Skating (athleticism / motor)
    pr  = _gc(pool, "Progressive runs per 90")
    ac  = _gc(pool, "Accelerations per 90")
    d   = _gc(pool, "Dribbles per 90")
    dp  = _gc(pool, "Successful dribbles, %", 50.0)
    skating_raw = pr + ac * 0.6 + d * (dp / 100.0) * 0.5

    # Physicality (duel dominance)
    od  = _gc(pool, "Offensive duels won, %", 50.0)
    dd  = _gc(pool, "Defensive duels won, %", 50.0)
    ae  = _gc(pool, "Aerial duels won, %", 50.0)
    phys_raw = (od + dd + ae * 0.7) / 3.0

    # Defence
    da  = _gc(pool, "Successful defensive actions per 90")
    ic  = _gc(pool, "Interceptions per 90")
    def_raw = da + ic * 1.1

    return pd.DataFrame({
        "H_Shooting":    to_pct(shooting_raw).round(1),
        "H_Playmaking":  to_pct(playmaking_raw).round(1),
        "H_Skating":     to_pct(skating_raw).round(1),
        "H_Physicality": to_pct(phys_raw).round(1),
        "H_Defence":     to_pct(def_raw).round(1),
    })


# ── Assemble all sheets ────────────────────────────────────────────────────────

def build_all(pool):
    war, war90, total_rate, repl = compute_war(pool)
    mins = _gc(pool, "Minutes played", 450.0)
    n90  = (mins / 90.0).round(2)

    base = pd.DataFrame({
        "Player":   pool["Player"],
        "Team":     pool.get("Team within selected timeframe",
                             pool.get("Team", "")),
        "League":   pool["_league"],
        "Position": pool["Position"],
        "Age":      _gc(pool, "Age", np.nan).replace(0, np.nan),
        "Minutes":  mins.round(0).astype(int),
        "90s":      n90,
    })

    # WAR + line
    base["WAR"]     = war.round(3)
    base["WAR/90"]  = war90.round(4)
    base["Line"]    = classify_line(war90)
    base["Traj"]    = base["Age"].apply(age_trajectory)

    # Game Score
    base["GS/90"]   = compute_game_score(pool).round(3)

    # Corsi/Fenwick
    cf = compute_corsi_fenwick(pool)

    # Points breakdown
    pts = compute_points(pool)

    # Zone scores
    zones = compute_zones(pool)

    # HERO
    hero = compute_hero(pool)

    # League rank within same league (by WAR)
    base["Lg_rank"] = (base.groupby("League")["WAR"]
                           .rank(ascending=False, method="min").astype(int))
    base["Lg_n"]    = base.groupby("League")["WAR"].transform("count").astype(int)

    # DB rank
    war_arr = war.values
    base["DB_rank"]  = base["WAR"].rank(ascending=False, method="min").astype(int)
    base["DB_pctile"]= base["WAR"].apply(
        lambda w: round(float(np.mean(war_arr < w) * 100), 1) if pd.notna(w) else np.nan
    )

    # ── Sheet 1: Main Ratings ────────────────────────────────────────────────
    main = pd.concat([base, cf[["CF%","FF%","xGF%","PDO"]], pts[["P/G","Primary%"]],
                      zones[["OZone","NZone","DZone","DomZone"]],
                      hero], axis=1)

    # ── Sheet 2: Corsi/Fenwick ───────────────────────────────────────────────
    s2 = pd.concat([base[["Player","Team","League","Age","Minutes",
                           "WAR","WAR/90","Line"]], cf], axis=1)

    # ── Sheet 3: Points Breakdown ────────────────────────────────────────────
    s3 = pd.concat([base[["Player","Team","League","Age","Minutes","WAR","Line"]], pts], axis=1)

    # ── Sheet 4: Zone Performance ────────────────────────────────────────────
    s4 = pd.concat([base[["Player","Team","League","Age","Minutes","WAR","Line","GS/90"]], zones], axis=1)

    # ── Sheet 5: Line Classification ────────────────────────────────────────
    line_cols = ["Player","Team","League","Position","Age","Minutes","WAR","WAR/90",
                 "Line","Traj","DB_rank","DB_pctile","Lg_rank","Lg_n"]
    s5 = base[line_cols].copy()
    s5["Line_Label"] = s5["Line"].map(LINE_LABELS)
    s5["GS/90"]      = base["GS/90"]
    s5 = pd.concat([s5, zones[["OZone","NZone","DZone","ZoneBalance"]], hero], axis=1)

    # Sort all by WAR
    for df in [main, s2, s3, s4, s5]:
        df.sort_values("WAR", ascending=False, inplace=True)
        df.reset_index(drop=True, inplace=True)
        df.insert(0, "Rank", range(1, len(df)+1))

    return main, s2, s3, s4, s5, repl


# ── Excel output ───────────────────────────────────────────────────────────────

METHODOLOGY = [
    ("WAR",         "Wins Above Replacement: (rate − 15th-pctile) × 90s ÷ 3 goals/win"),
    ("WAR/90",      "(rate − replacement) ÷ 3  — quality, not volume"),
    ("GS/90",       "Game Score: 0.75G + 0.65A + 0.45xA + 0.30SA + 0.20KP + 0.06Cr + 0.04PR + 0.02D − 0.10Fo"),
    ("Line",        "1st (WAR/90 ≥75th DB pctile) · 2nd (≥50th) · 3rd (≥25th) · 4th (<25th)"),
    ("Traj",        "Age trajectory: ↑ Rising ≤21  ↑ Developing ≤25  → Peak ≤29  ↓ Declining ≤32  ↓ Veteran 33+"),
    ("", ""),
    ("CF% (Corsi)", "Carry For%: (ProgRuns + DribWon) / (ProgRuns + DribWon + DribLost + ODuelsLost) × 100"),
    ("FF% (Fenwick)","Unblocked Carry%: DribWon / Dribble_Attempts × 100  (Dribble Success Rate)"),
    ("xGF%",        "Expected Goals For%: xG / (xG + OffDuelsLost×0.05) × 100"),
    ("SCF%",        "Scoring Chance For%: BoxTouches / (BoxTouches + 2) × 100  (box presence index)"),
    ("HDCF%",       "High Danger Carry For%: Shots on Target % (quality of shot selection)"),
    ("Shoot%",      "Goals / Shots × 100"),
    ("xShoot%",     "xG / Shots × 100"),
    ("PDO",         "Shoot% / xShoot% × 100  (conversion luck: 100 = neutral, >100 = overperforming xG)"),
    ("", ""),
    ("P/G",         "Points per game (Goals + Assists per 90-min appearance)"),
    ("Primary%",    "xA / (xA + ClippedShotAssists + KeyPasses×0.4) × 100  — direct creation dominance"),
    ("A1%",         "Assists / (Assists + 'secondary assists') × 100"),
    ("G+xA/90",     "Goals + Expected Assists per 90  (best single offensive rate stat)"),
    ("xG+xA/90",    "Expected Goals + Expected Assists per 90"),
    ("", ""),
    ("OZone",       "DB percentile: xG×2 + G×1.5 + Shots×0.3 + BoxTouches×0.5 + xA×1.8 + AccCrosses×0.8 + SA×0.5"),
    ("NZone",       "DB percentile: ProgRuns + Accels×0.6 + EffDrib×0.8 + OffDuel%×0.015 + KP×0.4"),
    ("DZone",       "DB percentile: DefActions×1.2 + Interceptions + DefDuel%×0.015 + Aerial%×0.008 − Fouls×0.1"),
    ("ZoneBalance", "100 − StdDev(O,N,D) × 1.5  — consistent across zones = high balance"),
    ("DomZone",     "Zone with highest score"),
    ("", ""),
    ("H_Shooting",  "HERO: DB pctile of xG×0.6 + (xG/Shot)×2 + G×0.4"),
    ("H_Playmaking","HERO: DB pctile of xA×2 + ShotAssists×0.7 + KeyPasses×0.4"),
    ("H_Skating",   "HERO: DB pctile of ProgRuns + Accels×0.6 + EffDrib×0.5"),
    ("H_Physicality","HERO: DB pctile of mean(OffDuel%, DefDuel%, Aerial%×0.7)"),
    ("H_Defence",   "HERO: DB pctile of DefActions + Interceptions×1.1"),
]


def _style(wb, sheet: str, df: pd.DataFrame,
           color_cols=None, left_ci=None, row_colors: dict | None = None):
    if not HAS_OPX: return
    ws = wb[sheet]
    hfont = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    hfill = PatternFill("solid", fgColor="1D4ED8")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    calign = Alignment(horizontal="center", vertical="center")
    lalign = Alignment(horizontal="left",   vertical="center")
    light  = PatternFill("solid", fgColor="F7F9FC")

    for cell in ws[1]:
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    cols = list(df.columns)
    line_ci = cols.index("Line") + 1 if "Line" in cols else None

    # Line → colour map
    line_colours_fill = {k: PatternFill("solid", fgColor=v) for k, v in LINE_COLOURS.items()}

    for ri, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        for ci, cell in enumerate(row, 1):
            cell.alignment = calign
            if ri % 2 == 0:
                cell.fill = light
        for ci in (left_ci or [2, 3, 4]):
            if ci <= len(row):
                row[ci-1].alignment = lalign
        # colour Line cell
        if line_ci and line_ci <= len(row):
            val = row[line_ci - 1].value
            if val in line_colours_fill:
                row[line_ci - 1].fill = line_colours_fill[val]
                row[line_ci - 1].font = Font(bold=True, color="FFFFFF",
                                              name="Calibri", size=9)

    for ci, col in enumerate(cols, 1):
        try:
            w = max(len(str(col)) + 2,
                    min(int(df[col].astype(str).str.len().max()) + 2, 38))
        except Exception:
            w = 12
        ws.column_dimensions[get_column_letter(ci)].width = w

    for col_name, direction in (color_cols or []):
        if col_name not in cols: continue
        ci = cols.index(col_name) + 1
        ltr = get_column_letter(ci)
        rng = f"{ltr}2:{ltr}{ws.max_row}"
        if direction == "rg":
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="percentile", start_value=5,  start_color="DC2626",
                mid_type="percentile",   mid_value=50,   mid_color="D97706",
                end_type="percentile",   end_value=95,   end_color="16A34A"))
        elif direction == "gr":
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="percentile", start_value=5,  start_color="16A34A",
                mid_type="percentile",   mid_value=50,   mid_color="D97706",
                end_type="percentile",   end_value=95,   end_color="DC2626"))
        elif direction == "50":   # centred on 100 (PDO-style)
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="num", start_value=70, start_color="DC2626",
                mid_type="num",   mid_value=100,  mid_color="D97706",
                end_type="num",   end_value=130,  end_color="16A34A"))


def save(main, s2, s3, s4, s5, repl, path):
    method_df = pd.DataFrame(METHODOLOGY, columns=["Metric", "Formula / Description"])

    with pd.ExcelWriter(path, engine="openpyxl") as w:
        main.to_excel(w, index=False, sheet_name="Main Ratings")
        s2.to_excel(w, index=False, sheet_name="Corsi-Fenwick")
        s3.to_excel(w, index=False, sheet_name="Points Breakdown")
        s4.to_excel(w, index=False, sheet_name="Zone Performance")
        s5.to_excel(w, index=False, sheet_name="Line Classification")
        method_df.to_excel(w, index=False, sheet_name="Methodology")

        if HAS_OPX:
            rg = "rg"; gr = "gr"
            _style(w.book, "Main Ratings", main,
                   [("WAR",rg),("WAR/90",rg),("GS/90",rg),
                    ("CF%",rg),("FF%",rg),("PDO","50"),
                    ("P/G",rg),("Primary%",rg),
                    ("OZone",rg),("NZone",rg),("DZone",rg),
                    ("H_Shooting",rg),("H_Playmaking",rg),
                    ("H_Skating",rg),("H_Physicality",rg),("H_Defence",rg),
                    ("DB_rank",gr),("Lg_rank",gr)])

            _style(w.book, "Corsi-Fenwick", s2,
                   [("CF%",rg),("FF%",rg),("xGF%",rg),
                    ("SCF%",rg),("HDCF%",rg),("PDO","50"),
                    ("WAR",rg),("WAR/90",rg)])

            _style(w.book, "Points Breakdown", s3,
                   [("G_per_90",rg),("A_per_90",rg),("xA_per_90",rg),
                    ("P/G",rg),("Primary%",rg),("A1%",rg),
                    ("G+xA/90",rg),("xG+xA/90",rg)])

            _style(w.book, "Zone Performance", s4,
                   [("OZone",rg),("NZone",rg),("DZone",rg),
                    ("ZoneBalance",rg),("OvrlZone",rg),("GS/90",rg)])

            _style(w.book, "Line Classification", s5,
                   [("WAR",rg),("WAR/90",rg),("DB_pctile",rg),
                    ("DB_rank",gr),("Lg_rank",gr),
                    ("OZone",rg),("NZone",rg),("DZone",rg),("ZoneBalance",rg),
                    ("H_Shooting",rg),("H_Playmaking",rg),
                    ("H_Skating",rg),("H_Physicality",rg),("H_Defence",rg)])

            # Methodology sheet
            ws_m = w.book["Methodology"]
            for cell in ws_m[1]:
                cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
                cell.fill = PatternFill("solid", fgColor="1D4ED8")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws_m.columns:
                ww = max(len(str(c.value or "")) for c in col) + 4
                ws_m.column_dimensions[col[0].column_letter].width = min(ww, 100)

    print(f"  Saved → {path}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("Building Hockey-Style Scouting database …")
    pool = load_pool()

    # Deduplicate cleanly
    team_col = "Team within selected timeframe" if "Team within selected timeframe" in pool.columns else "Team"
    before   = len(pool)
    pool     = (pool.sort_values("Minutes played", ascending=False)
                    .drop_duplicates(subset=["Player", team_col], keep="first")
                    .reset_index(drop=True))
    print(f"  Deduplicated: {before - len(pool)} duplicates removed  →  {len(pool):,} players")

    main_df, s2, s3, s4, s5, repl = build_all(pool)

    print(f"\n  Top 10 by GS/90:")
    for _, r in main_df.nlargest(10, "GS/90").iterrows():
        print(f"    {str(r['Player']):<28}  {str(r['League']):<18}  "
              f"GS/90={r['GS/90']:+.3f}  Line={r['Line']}")

    # Print Barát
    barat = main_df[main_df["Player"].astype(str).str.startswith("D. Bar") &
                    main_df["Team"].astype(str).str.contains("Slov")]
    if not barat.empty:
        r = barat.iloc[0]
        print(f"\n  D. Barát:")
        print(f"    WAR={r['WAR']:+.3f}  GS/90={r['GS/90']:+.3f}  Line={r['Line']}")
        print(f"    CF%={r['CF%']}  FF%={r['FF%']}  PDO={r['PDO']}")
        print(f"    OZone={r['OZone']}  NZone={r['NZone']}  DZone={r['DZone']}")
        print(f"    HERO: Shoot={r['H_Shooting']}  Play={r['H_Playmaking']}  "
              f"Skate={r['H_Skating']}  Phys={r['H_Physicality']}  Def={r['H_Defence']}")

    out = OUT_DIR / "Hockey_Style_Scouting.xlsx"
    save(main_df, s2, s3, s4, s5, repl, out)
    print(f"\n  {len(main_df):,} players  ·  {main_df['League'].nunique()} leagues")
    print("  Sheets: Main Ratings | Corsi-Fenwick | Points Breakdown | "
          "Zone Performance | Line Classification | Methodology")


if __name__ == "__main__":
    main()
