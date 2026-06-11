"""
Anomaly Platform — Statistical outlier detection on Wyscout per-90 data.
z-score / MAD / IQR detection with position-specific metric sets.
"""
from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import altair as alt
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import streamlit as st
from scipy.stats import norm as scipy_norm

APP_DIR = Path(__file__).parent
os.environ.setdefault("MPLCONFIGDIR", "/tmp/anomaly_platform_matplotlib")
matplotlib.use("Agg")
plt.rcParams.update({
    "font.family": "monospace",
    "axes.facecolor":  "#090c18",
    "figure.facecolor": "#090c18",
    "text.color": "#8898b8",
    "axes.labelcolor": "#8898b8",
    "xtick.color": "#4a5880",
    "ytick.color": "#4a5880",
    "axes.edgecolor": "#141c30",
    "grid.color": "#141c30",
    "grid.linewidth": 0.6,
})

from scouting_model import (
    AnomalyEngine,
    SimilarityEngine,
    SetPieceAnalyzer,
    SET_PIECE_ROLES,
)
from wyscout_model import (
    load_wyscout_raw,
    available_leagues as wyscout_available_leagues,
)

# ── Position-specific metric sets ─────────────────────────────────────────────
POSITION_METRICS: dict[str, list[str]] = {
    "ST": [
        "Goals per 90", "Non-penalty goals per 90", "xG per 90",
        "Shots per 90", "Shots on target, %", "Goal conversion, %",
        "Touches in box per 90", "Aerial duels won, %",
        "Dribbles per 90", "Successful dribbles, %",
        "Progressive runs per 90", "xA per 90",
    ],
    "W": [
        "Goals per 90", "xG per 90", "Assists per 90", "xA per 90",
        "Key passes per 90", "Dribbles per 90", "Successful dribbles, %",
        "Crosses per 90", "Accurate crosses, %",
        "Progressive runs per 90", "Touches in box per 90",
        "Successful defensive actions per 90",
    ],
    "AM": [
        "Key passes per 90", "xA per 90", "Assists per 90",
        "Smart passes per 90", "Accurate smart passes, %",
        "Goals per 90", "xG per 90", "Touches in box per 90",
        "Dribbles per 90", "Progressive passes per 90",
        "Through passes per 90", "Successful dribbles, %",
    ],
    "CM": [
        "Passes per 90", "Accurate passes, %",
        "Forward passes per 90", "Accurate forward passes, %",
        "Progressive passes per 90", "Key passes per 90",
        "xA per 90", "Progressive runs per 90",
        "Successful defensive actions per 90", "Defensive duels won, %",
        "Interceptions per 90", "Duels won, %",
    ],
    "DM": [
        "Successful defensive actions per 90", "Defensive duels per 90",
        "Defensive duels won, %", "Interceptions per 90",
        "PAdj Interceptions", "Aerial duels won, %", "Duels won, %",
        "Passes per 90", "Accurate passes, %",
        "Progressive passes per 90", "Key passes per 90", "Fouls per 90",
    ],
    "FB": [
        "Crosses per 90", "Accurate crosses, %",
        "xA per 90", "Assists per 90", "Key passes per 90",
        "Progressive runs per 90", "Dribbles per 90",
        "Successful defensive actions per 90", "Defensive duels won, %",
        "Aerial duels won, %", "Accurate passes, %",
        "Progressive passes per 90",
    ],
    "CB": [
        "Successful defensive actions per 90", "Defensive duels per 90",
        "Defensive duels won, %", "Aerial duels per 90", "Aerial duels won, %",
        "Interceptions per 90", "PAdj Interceptions", "Shots blocked per 90",
        "Fouls per 90", "Accurate passes, %",
        "Accurate forward passes, %", "Progressive passes per 90",
    ],
    "GK": [
        "Save rate, %", "Prevented goals per 90", "Conceded goals per 90",
        "Shots against per 90", "Clean sheets",
        "Exits per 90", "Aerial duels per 90.1",
        "Accurate passes, %", "Accurate long passes, %",
        "Back passes received as GK per 90",
    ],
}

SIMILARITY_FEATURES = [
    "Goals per 90", "xG per 90", "Assists per 90", "xA per 90",
    "Key passes per 90", "Passes per 90", "Accurate passes, %",
    "Dribbles per 90", "Successful dribbles, %",
    "Successful defensive actions per 90", "Defensive duels won, %",
    "Aerial duels won, %", "Progressive passes per 90",
    "Progressive runs per 90", "Touches in box per 90",
]

ANOMALY_COLORS: dict[str, str] = {
    "Hidden Gem":               "#00e5c4",
    "Specialist Elite":         "#ff3b5c",
    "Multi-dimensional":        "#ff8c00",
    "Age-adjusted Gem":         "#38b6ff",
    "Consistent Overperformer": "#a78bfa",
}

POSITION_ORDER = ["GK", "CB", "FB", "DM", "CM", "AM", "W", "ST"]
POSITION_COLORS = {
    "GK": "#556070", "CB": "#2f5f98", "FB": "#007ca6",
    "DM": "#4a7a23", "CM": "#2f7a50", "AM": "#c07800",
    "W":  "#cc5030", "ST": "#aa2000",
}

BG      = "#060810"
SURFACE = "#090c18"
SURF2   = "#0d1120"
BORDER  = "#151c30"
ACCENT  = "#3d7fff"
DANGER  = "#ff3b5c"
TEAL    = "#00e5c4"
TEXT    = "#c4ceea"
MUTED   = "#3d4b6a"
MONO    = "'JetBrains Mono','Fira Code','Cascadia Code',monospace"


# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Anomaly Platform",
    page_icon="⬡",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;700&display=swap');

html, body,
[data-testid="stApp"],
[data-testid="stAppViewContainer"],
[data-testid="stMain"] {{
    background-color: {BG} !important;
    color: {TEXT} !important;
    font-family: 'Inter', system-ui, sans-serif !important;
}}
[data-testid="stSidebar"] {{
    background-color: #050710 !important;
    border-right: 1px solid {BORDER} !important;
}}
[data-testid="stSidebar"] * {{ color: #8898b8 !important; }}
[data-testid="stSidebar"] .stMarkdown p {{
    font-size: 0.72rem !important;
    color: {MUTED} !important;
    font-family: {MONO} !important;
}}
[data-testid="stSidebar"] label {{
    font-size: 0.62rem !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.11em !important;
    color: {MUTED} !important;
}}
.main .block-container {{
    padding: 0 2.4rem 2rem !important;
    max-width: 100% !important;
}}
#MainMenu, footer, header,
[data-testid="stToolbar"],
[data-testid="stDecoration"],
[data-testid="stStatusWidget"] {{ display: none !important; }}

/* ── Page header strip ── */
.page-header {{
    background: linear-gradient(90deg, #0d1120 0%, {BG} 100%);
    border-bottom: 1px solid {BORDER};
    padding: 22px 0 18px 0;
    margin-bottom: 28px;
}}
.page-header-eyebrow {{
    font-size: 0.6rem;
    font-weight: 700;
    letter-spacing: 0.16em;
    text-transform: uppercase;
    color: {MUTED};
    font-family: {MONO};
    margin-bottom: 4px;
}}
.page-header-title {{
    font-size: 1.85rem;
    font-weight: 800;
    color: #e8eeff;
    line-height: 1.1;
    letter-spacing: -0.02em;
}}
.page-header-sub {{
    font-size: 0.78rem;
    color: {MUTED};
    margin-top: 5px;
    font-family: {MONO};
    font-size: 0.72rem;
}}
.page-accent {{ color: {ACCENT}; }}

/* ── KPI cards ── */
.kpi-grid {{ display: flex; gap: 12px; margin-bottom: 28px; }}
.kpi-card {{
    background: {SURFACE};
    border: 1px solid {BORDER};
    border-top: 2px solid {ACCENT};
    border-radius: 6px;
    padding: 16px 20px;
    flex: 1;
    min-width: 0;
}}
.kpi-card.danger {{ border-top-color: {DANGER}; }}
.kpi-card.teal   {{ border-top-color: {TEAL}; }}
.kpi-card.purple {{ border-top-color: #a78bfa; }}
.kpi-num {{
    font-size: 2.1rem;
    font-weight: 800;
    color: #e8eeff;
    line-height: 1;
    font-family: {MONO};
    letter-spacing: -0.03em;
}}
.kpi-label {{
    font-size: 0.6rem;
    color: {MUTED};
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-top: 5px;
    font-weight: 600;
}}
.kpi-sub {{
    font-size: 0.65rem;
    color: #4a5880;
    margin-top: 2px;
    font-family: {MONO};
}}

/* ── Anomaly type badges ── */
.type-row {{ display: flex; gap: 10px; margin-bottom: 24px; flex-wrap: wrap; }}
.type-badge {{
    background: {SURF2};
    border-radius: 6px;
    padding: 12px 18px;
    border: 1px solid {BORDER};
    border-left: 3px solid;
    flex: 1;
    min-width: 120px;
}}
.type-count {{
    font-size: 1.65rem;
    font-weight: 800;
    font-family: {MONO};
    line-height: 1;
    letter-spacing: -0.03em;
}}
.type-name {{
    font-size: 0.58rem;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    font-weight: 700;
    color: {MUTED};
    margin-top: 5px;
}}
.type-pct {{
    font-size: 0.62rem;
    font-family: {MONO};
    color: #3d4b6a;
    margin-top: 2px;
}}

/* ── Stat table (μ σ z) ── */
.stat-table {{
    width: 100%;
    border-collapse: collapse;
    font-family: {MONO};
    font-size: 0.75rem;
}}
.stat-table th {{
    background: {SURF2};
    color: {MUTED};
    text-align: right;
    padding: 6px 10px;
    font-size: 0.6rem;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    border-bottom: 1px solid {BORDER};
    font-weight: 700;
}}
.stat-table th:first-child {{ text-align: left; }}
.stat-table td {{
    padding: 5px 10px;
    border-bottom: 1px solid #0f1422;
    text-align: right;
    color: {TEXT};
}}
.stat-table td:first-child {{ text-align: left; color: #8898b8; max-width: 200px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
.stat-table tr:hover td {{ background: #0d1120; }}
.z-high  {{ color: {DANGER} !important; font-weight: 700; }}
.z-med   {{ color: #ff8c00 !important; }}
.z-pos   {{ color: {TEAL} !important; }}
.z-neg   {{ color: {MUTED} !important; }}

/* ── Player header card ── */
.player-card {{
    background: {SURFACE};
    border: 1px solid {BORDER};
    border-radius: 8px;
    padding: 22px 28px;
    margin-bottom: 24px;
    position: relative;
    overflow: hidden;
}}
.player-card::before {{
    content: '';
    position: absolute;
    top: 0; left: 0;
    width: 4px; height: 100%;
    background: {ACCENT};
}}
.player-name {{
    font-size: 1.55rem;
    font-weight: 800;
    color: #eef2ff;
    letter-spacing: -0.02em;
}}
.player-meta {{
    font-size: 0.72rem;
    color: {MUTED};
    margin-top: 6px;
    font-family: {MONO};
}}
.stat-pill {{
    display: inline-block;
    background: {SURF2};
    border: 1px solid {BORDER};
    border-radius: 4px;
    padding: 4px 10px;
    font-family: {MONO};
    font-size: 0.68rem;
    color: {TEXT};
    margin-right: 8px;
    margin-top: 10px;
}}
.stat-pill .label {{ color: {MUTED}; font-size: 0.6rem; display: block; margin-bottom: 1px; }}
.stat-pill .value {{ font-weight: 700; color: #e8eeff; }}

/* ── Formula block ── */
.formula-box {{
    background: {SURF2};
    border: 1px solid {BORDER};
    border-left: 3px solid {ACCENT};
    border-radius: 0 6px 6px 0;
    padding: 12px 16px;
    font-family: {MONO};
    font-size: 0.72rem;
    color: #8898b8;
    margin: 12px 0;
    line-height: 1.8;
}}
.formula-box .highlight {{ color: {ACCENT}; font-weight: 700; }}
.formula-box .danger    {{ color: {DANGER}; font-weight: 700; }}

/* ── Section label ── */
.section-label {{
    font-size: 0.6rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.14em;
    color: {MUTED};
    border-bottom: 1px solid {BORDER};
    padding-bottom: 6px;
    margin: 20px 0 14px 0;
    font-family: {MONO};
}}

/* ── Sidebar brand ── */
.sidebar-brand {{
    padding: 20px 0 14px 2px;
    border-bottom: 1px solid {BORDER};
    margin-bottom: 14px;
}}
.brand-title {{
    font-size: 0.78rem;
    font-weight: 800;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: {ACCENT} !important;
    font-family: {MONO};
}}
.brand-sub {{
    font-size: 0.58rem;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: {MUTED} !important;
    margin-top: 2px;
    font-family: {MONO};
}}

div[data-testid="stDataFrame"] {{
    border: 1px solid {BORDER} !important;
    border-radius: 6px !important;
}}
[data-testid="stTabs"] [role="tab"] {{
    font-family: {MONO} !important;
    font-size: 0.72rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.06em !important;
}}
</style>
""", unsafe_allow_html=True)


# ── Matplotlib figure helpers ─────────────────────────────────────────────────

def _fig_z_distribution(peak_z: pd.Series, threshold: float) -> plt.Figure:
    """Empirical peak-z distribution vs. standard normal with threshold annotation."""
    fig, ax = plt.subplots(figsize=(9, 2.8))
    vals = peak_z.dropna().values
    bins = np.linspace(vals.min() - 0.2, vals.max() + 0.2, 55)
    ax.hist(vals, bins=bins, density=True, color="#1a2540", edgecolor="#242f50", linewidth=0.5, zorder=2)
    x = np.linspace(-4.5, vals.max() + 0.5, 300)
    ax.plot(x, scipy_norm.pdf(x), color=ACCENT, linewidth=1.4, alpha=0.85, label="𝒩(0,1)", zorder=3)
    ax.axvline(threshold, color=DANGER, linewidth=1.5, linestyle="--", zorder=4)
    ax.fill_between(x[x >= threshold], scipy_norm.pdf(x[x >= threshold]),
                    alpha=0.14, color=DANGER, zorder=1)
    ymax = ax.get_ylim()[1]
    ax.text(threshold + 0.05, ymax * 0.82,
            f"α = {threshold:.1f}\np ≈ {(1 - scipy_norm.cdf(threshold)) * 100:.1f}%",
            color=DANGER, fontsize=7.5, va="top", fontfamily="monospace")
    ax.set_xlabel("Peak z-score  max_j z_ij", fontsize=8)
    ax.set_ylabel("Density", fontsize=8)
    ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.2f"))
    ax.grid(axis="y", zorder=0)
    ax.legend(fontsize=7.5, framealpha=0, labelcolor=ACCENT)
    fig.tight_layout(pad=0.6)
    return fig


def _fig_anomaly_space(anomalies: pd.DataFrame) -> plt.Figure:
    """Scatter: Peak Z (y) × Breadth (x), coloured by anomaly type."""
    fig, ax = plt.subplots(figsize=(9, 3.4))
    for atype, color in ANOMALY_COLORS.items():
        sub = anomalies[anomalies["_anomaly_type"] == atype] if "_anomaly_type" in anomalies.columns else pd.DataFrame()
        if sub.empty:
            continue
        ax.scatter(
            sub["_anomaly_breadth"], sub["_peak_z"],
            c=color, s=14, alpha=0.72, linewidths=0, label=atype, zorder=3,
        )
    ax.set_xlabel("Anomaly breadth  |{j : z_ij ≥ α}|", fontsize=8)
    ax.set_ylabel("Peak z-score  max_j z_ij", fontsize=8)
    ax.grid(True, zorder=0)
    ax.legend(fontsize=7, framealpha=0, ncol=3, loc="upper left",
              labelcolor="linecolor", markerscale=1.8)
    fig.tight_layout(pad=0.6)
    return fig


def _fig_player_zscore(z_df: pd.DataFrame, threshold: float, pos_group: str) -> plt.Figure:
    """Horizontal bar chart of per-metric z-scores."""
    n = len(z_df)
    fig, ax = plt.subplots(figsize=(9, max(3.0, n * 0.38)))
    colors = [
        DANGER    if v >= threshold      else
        "#ff8c00" if v >= threshold * 0.7 else
        TEAL      if v >= 0              else
        "#2a3560"
        for v in z_df["Z"].values
    ]
    bars = ax.barh(z_df["Metric"], z_df["Z"], color=colors, height=0.65, zorder=3)
    ax.axvline(0,         color=BORDER,  linewidth=1.0, zorder=2)
    ax.axvline(threshold, color=DANGER,  linewidth=1.0, linestyle="--", zorder=4, alpha=0.7)
    ax.axvline(-threshold, color=MUTED, linewidth=0.8, linestyle=":", zorder=4, alpha=0.4)
    for bar, val in zip(bars, z_df["Z"].values):
        ax.text(val + (0.04 if val >= 0 else -0.04),
                bar.get_y() + bar.get_height() / 2,
                f"{val:+.3f}", va="center",
                ha="left" if val >= 0 else "right",
                fontsize=7, color="#6a7a9a", fontfamily="monospace")
    ax.set_xlabel(f"z-score  (z = (x − μ) / σ)    threshold α = {threshold:.1f}", fontsize=8)
    ax.set_title(f"Position group: {pos_group}", fontsize=8.5, loc="left", pad=6)
    ax.invert_yaxis()
    ax.grid(axis="x", zorder=0)
    fig.tight_layout(pad=0.7)
    return fig


# ── Data / anomaly loaders ────────────────────────────────────────────────────

@st.cache_data(show_spinner=False, ttl=600)
def _load_raw(leagues_key: tuple[str, ...] | None, min_minutes: int) -> pd.DataFrame:
    df = load_wyscout_raw(min_minutes=min_minutes, leagues=list(leagues_key) if leagues_key else None)
    if df.empty:
        return df
    if "AgeYears" not in df.columns and "Age" in df.columns:
        df["AgeYears"] = pd.to_numeric(df["Age"], errors="coerce")
    mins_col = next((c for c in ["Minutes played", "MinutesPlayed"] if c in df.columns), None)
    if mins_col and "CompositeRecruitmentScore" not in df.columns:
        mins = pd.to_numeric(df[mins_col], errors="coerce").fillna(0)
        df["CompositeRecruitmentScore"] = (
            (mins - mins.min()) / (mins.max() - mins.min() + 1e-9) * 100
        ).clip(0, 100)
    return df.reset_index(drop=True)


@st.cache_data(show_spinner=False, ttl=600)
def _run_anomaly_detection(data_json: str, threshold: float, method: str) -> pd.DataFrame:
    df = pd.read_json(data_json)
    frames: list[pd.DataFrame] = []
    for pos_group, grp in df.groupby("PositionGroup"):
        metrics = [m for m in POSITION_METRICS.get(str(pos_group), []) if m in grp.columns]
        if not metrics or len(grp) < 5:
            frames.append(grp)
            continue
        try:
            frames.append(AnomalyEngine(threshold=threshold, method=method, groupby=None).fit_transform(grp, metrics))
        except Exception:
            frames.append(grp)
    return pd.concat(frames, ignore_index=True) if frames else df


# ── Excel export ──────────────────────────────────────────────────────────────

def _build_anomaly_excel(anomalies: pd.DataFrame, zdf: pd.DataFrame, threshold: float) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb  = Workbook()
    hdr = PatternFill("solid", fgColor="090C18")
    acc = PatternFill("solid", fgColor="3D7FFF")
    wf  = Font(color="FFFFFF", bold=True, size=10)
    tb  = Border(
        left=Side(style="thin", color="151C30"), right=Side(style="thin", color="151C30"),
        top=Side(style="thin", color="151C30"),  bottom=Side(style="thin", color="151C30"),
    )

    DISPLAY_COLS = [
        "Player", "Team", "Position", "PositionGroup", "Age", "_League",
        "Minutes played", "Matches played",
        "_anomaly_type", "_anomaly_score", "_peak_z", "_mean_z", "_anomaly_breadth",
        "Goals per 90", "xG per 90", "Assists per 90", "xA per 90",
        "Key passes per 90", "Passes per 90", "Accurate passes, %",
        "Successful defensive actions per 90", "Progressive passes per 90",
    ]
    RENAME = {
        "_League": "League", "_anomaly_type": "Anomaly Type",
        "_anomaly_score": "Ω Score", "_peak_z": "z_peak",
        "_mean_z": "z_mean", "_anomaly_breadth": "Breadth k",
    }

    def _ws(title: str, df: pd.DataFrame) -> None:
        if df is None or df.empty:
            return
        cols = [c for c in DISPLAY_COLS if c in df.columns]
        out  = df[cols].rename(columns=RENAME).round(4).reset_index(drop=True)
        ws = wb.create_sheet(title=title[:31])
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=11)
        ws.cell(row=1, column=1).fill = hdr
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out.columns))
        ws.row_dimensions[1].height = 20
        for ci, cn in enumerate(out.columns, 1):
            c = ws.cell(row=2, column=ci, value=cn)
            c.font = wf; c.fill = acc; c.alignment = Alignment(horizontal="center"); c.border = tb
        for ri, row in enumerate(dataframe_to_rows(out, index=False, header=False), 3):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = tb; c.alignment = Alignment(horizontal="left" if ci <= 3 else "center")
                if ri % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="0D1020")
        for col_cells in ws.columns:
            w = max(len(str(col_cells[1].value or "")), *(len(str(c.value or "")) for c in col_cells[2:10]))
            ws.column_dimensions[col_cells[0].column_letter].width = min(w + 3, 42)

    def _filt(t: str) -> pd.DataFrame:
        if "_anomaly_type" not in anomalies.columns:
            return pd.DataFrame()
        return anomalies[anomalies["_anomaly_type"] == t]

    ws0 = wb.active
    ws0.title = "Overview"
    ws0.cell(row=1, column=1, value=f"Anomaly Platform  ·  threshold z ≥ {threshold:.2f}  ·  n = {len(anomalies):,}  ·  p(z≥α) ≈ {(1-scipy_norm.cdf(threshold))*100:.2f}%")
    ws0.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=12)
    ws0.cell(row=1, column=1).fill = hdr
    ws0.merge_cells("A1:H1")
    ws0.row_dimensions[1].height = 24
    for i, (atype, color) in enumerate(ANOMALY_COLORS.items(), start=3):
        n = len(_filt(atype))
        ws0.cell(row=i, column=1, value=atype).fill = PatternFill("solid", fgColor=color.lstrip("#"))
        ws0.cell(row=i, column=1).font = Font(color="000000", bold=True)
        ws0.cell(row=i, column=2, value=n)

    _ws("All Anomalies",         anomalies.head(3000))
    _ws("Hidden Gems",           _filt("Hidden Gem"))
    _ws("Specialist Elite",      _filt("Specialist Elite"))
    _ws("Multi-dimensional",     _filt("Multi-dimensional"))
    _ws("Age-adjusted Gems",     _filt("Age-adjusted Gem"))
    _ws("Consistent Overperf",   _filt("Consistent Overperformer"))
    try:
        sp = SetPieceAnalyzer(threshold=threshold * 0.85)
        _ws("Set Piece Anomalies", sp.anomaly_table(sp.fit_transform(zdf), top_n=300))
    except Exception:
        pass

    buf = BytesIO(); wb.save(buf); return buf.getvalue()


# ── Shared helpers ────────────────────────────────────────────────────────────

def _pc(df: pd.DataFrame) -> str:
    return "Player" if "Player" in df.columns else "PlayerName"

def _tc(df: pd.DataFrame) -> str:
    return "Team" if "Team" in df.columns else "TeamName"


def _percentile_col(z_series: pd.Series) -> pd.Series:
    return (scipy_norm.cdf(z_series) * 100).round(1)


def _anomaly_table(df: pd.DataFrame, height: int = 520) -> None:
    disp = [c for c in [
        _pc(df), _tc(df), "PositionGroup", "_League", "Age",
        "_anomaly_type", "_anomaly_score", "_peak_z", "_mean_z", "_anomaly_breadth",
        "Goals per 90", "xG per 90", "Assists per 90", "xA per 90",
        "Key passes per 90", "Successful defensive actions per 90",
        "Progressive passes per 90",
    ] if c in df.columns]
    out = df[disp].copy()
    if "_peak_z" in out.columns:
        out.insert(out.columns.get_loc("_peak_z") + 1, "Pct", _percentile_col(out["_peak_z"]))
    rename = {
        _pc(df): "Player", _tc(df): "Team",
        "_League": "League", "_anomaly_type": "Type",
        "_anomaly_score": "Ω", "_peak_z": "z_peak",
        "_mean_z": "z_mean", "_anomaly_breadth": "k",
    }
    st.dataframe(
        out.rename(columns=rename).round(3),
        use_container_width=True, height=height, hide_index=True,
        column_config={
            "Ω":      st.column_config.ProgressColumn("Ω",      min_value=0, max_value=20, format="%.3f"),
            "z_peak": st.column_config.NumberColumn("z_peak",   format="%.3f"),
            "z_mean": st.column_config.NumberColumn("z_mean",   format="%.3f"),
            "Pct":    st.column_config.ProgressColumn("Pct %",  min_value=0, max_value=100, format="%.1f"),
        },
    )


def _page_header(eyebrow: str, title: str, sub: str) -> None:
    st.markdown(
        f"<div class='page-header'>"
        f"<div class='page-header-eyebrow'>{eyebrow}</div>"
        f"<div class='page-header-title'>{title}</div>"
        f"<div class='page-header-sub'>{sub}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )


# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
    <div class='sidebar-brand'>
        <div class='brand-title'>⬡ ANOMALY PLATFORM</div>
        <div class='brand-sub'>Statistical outlier detection</div>
    </div>
    """, unsafe_allow_html=True)

    page = st.radio(
        "Navigate",
        ["Σ  Anomaly Hub", "∂  Explorer", "z  Player Profile",
         "~  Similarity", "⊕  Set Pieces", "↓  Export"],
        label_visibility="collapsed",
    )

    st.markdown("<div class='section-label'>Dataset</div>", unsafe_allow_html=True)
    all_leagues = wyscout_available_leagues()
    sel_leagues = st.multiselect(
        "Leagues", all_leagues, default=[],
        placeholder="All leagues", key="sel_leagues", label_visibility="collapsed",
    )
    leagues_key: tuple[str, ...] | None = tuple(sorted(sel_leagues)) if sel_leagues else None
    min_minutes = st.slider("Min minutes", 0, 2500, 450, 50, key="min_min")

    st.markdown("<div class='section-label'>Detection parameters</div>", unsafe_allow_html=True)
    threshold = st.slider("Threshold α", 1.0, 3.5, 1.8, 0.05, key="threshold",
                          help="z-score cutoff. Players with any metric z ≥ α are flagged.")
    method_opts = {"Z-score  z = (x−μ)/σ": "z-score", "MAD  0.6745(x−m̃)/MAD": "mad", "IQR  (x−Q₃)/IQR": "iqr"}
    method_label = st.selectbox("Method", list(method_opts.keys()), key="method", label_visibility="collapsed")
    method = method_opts[method_label]

    # Detection model box
    p_val = (1 - scipy_norm.cdf(threshold)) * 100
    st.markdown(
        f"<div class='formula-box'>"
        f"<span class='highlight'>z</span><sub>ij</sub> = (x<sub>ij</sub> − μ<sub>j,pos</sub>) / σ<sub>j,pos</sub><br>"
        f"flag ← <span class='danger'>z</span><sub>ij</sub> ≥ <span class='danger'>{threshold:.2f}</span><br>"
        f"p(z ≥ α | 𝒩) ≈ <span class='highlight'>{p_val:.2f}%</span><br>"
        f"groupby: position class"
        f"</div>",
        unsafe_allow_html=True,
    )


# ── Load & detect ─────────────────────────────────────────────────────────────

with st.spinner("Loading Wyscout data…"):
    df = _load_raw(leagues_key, min_minutes)

if df.empty:
    st.error("No data. Add Wyscout `.xlsx` files to `data/Wyscout DB/`.")
    st.stop()

with st.spinner("Running anomaly detection…"):
    zdf = _run_anomaly_detection(df.to_json(), threshold, method)

anomalies = (
    zdf[zdf["_peak_z"] >= threshold]
    .sort_values("_anomaly_score", ascending=False)
    .reset_index(drop=True)
) if "_peak_z" in zdf.columns else pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# Σ  Anomaly Hub
# ══════════════════════════════════════════════════════════════════════════════
if page == "Σ  Anomaly Hub":
    _page_header(
        "STATISTICAL OUTLIER INTELLIGENCE",
        "Anomaly Hub",
        f"n = {len(df):,} players  ·  z ≥ {threshold:.2f}  ·  p(z≥α) ≈ {(1-scipy_norm.cdf(threshold))*100:.2f}%  ·  method: {method}",
    )

    n_anom = len(anomalies)
    rate   = n_anom / len(df) * 100 if len(df) else 0

    c1, c2, c3, c4 = st.columns(4)
    kpis = [
        (c1, f"{len(df):,}",      "Players  n",           "n ∈ ℕ",   ""),
        (c2, f"{n_anom:,}",       "Anomalies  |Â|",        f"{rate:.1f}% of n", "danger"),
        (c3, df["_League"].nunique() if "_League" in df.columns else 0, "Leagues",  "", "teal"),
        (c4, df["PositionGroup"].nunique() if "PositionGroup" in df.columns else 0, "Position groups", "k classes", "purple"),
    ]
    for w, num, lbl, sub, cls in kpis:
        w.markdown(
            f"<div class='kpi-card {cls}'>"
            f"<div class='kpi-num'>{num}</div>"
            f"<div class='kpi-label'>{lbl}</div>"
            f"<div class='kpi-sub'>{sub}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )

    if anomalies.empty:
        st.info("No anomalies at this threshold. Reduce α in the sidebar.")
        st.stop()

    # ── Type badges ───────────────────────────────────────────────────────────
    type_counts = anomalies["_anomaly_type"].value_counts() if "_anomaly_type" in anomalies.columns else pd.Series()
    badges_html = "<div class='type-row'>"
    for atype, color in ANOMALY_COLORS.items():
        cnt = int(type_counts.get(atype, 0))
        pct = cnt / n_anom * 100 if n_anom else 0
        badges_html += (
            f"<div class='type-badge' style='border-left-color:{color};'>"
            f"<div class='type-count' style='color:{color};'>{cnt}</div>"
            f"<div class='type-name'>{atype}</div>"
            f"<div class='type-pct'>{pct:.1f}% of anomalies</div>"
            f"</div>"
        )
    badges_html += "</div>"
    st.markdown(badges_html, unsafe_allow_html=True)

    # ── Z-score distribution ──────────────────────────────────────────────────
    if "_peak_z" in zdf.columns:
        st.markdown("<div class='section-label'>Peak z-score distribution  P(z_peak)</div>", unsafe_allow_html=True)
        peak_z_all = pd.to_numeric(zdf["_peak_z"], errors="coerce").dropna()
        if not peak_z_all.empty:
            fig_dist = _fig_z_distribution(peak_z_all, threshold)
            st.pyplot(fig_dist, clear_figure=True, use_container_width=True)

    # ── Anomaly space scatter ─────────────────────────────────────────────────
    if "_peak_z" in anomalies.columns and "_anomaly_breadth" in anomalies.columns:
        st.markdown("<div class='section-label'>Anomaly space  (breadth k × peak z)</div>", unsafe_allow_html=True)
        fig_space = _fig_anomaly_space(anomalies)
        st.pyplot(fig_space, clear_figure=True, use_container_width=True)

    # ── Stacked breakdown + top table ─────────────────────────────────────────
    st.markdown("<div class='section-label'>Breakdown by position class</div>", unsafe_allow_html=True)
    col_left, col_right = st.columns([3, 2])

    with col_left:
        if "_anomaly_type" in anomalies.columns and "PositionGroup" in anomalies.columns:
            bd = anomalies.groupby(["PositionGroup", "_anomaly_type"]).size().reset_index(name="n")
            pos_ord = [p for p in POSITION_ORDER[::-1] if p in bd["PositionGroup"].values]
            ch = (
                alt.Chart(bd)
                .mark_bar(cornerRadiusTopRight=3, cornerRadiusBottomRight=3)
                .encode(
                    y=alt.Y("PositionGroup:N", sort=pos_ord, title=None,
                             axis=alt.Axis(labelColor="#6a7a9a", labelFontSize=11, labelFont="monospace")),
                    x=alt.X("n:Q", title="n anomalies",
                             axis=alt.Axis(labelColor="#6a7a9a", gridColor=BORDER, titleFont="monospace")),
                    color=alt.Color(
                        "_anomaly_type:N",
                        scale=alt.Scale(domain=list(ANOMALY_COLORS.keys()), range=list(ANOMALY_COLORS.values())),
                        title=None,
                        legend=alt.Legend(labelColor="#8898b8", orient="top", labelFont="monospace", labelFontSize=10),
                    ),
                    tooltip=["PositionGroup", "_anomaly_type", "n"],
                )
                .properties(height=300)
                .configure_view(fill=SURFACE, stroke=None)
                .configure(background=SURFACE)
            )
            st.altair_chart(ch, use_container_width=True)

    with col_right:
        st.markdown("<div class='section-label'>Top 20 by Ω score</div>", unsafe_allow_html=True)
        top20 = anomalies.head(20)[[c for c in [_pc(anomalies), "PositionGroup", "_anomaly_type", "_anomaly_score", "_peak_z", "_anomaly_breadth"] if c in anomalies.columns]].copy()
        top20.insert(0, "#", range(1, len(top20) + 1))
        st.dataframe(
            top20.rename(columns={_pc(anomalies): "Player", "_anomaly_type": "Type", "_anomaly_score": "Ω", "_peak_z": "z_peak", "_anomaly_breadth": "k"}).round(3),
            use_container_width=True, height=346, hide_index=True,
            column_config={"Ω": st.column_config.ProgressColumn("Ω", min_value=0, max_value=20, format="%.3f")},
        )

    # ── League density table ──────────────────────────────────────────────────
    if "_League" in anomalies.columns and "_League" in df.columns:
        st.markdown("<div class='section-label'>League anomaly density</div>", unsafe_allow_html=True)
        ld = (
            anomalies["_League"].value_counts().rename("A")
            .to_frame().join(df["_League"].value_counts().rename("N"))
            .dropna().reset_index()
        )
        ld.columns = ["League", "A", "N"]
        ld["ρ (%)"] = (ld["A"] / ld["N"] * 100).round(2)
        ld["μ_Ω"] = ld["League"].map(anomalies.groupby("_League")["_anomaly_score"].mean().round(3))
        ld = ld.sort_values("ρ (%)", ascending=False).head(30)

        l1, l2 = st.columns([3, 2])
        with l1:
            lc = (
                alt.Chart(ld)
                .mark_bar(cornerRadiusTopRight=3, cornerRadiusBottomRight=3, color=DANGER)
                .encode(
                    y=alt.Y("League:N", sort="-x", title=None,
                             axis=alt.Axis(labelColor="#6a7a9a", labelFont="monospace", labelFontSize=10)),
                    x=alt.X("ρ (%):Q", title="Anomaly density ρ = |Â_l| / N_l  (%)",
                             axis=alt.Axis(labelColor="#6a7a9a", gridColor=BORDER, titleFont="monospace")),
                    tooltip=["League", "A", "N", "ρ (%)", "μ_Ω"],
                )
                .properties(height=max(280, len(ld) * 19))
                .configure_view(fill=SURFACE, stroke=None)
                .configure(background=SURFACE)
            )
            st.altair_chart(lc, use_container_width=True)
        with l2:
            st.dataframe(
                ld.reset_index(drop=True),
                use_container_width=True, height=380, hide_index=True,
                column_config={"ρ (%)": st.column_config.ProgressColumn("ρ (%)", min_value=0, max_value=100, format="%.2f")},
            )


# ══════════════════════════════════════════════════════════════════════════════
# ∂  Explorer
# ══════════════════════════════════════════════════════════════════════════════
elif page == "∂  Explorer":
    _page_header(
        "ANOMALY EXPLORER",
        "Filtered Anomaly Database",
        f"Full z-scored dataset  ·  α = {threshold:.2f}  ·  n_total = {len(anomalies):,}",
    )

    if anomalies.empty:
        st.info("No anomalies at this threshold.")
        st.stop()

    f1, f2, f3, f4, f5 = st.columns(5)
    with f1:
        sel_pos = st.multiselect("Position group", POSITION_ORDER, default=[], key="exp_pos")
    with f2:
        all_types = sorted(anomalies["_anomaly_type"].unique()) if "_anomaly_type" in anomalies.columns else []
        sel_types = st.multiselect("Anomaly type", all_types, default=[], key="exp_type")
    with f3:
        leagues_avail = sorted(anomalies["_League"].unique()) if "_League" in anomalies.columns else []
        sel_exp_leagues = st.multiselect("League", leagues_avail, default=[], key="exp_leagues")
    with f4:
        age_range = st.slider("Age range", 15, 45, (16, 36), key="exp_age")
    with f5:
        min_z = st.slider("Min z_peak", float(threshold), 6.0, float(threshold), 0.1, key="exp_minz")

    filt = anomalies.copy()
    if sel_pos and "PositionGroup" in filt.columns:
        filt = filt[filt["PositionGroup"].isin(sel_pos)]
    if sel_types and "_anomaly_type" in filt.columns:
        filt = filt[filt["_anomaly_type"].isin(sel_types)]
    if sel_exp_leagues and "_League" in filt.columns:
        filt = filt[filt["_League"].isin(sel_exp_leagues)]
    age_col_f = "Age" if "Age" in filt.columns else ("AgeYears" if "AgeYears" in filt.columns else None)
    if age_col_f:
        filt = filt[pd.to_numeric(filt[age_col_f], errors="coerce").between(age_range[0], age_range[1])]
    if "_peak_z" in filt.columns:
        filt = filt[filt["_peak_z"] >= min_z]
    filt = filt.reset_index(drop=True)

    # Summary stats row
    if not filt.empty and "_peak_z" in filt.columns:
        pz = filt["_peak_z"]
        s1, s2, s3, s4, s5 = st.columns(5)
        for w, val, lbl in [
            (s1, f"{len(filt):,}", "n anomalies"),
            (s2, f"{pz.mean():.3f}", "μ(z_peak)"),
            (s3, f"{pz.std():.3f}", "σ(z_peak)"),
            (s4, f"{pz.max():.3f}", "max(z_peak)"),
            (s5, f"{filt['_anomaly_breadth'].mean():.2f}" if "_anomaly_breadth" in filt.columns else "—", "μ(breadth k)"),
        ]:
            w.markdown(
                f"<div style='background:{SURF2};border:1px solid {BORDER};border-radius:6px;"
                f"padding:8px 14px;font-family:{MONO};'>"
                f"<div style='font-size:1.1rem;font-weight:700;color:#e8eeff;'>{val}</div>"
                f"<div style='font-size:0.58rem;color:{MUTED};text-transform:uppercase;letter-spacing:.09em;margin-top:2px;'>{lbl}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )
        st.markdown("")

    pos_present = [p for p in POSITION_ORDER if p in filt.get("PositionGroup", pd.Series()).values]
    tabs = st.tabs(["ALL  n=" + str(len(filt))] + [f"{p}  n={int((filt['PositionGroup']==p).sum())}" for p in pos_present])

    for tab_w, label in zip(tabs, ["All"] + pos_present):
        with tab_w:
            sub = filt if label == "All" else filt[filt["PositionGroup"] == label]
            if sub.empty:
                st.info("No results.")
            else:
                _anomaly_table(sub, height=560)

    st.markdown("---")
    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button("↓ CSV", data=filt.to_csv(index=False).encode(), file_name="anomalies.csv", mime="text/csv")
    with dl2:
        st.download_button("↓ Excel", data=_build_anomaly_excel(filt, zdf, threshold),
                           file_name="anomaly_report.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════════════════════
# z  Player Profile
# ══════════════════════════════════════════════════════════════════════════════
elif page == "z  Player Profile":
    _page_header(
        "INDIVIDUAL ANOMALY PROFILE",
        "Player z-Score Analysis",
        "Per-metric z-scores relative to position group population  ·  zᵢⱼ = (xᵢⱼ − μⱼ) / σⱼ",
    )

    if anomalies.empty:
        st.info("No anomalies detected at this threshold.")
        st.stop()

    pc = _pc(anomalies)
    tc = _tc(anomalies)

    fc1, fc2 = st.columns([4, 1])
    with fc2:
        pos_filt = st.selectbox("Position", ["All"] + POSITION_ORDER, key="prof_pos")
    with fc1:
        pool = anomalies if pos_filt == "All" else anomalies[anomalies.get("PositionGroup", pd.Series()) == pos_filt]
        opts = (pool[pc].astype(str) + "  |  " + pool.get(tc, pd.Series("", index=pool.index)).astype(str)
                + "  |  " + pool.get("PositionGroup", pd.Series("", index=pool.index)).astype(str)).tolist()
        sel  = st.selectbox("Select anomaly", opts if opts else ["—"], key="prof_sel")

    if not opts:
        st.info("No anomalies for this position.")
        st.stop()

    sel_name  = sel.split("  |  ")[0].strip()
    row       = pool.loc[pool[pc] == sel_name].iloc[0]
    pos_group = str(row.get("PositionGroup", ""))
    atype     = str(row.get("_anomaly_type", ""))
    acolor    = ANOMALY_COLORS.get(atype, "#888")
    team      = row.get(tc, "")
    league    = row.get("_League", "")
    age       = row.get("Age", row.get("AgeYears", ""))
    peak_z    = float(row.get("_peak_z", 0))
    score     = float(row.get("_anomaly_score", 0))
    breadth   = int(row.get("_anomaly_breadth", 0))
    mean_z    = float(row.get("_mean_z", 0))
    mins      = row.get("Minutes played", row.get("MinutesPlayed", "—"))
    pct       = scipy_norm.cdf(peak_z) * 100

    # ── Player card ───────────────────────────────────────────────────────────
    st.markdown(
        f"<div class='player-card' style='--accent:{acolor};'>"
        f"<div style='position:absolute;top:0;left:0;width:4px;height:100%;background:{acolor};border-radius:6px 0 0 6px;'></div>"
        f"<div class='player-name'>{sel_name}</div>"
        f"<div class='player-meta'>{team}  ·  {league}  ·  {pos_group}  ·  age {age}  ·  {mins} min</div>"
        f"<div style='display:flex;flex-wrap:wrap;gap:0;margin-top:14px;'>"
        f"<span class='stat-pill'><span class='label'>TYPE</span><span class='value' style='color:{acolor};'>{atype}</span></span>"
        f"<span class='stat-pill'><span class='label'>Ω SCORE</span><span class='value'>{score:.4f}</span></span>"
        f"<span class='stat-pill'><span class='label'>z_peak</span><span class='value' style='color:{DANGER};'>{peak_z:.4f}</span></span>"
        f"<span class='stat-pill'><span class='label'>z_mean</span><span class='value'>{mean_z:.4f}</span></span>"
        f"<span class='stat-pill'><span class='label'>BREADTH k</span><span class='value'>{breadth}</span></span>"
        f"<span class='stat-pill'><span class='label'>Φ(z_peak)</span><span class='value' style='color:{TEAL};'>{pct:.2f}th pct</span></span>"
        f"</div></div>",
        unsafe_allow_html=True,
    )

    # ── Metric breakdown ──────────────────────────────────────────────────────
    z_metrics = POSITION_METRICS.get(pos_group, [])

    # Compute μ and σ for the position group
    pos_pool = zdf[zdf.get("PositionGroup", pd.Series()) == pos_group] if "PositionGroup" in zdf.columns else zdf

    stat_rows = []
    for m in z_metrics:
        z_col = f"_z_{m}"
        if z_col not in row.index:
            continue
        z_val = float(row[z_col])
        raw   = row.get(m)
        mu_j  = pos_pool[m].mean() if m in pos_pool.columns else np.nan
        sig_j = pos_pool[m].std()  if m in pos_pool.columns else np.nan
        pct_j = scipy_norm.cdf(z_val) * 100
        stat_rows.append({
            "Metric": m,
            "xᵢ": round(float(raw), 4) if pd.notna(raw) else np.nan,
            "μⱼ": round(mu_j, 4) if pd.notna(mu_j) else np.nan,
            "σⱼ": round(sig_j, 4) if pd.notna(sig_j) else np.nan,
            "zᵢⱼ": round(z_val, 4),
            "Φ(z)": round(pct_j, 1),
            "flag": z_val >= threshold,
        })

    ch1, ch2 = st.columns([3, 2])

    with ch1:
        if stat_rows:
            z_df = pd.DataFrame(stat_rows).sort_values("zᵢⱼ", ascending=False)
            fig  = _fig_player_zscore(z_df.rename(columns={"Metric": "Metric", "zᵢⱼ": "Z"}), threshold, pos_group)
            st.markdown("<div class='section-label'>Per-metric z-score  zᵢⱼ = (xᵢ − μⱼ) / σⱼ</div>", unsafe_allow_html=True)
            st.pyplot(fig, clear_figure=True, use_container_width=True)

    with ch2:
        if stat_rows:
            st.markdown("<div class='section-label'>Statistical table  (population parameters)</div>", unsafe_allow_html=True)
            z_df2 = pd.DataFrame(stat_rows).sort_values("zᵢⱼ", ascending=False)
            rows_html = ""
            for _, r in z_df2.iterrows():
                z = r["zᵢⱼ"]
                flag_cls = "z-high" if z >= threshold else ("z-med" if z >= threshold * 0.7 else ("z-pos" if z >= 0 else "z-neg"))
                star = " ★" if r["flag"] else ""
                rows_html += (
                    f"<tr>"
                    f"<td title='{r['Metric']}'>{r['Metric'][:30]}{star}</td>"
                    f"<td>{r['xᵢ']:.3f}</td>"
                    f"<td>{r['μⱼ']:.3f}</td>"
                    f"<td>{r['σⱼ']:.3f}</td>"
                    f"<td class='{flag_cls}'>{z:.4f}</td>"
                    f"<td style='color:{TEAL};'>{r['Φ(z)']:.1f}</td>"
                    f"</tr>"
                )
            st.markdown(
                f"<table class='stat-table'>"
                f"<thead><tr><th>Metric</th><th>xᵢ</th><th>μⱼ</th><th>σⱼ</th><th>zᵢⱼ</th><th>Φ(z)%</th></tr></thead>"
                f"<tbody>{rows_html}</tbody>"
                f"</table>",
                unsafe_allow_html=True,
            )
            n_pos_grp = len(pos_pool)
            st.markdown(
                f"<div style='font-family:{MONO};font-size:0.62rem;color:{MUTED};margin-top:8px;'>"
                f"★ z ≥ α = {threshold:.2f}  ·  n({pos_group}) = {n_pos_grp}  ·  μ,σ computed within position group"
                f"</div>",
                unsafe_allow_html=True,
            )

    # ── Position peers ────────────────────────────────────────────────────────
    st.markdown("<div class='section-label'>Position group peers  (top 10 by Ω)</div>", unsafe_allow_html=True)
    peers = anomalies[anomalies.get("PositionGroup", pd.Series()) == pos_group]
    peers = peers[peers[pc] != sel_name].head(10)
    if not peers.empty:
        _anomaly_table(peers, height=360)


# ══════════════════════════════════════════════════════════════════════════════
# ~  Similarity
# ══════════════════════════════════════════════════════════════════════════════
elif page == "~  Similarity":
    _page_header(
        "FEATURE-SPACE SIMILARITY",
        "Similarity Engine",
        "Standardised feature vectors  ·  z* = (x−μ)/σ  ·  cosine / Pearson / Euclidean distance",
    )

    if anomalies.empty:
        st.info("No anomalies available.")
        st.stop()

    pc = _pc(anomalies)
    tc = _tc(anomalies)
    opts = (
        anomalies[pc].astype(str) + "  |  "
        + anomalies.get(tc, pd.Series("", index=anomalies.index)).astype(str)
        + "  |  " + anomalies.get("PositionGroup", pd.Series("", index=anomalies.index)).astype(str)
    ).tolist()

    sc1, sc2, sc3, sc4 = st.columns([4, 1, 1, 1])
    with sc1:
        sel = st.selectbox("Target player", opts, key="sim_player")
    with sc2:
        n_k = st.slider("k", 5, 30, 10, key="sim_n")
    with sc3:
        same_pos = st.checkbox("Same position", value=True, key="sim_pos")
    with sc4:
        sim_method = st.selectbox("Metric", ["cosine", "pearson", "euclidean"], key="sim_method", label_visibility="collapsed")

    sel_name  = sel.split("  |  ")[0].strip()
    target    = anomalies.loc[anomalies[pc] == sel_name].iloc[0]
    pos_group = str(target.get("PositionGroup", ""))
    atype     = str(target.get("_anomaly_type", ""))
    acolor    = ANOMALY_COLORS.get(atype, "#888")

    # Formula box for chosen method
    formulas = {
        "cosine":    "sim(A,B) = A·B / (‖A‖₂ · ‖B‖₂)",
        "pearson":   "r(A,B) = Σ(Aᵢ−Ā)(Bᵢ−B̄) / √[Σ(Aᵢ−Ā)² · Σ(Bᵢ−B̄)²]",
        "euclidean": "d(A,B) = ‖A−B‖₂ = √Σ(Aᵢ−Bᵢ)²   →   sim = 1 − d/d_max",
    }
    st.markdown(
        f"<div style='display:flex;gap:14px;align-items:center;margin-bottom:16px;flex-wrap:wrap;'>"
        f"<div style='background:{SURF2};border:1px solid {BORDER};border-left:3px solid {acolor};"
        f"border-radius:0 6px 6px 0;padding:10px 16px;flex:1;'>"
        f"<span style='font-weight:700;color:#e8eeff;font-size:.9rem;'>{sel_name}</span>&nbsp;&nbsp;"
        f"<span style='background:{acolor};color:#000;padding:1px 9px;border-radius:3px;font-size:.65rem;font-weight:800;'>{atype}</span>&nbsp;&nbsp;"
        f"<span style='color:{MUTED};font-family:{MONO};font-size:.72rem;'>{target.get(tc,'')} · {pos_group}</span>"
        f"</div>"
        f"<div class='formula-box' style='flex:1.2;margin:0;'>"
        f"<span class='highlight'>{sim_method}:</span> {formulas[sim_method]}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    feats   = [f for f in SIMILARITY_FEATURES if f in zdf.columns]
    engine  = SimilarityEngine(feats)
    results = engine.find_similar(zdf, target, method=sim_method, n=n_k, same_position=same_pos)

    if results.empty:
        st.info("No similar players found.")
    else:
        pc_r = _pc(results); tc_r = _tc(results)
        disp = [c for c in [pc_r, tc_r, "PositionGroup", "_League", "Age", "_similarity"] + feats[:8] if c in results.columns]
        out  = results[disp].copy()
        out.insert(0, "rank", range(1, len(out) + 1))
        st.dataframe(
            out.rename(columns={pc_r: "Player", tc_r: "Team", "_League": "League", "_similarity": f"sim ({sim_method})"}).round(4),
            use_container_width=True, height=420, hide_index=True,
            column_config={f"sim ({sim_method})": st.column_config.ProgressColumn(f"sim ({sim_method})", min_value=-1, max_value=1, format="%.4f")},
        )
        st.markdown(
            f"<div style='font-family:{MONO};font-size:.62rem;color:{MUTED};margin-top:4px;'>"
            f"k = {len(results)}  ·  feature space dim = {len(feats)}  ·  {'same position class' if same_pos else 'all positions'}"
            f"</div>",
            unsafe_allow_html=True,
        )

    with st.expander("All three distance metrics side by side"):
        mt = st.tabs(["Cosine", "Pearson", "Euclidean"])
        for tab_w, mn in zip(mt, ["cosine", "pearson", "euclidean"]):
            with tab_w:
                st.markdown(
                    f"<div class='formula-box'>{formulas[mn]}</div>",
                    unsafe_allow_html=True,
                )
                rm = engine.find_similar(zdf, target, method=mn, n=n_k, same_position=same_pos)
                if rm.empty:
                    st.info("No results.")
                else:
                    pc_m = _pc(rm)
                    dm   = [c for c in [pc_m, "PositionGroup", "_League", "Age", "_similarity"] if c in rm.columns]
                    rm.insert(0, "rank", range(1, len(rm) + 1))
                    st.dataframe(
                        rm[["rank"] + dm].rename(columns={pc_m: "Player", "_League": "League", "_similarity": "sim"}).round(4),
                        use_container_width=True, height=300, hide_index=True,
                    )


# ══════════════════════════════════════════════════════════════════════════════
# ⊕  Set Pieces
# ══════════════════════════════════════════════════════════════════════════════
elif page == "⊕  Set Pieces":
    _page_header(
        "SET-PIECE ANOMALY DETECTION",
        "Set Piece Scanner",
        f"Role-weighted z-score profiling  ·  threshold α × 0.85 = {threshold * 0.85:.3f}  ·  6 role archetypes",
    )

    sp_analyzer  = SetPieceAnalyzer(threshold=threshold * 0.85)
    ws_enriched  = sp_analyzer.fit_transform(df)
    sp_anom_df   = sp_analyzer.anomaly_table(ws_enriched, top_n=300)
    role_leaders = sp_analyzer.top_players_by_role(ws_enriched, top_n=25)

    n_sp = len(sp_anom_df)
    st.markdown(
        f"<div class='kpi-card danger' style='max-width:240px;margin-bottom:20px;'>"
        f"<div class='kpi-num'>{n_sp}</div>"
        f"<div class='kpi-label'>Set-piece anomalies</div>"
        f"<div class='kpi-sub'>α_SP = {threshold * 0.85:.3f}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # Role formula
    role_formulas = {k: "  +  ".join(f"{w}×z({m[:20]})" for m, w in v.items())
                     for k, v in SET_PIECE_ROLES.items()}
    with st.expander("Role composite score formulas"):
        for role, formula in role_formulas.items():
            st.markdown(
                f"<div class='formula-box'>"
                f"<span class='highlight'>{role}</span>: S = ({formula}) / Σwᵢ"
                f"</div>",
                unsafe_allow_html=True,
            )

    if not sp_anom_df.empty:
        st.markdown("<div class='section-label'>Set-piece anomaly table  (sorted by z_peak)</div>", unsafe_allow_html=True)
        st.dataframe(sp_anom_df.round(4), use_container_width=True, height=380, hide_index=True)

    st.markdown("<div class='section-label'>Top players by role archetype</div>", unsafe_allow_html=True)
    role_names = list(SET_PIECE_ROLES.keys())
    role_tabs  = st.tabs(role_names)
    for rtab, role_name in zip(role_tabs, role_names):
        with rtab:
            rdf = role_leaders.get(role_name, pd.DataFrame())
            if rdf.empty:
                st.info(f"No data for {role_name}.")
            else:
                st.markdown(
                    f"<div class='formula-box'><span class='highlight'>{role_name}</span>:  "
                    f"S = {role_formulas.get(role_name, '')}</div>",
                    unsafe_allow_html=True,
                )
                st.dataframe(rdf.round(4), use_container_width=True, height=460, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# ↓  Export
# ══════════════════════════════════════════════════════════════════════════════
elif page == "↓  Export":
    _page_header(
        "DATA EXPORT",
        "Anomaly Report",
        f"Multi-sheet Excel  ·  flat CSV  ·  n = {len(anomalies):,} anomalies  ·  z ≥ {threshold:.2f}",
    )

    n_total = len(anomalies)
    p_val   = (1 - scipy_norm.cdf(threshold)) * 100

    st.markdown(
        f"<div class='formula-box' style='max-width:600px;'>"
        f"Detection:  z = (x − μ_pos) / σ_pos    method = {method}<br>"
        f"Threshold:  α = {threshold:.4f}    ⟹    p(z ≥ α | 𝒩) ≈ {p_val:.4f}%<br>"
        f"Sample:     n = {len(df):,}    anomalies = {n_total:,}    rate = {n_total/len(df)*100:.3f}%"
        f"</div>",
        unsafe_allow_html=True,
    )

    if not anomalies.empty and "_anomaly_type" in anomalies.columns:
        tc_s = anomalies["_anomaly_type"].value_counts()
        badges = "<div class='type-row'>"
        for atype, color in ANOMALY_COLORS.items():
            cnt = int(tc_s.get(atype, 0))
            pct = cnt / n_total * 100 if n_total else 0
            badges += (
                f"<div class='type-badge' style='border-left-color:{color};'>"
                f"<div class='type-count' style='color:{color};'>{cnt}</div>"
                f"<div class='type-name'>{atype}</div>"
                f"<div class='type-pct'>{pct:.1f}%</div>"
                f"</div>"
            )
        badges += "</div>"
        st.markdown(badges, unsafe_allow_html=True)

    ec1, ec2 = st.columns(2)
    with ec1:
        st.markdown("<div class='section-label'>Excel workbook — multi-sheet</div>", unsafe_allow_html=True)
        st.markdown(
            f"<div style='font-family:{MONO};font-size:.72rem;color:{MUTED};line-height:2;'>"
            "Overview  ·  All Anomalies  ·  Hidden Gems<br>"
            "Specialist Elite  ·  Multi-dimensional<br>"
            "Age-adjusted Gems  ·  Consistent Overperf<br>"
            "Set Piece Anomalies"
            "</div>",
            unsafe_allow_html=True,
        )
        if n_total > 0:
            st.download_button(
                "↓ Download .xlsx",
                data=_build_anomaly_excel(anomalies, zdf, threshold),
                file_name=f"anomaly_report_z{threshold:.2f}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    with ec2:
        st.markdown("<div class='section-label'>CSV — flat table</div>", unsafe_allow_html=True)
        st.markdown(
            f"<div style='font-family:{MONO};font-size:.72rem;color:{MUTED};line-height:2;'>"
            f"All {n_total:,} anomalies  ·  all z-score columns<br>"
            "Columns: Player, Team, Position, League, Age,<br>"
            "Anomaly type, Ω score, z_peak, z_mean, k,<br>"
            "all per-90 metrics"
            "</div>",
            unsafe_allow_html=True,
        )
        if n_total > 0:
            st.download_button(
                "↓ Download .csv",
                data=anomalies.to_csv(index=False).encode(),
                file_name=f"anomaly_platform_z{threshold:.2f}.csv",
                mime="text/csv",
                use_container_width=True,
            )
